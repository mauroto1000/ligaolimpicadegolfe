from flask import Flask, render_template, request, redirect, url_for, flash, jsonify, session
import sqlite3
from datetime import datetime, timedelta
import os
from functools import wraps
import hashlib
from werkzeug.utils import secure_filename  # Adicione esta linha
import json
from flask import send_file
from io import BytesIO

# Adicionando session config
app = Flask(__name__)
app.secret_key = 'liga_olimpica_golfe_2025'
app.config['PERMANENT_SESSION_LIFETIME'] = timedelta(hours=24)  # Sessão válida por 24 horas

DATABASE = 'golf_league.db'

CHAT_STATES = {}
CHAT_STATE_TIMEOUT = 10


# Função para obter conexão com o banco de dados
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# Função decoradora para verificar autenticação
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'error')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# Configurações para upload de arquivos
UPLOAD_FOLDER = 'static/profile_photos'
ALLOWED_EXTENSIONS = {'png', 'jpg', 'jpeg', 'gif'}

# Certifique-se de que a pasta existe
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

# Adicionar configuração à aplicação
app.config['UPLOAD_FOLDER'] = UPLOAD_FOLDER
app.config['MAX_CONTENT_LENGTH'] = 5 * 1024 * 1024  # Limitar tamanho para 5MB

def allowed_file(filename):
    return '.' in filename and \
           filename.rsplit('.', 1)[1].lower() in ALLOWED_EXTENSIONS

@app.route('/upload_profile_photo/<int:player_id>', methods=['POST'])
@login_required
def upload_profile_photo(player_id):
    try:
        # Verificar se é o próprio jogador ou um admin
        if not (session.get('user_id') == player_id or session.get('is_admin', False)):
            flash('Acesso negado. Você só pode alterar sua própria foto de perfil.', 'error')
            return redirect(url_for('player_detail', player_id=player_id))
        
        # Verificar se o arquivo foi enviado
        if 'profile_photo' not in request.files:
            flash('Nenhum arquivo enviado.', 'error')
            return redirect(url_for('player_detail', player_id=player_id))
        
        file = request.files['profile_photo']
        
        # Se usuário não selecionar um arquivo, o navegador envia um arquivo vazio
        if file.filename == '':
            flash('Nenhum arquivo selecionado.', 'error')
            return redirect(url_for('player_detail', player_id=player_id))
        
        if file and allowed_file(file.filename):
            # Criar nome de arquivo seguro e único
            filename = secure_filename(f"player_{player_id}_{int(datetime.now().timestamp())}.{file.filename.rsplit('.', 1)[1].lower()}")
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], filename)
            
            # Salvar o arquivo
            file.save(file_path)
            
            # Atualizar o banco de dados
            conn = get_db_connection()
            
            # Obter o caminho da foto anterior (se existir)
            old_photo = conn.execute('SELECT profile_photo FROM players WHERE id = ?', 
                                  (player_id,)).fetchone()
            
            # Atualizar para o novo caminho
            conn.execute('UPDATE players SET profile_photo = ? WHERE id = ?', 
                        (filename, player_id))
            conn.commit()
            
            # Remover a foto antiga se existir
            if old_photo and old_photo['profile_photo']:
                try:
                    old_file_path = os.path.join(app.config['UPLOAD_FOLDER'], old_photo['profile_photo'])
                    if os.path.exists(old_file_path):
                        os.remove(old_file_path)
                except Exception as e:
                    print(f"Erro ao remover arquivo antigo: {e}")
            
            conn.close()
            
            flash('Foto de perfil atualizada com sucesso!', 'success')
            return redirect(url_for('player_detail', player_id=player_id))
        else:
            flash('Tipo de arquivo não permitido. Use apenas JPG, PNG ou GIF.', 'error')
            return redirect(url_for('player_detail', player_id=player_id))
    
    except Exception as e:
        # Imprimir o erro no console do servidor para depuração
        import traceback
        print(f"Erro no upload de foto: {str(e)}")
        print(traceback.format_exc())
        flash(f'Erro ao processar o upload: {str(e)}', 'error')
        return redirect(url_for('player_detail', player_id=player_id))



@app.route('/remove_profile_photo/<int:player_id>', methods=['POST'])
@login_required
def remove_profile_photo(player_id):
    # Verificar se é o próprio jogador ou um admin
    if not (session.get('user_id') == player_id or session.get('is_admin', False)):
        flash('Acesso negado. Você só pode remover sua própria foto de perfil.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    conn = get_db_connection()
    
    # Obter o caminho da foto
    photo = conn.execute('SELECT profile_photo FROM players WHERE id = ?', 
                      (player_id,)).fetchone()
    
    if photo and photo['profile_photo']:
        # Remover do banco de dados
        conn.execute('UPDATE players SET profile_photo = NULL WHERE id = ?', (player_id,))
        conn.commit()
        
        # Remover arquivo físico
        try:
            file_path = os.path.join(app.config['UPLOAD_FOLDER'], photo['profile_photo'])
            if os.path.exists(file_path):
                os.remove(file_path)
        except Exception as e:
            print(f"Erro ao remover arquivo: {e}")
    
    conn.close()
    
    flash('Foto de perfil removida com sucesso!', 'success')
    return redirect(url_for('player_detail', player_id=player_id))




# Função para obter a data atual para uso nos templates
@app.context_processor
def utility_processor():
    def now():
        return datetime.now()
    return dict(now=now)

# Registrando a biblioteca datetime para que esteja disponível nos templates
@app.context_processor
def utility_processor_datetime():
    return dict(datetime=datetime)


# Após a linha onde você cria a aplicação Flask:
# app = Flask(__name__)

# Filtro para formatar data e hora
@app.template_filter('datetime')
def format_datetime(value, format='%d/%m/%Y %H:%M'):
    """Formata uma string de data para exibição."""
    if value is None:
        return ""
    
    try:
        if isinstance(value, str):
            dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        else:
            dt = value
        return dt.strftime(format)
    except:
        return value

@app.template_filter('country_code')
def country_code_filter(country_name):
    """
    Converte o nome do país para o código ISO de 2 letras usado para exibir bandeiras.
    """
    # Mapeamento de nomes de países para códigos ISO de 2 letras
    country_mapping = {
        'Brasil': 'br',
        'Argentina': 'ar',
        'Portugal': 'pt',
        'Estados Unidos': 'us',
        'Espanha': 'es',
        'Itália': 'it',
        'França': 'fr',
        'Alemanha': 'de',
        'Reino Unido': 'gb',
        'Inglaterra': 'gb-eng',
        'Escócia': 'gb-sct',
        'País de Gales': 'gb-wls',
        'Irlanda do Norte': 'gb-nir',
        'Japão': 'jp',
        'Coreia do Sul': 'kr',
        'China': 'cn',
        'Austrália': 'au',
        'Canadá': 'ca',
        'México': 'mx',
        'Chile': 'cl',
        'Colômbia': 'co',
        'Uruguai': 'uy',
        'Paraguai': 'py',
        'Peru': 'pe',
        'Venezuela': 've',  # ← ADICIONE ESTA LINHA
        'África do Sul': 'za',
        'Suíça': 'ch',
        'Suécia': 'se',
        'Noruega': 'no',
        'Dinamarca': 'dk',
        'Holanda': 'nl',
        'Países Baixos': 'nl',
        'Bélgica': 'be',
        'Irlanda': 'ie',
        'Nova Zelândia': 'nz',
        'Índia': 'in',
        'Rússia': 'ru',
        'Polônia': 'pl',
        'Áustria': 'at',
        'Grécia': 'gr',
        'Turquia': 'tr'
    }
    
    # Retorna o código ISO ou o nome do país em minúsculas como fallback
    return country_mapping.get(country_name, country_name.lower())

# Adicione este código perto do início do seu arquivo app.py, após a definição da aplicação Flask

# Filtro para formatar data e hora
@app.template_filter('datetime')
def format_datetime(value, format='%d/%m/%Y %H:%M'):
    """Formata uma string de data para exibição."""
    if value is None:
        return ""
    
    try:
        if isinstance(value, str):
            dt = datetime.strptime(value, '%Y-%m-%d %H:%M:%S')
        else:
            dt = value
        return dt.strftime(format)
    except:
        return value

# Função decoradora para verificar autenticação
def login_required(f):
    @wraps(f)
    def decorated_function(*args, **kwargs):
        if 'user_id' not in session:
            flash('Por favor, faça login para acessar esta página.', 'error')
            return redirect(url_for('login', next=request.url))
        return f(*args, **kwargs)
    return decorated_function

# Função para obter conexão com o banco de dados
def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# Função auxiliar para gerar hash de senha
def hash_password(password):
    """
    Método consistente de hash para senhas usando SHA-256.
    Garante que o mesmo password sempre produza o mesmo hash.
    """
    # Garante que a senha é uma string
    if not isinstance(password, str):
        password = str(password)
    
    # Codifica a senha para bytes e aplica o hash
    encoded_password = password.encode('utf-8')
    hashed = hashlib.sha256(encoded_password).hexdigest()
    
    return hashed

# Função para criar tabela de usuários e campos de senha na tabela players
def create_authentication_tables():
    conn = get_db_connection()
    
    # Adicionar coluna de senha à tabela players se não existir
    columns_info = conn.execute('PRAGMA table_info(players)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    if 'password' not in column_names:
        conn.execute('ALTER TABLE players ADD COLUMN password TEXT')
        print("Coluna 'password' adicionada à tabela players.")
    
    if 'last_login' not in column_names:
        conn.execute('ALTER TABLE players ADD COLUMN last_login DATETIME')
        print("Coluna 'last_login' adicionada à tabela players.")
    
    if 'reset_token' not in column_names:
        conn.execute('ALTER TABLE players ADD COLUMN reset_token TEXT')
        print("Coluna 'reset_token' adicionada à tabela players.")
    
    if 'reset_token_expiry' not in column_names:
        conn.execute('ALTER TABLE players ADD COLUMN reset_token_expiry DATETIME')
        print("Coluna 'reset_token_expiry' adicionada à tabela players.")
    
    # Verificar se a tabela admins existe
    tables = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='admins'").fetchall()
    
    # Se a tabela não existir ou se precisar recriar por falta de estrutura correta
    if not tables:
        print("Criando tabela de administradores...")
        # Tabela para administradores
        conn.execute('''
        CREATE TABLE IF NOT EXISTS admins (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            username TEXT NOT NULL UNIQUE,
            password TEXT NOT NULL,
            name TEXT NOT NULL,
            email TEXT,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            last_login DATETIME
        )
        ''')
        
        # Criar admin padrão (username: admin, senha: liga2025)
        conn.execute('INSERT INTO admins (username, password, name) VALUES (?, ?, ?)', 
                    ('admin', hash_password('liga2025'), 'Administrador'))
        print("Administrador padrão criado (usuário: admin, senha: liga2025).")
    else:
        # Verificar se a estrutura da tabela está correta
        admin_columns = conn.execute('PRAGMA table_info(admins)').fetchall()
        admin_column_names = [col[1] for col in admin_columns]
        
        # Se a coluna username não existir, recriar a tabela
        if 'username' not in admin_column_names:
            print("Reestruturando tabela de administradores...")
            # Fazer backup dos dados existentes, se houver
            try:
                admin_data = conn.execute('SELECT * FROM admins').fetchall()
            except:
                admin_data = []
            
            # Dropar e recriar a tabela com a estrutura correta
            conn.execute('DROP TABLE IF EXISTS admins')
            conn.execute('''
            CREATE TABLE admins (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                username TEXT NOT NULL UNIQUE,
                password TEXT NOT NULL,
                name TEXT NOT NULL,
                email TEXT,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                last_login DATETIME
            )
            ''')
            
            # Recriar admin padrão
            conn.execute('INSERT INTO admins (username, password, name) VALUES (?, ?, ?)', 
                        ('admin', hash_password('liga2025'), 'Administrador'))
            print("Administrador padrão recriado (usuário: admin, senha: liga2025).")
    
    # Verificar se já temos algum admin padrão
    try:
        admin = conn.execute('SELECT * FROM admins WHERE username = ?', ('admin',)).fetchone()
        if not admin:
            # Criar admin padrão (username: admin, senha: liga2025)
            conn.execute('INSERT INTO admins (username, password, name) VALUES (?, ?, ?)', 
                        ('admin', hash_password('liga2025'), 'Administrador'))
            print("Administrador padrão criado (usuário: admin, senha: liga2025).")
    except Exception as e:
        print(f"Erro ao verificar admin: {e}")
    
    # Definir senhas iniciais para todos os jogadores se a senha estiver vazia
    players = conn.execute('SELECT id, name, password FROM players WHERE active = 1').fetchall()
    for player in players:
        if not player['password']:
            # Senha inicial: 3 primeiras letras do nome em minúsculas
            default_password = player['name'].strip().lower()[:3]
            hashed_password = hash_password(default_password)
            
            conn.execute('UPDATE players SET password = ? WHERE id = ?', 
                       (hashed_password, player['id']))
            print(f"Senha inicial definida para o jogador {player['name']}")
    
    conn.commit()
    conn.close()
    print("Tabelas de autenticação verificadas com sucesso.")


def create_system_settings_table():
    conn = get_db_connection()
    conn.execute('''
    CREATE TABLE IF NOT EXISTS system_settings (
        key TEXT PRIMARY KEY,
        value TEXT,
        updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
    )
    ''')
    
    # Inserir configuração padrão para desafios
    conn.execute('''
    INSERT OR IGNORE INTO system_settings (key, value)
    VALUES ('challenges_locked', 'false')
    ''')
    
    conn.commit()
    conn.close()
    print("Tabela de configurações do sistema criada/verificada com sucesso.")



# Rota de login
@app.route('/login', methods=['GET', 'POST'])
def login():
    # Se o usuário já está logado, redirecionar para o dashboard
    if 'user_id' in session:
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        player_code = request.form.get('player_code', '').strip()  # Não converter para uppercase
        password = request.form.get('password', '')
        
        conn = get_db_connection()
        
        # SOLUÇÃO MELHORADA: 
        # 1. Verificar na tabela de administradores primeiro
        admin = conn.execute('SELECT * FROM admins WHERE username = ?', (player_code,)).fetchone()
        
        if admin:
            # É um administrador, verificar a senha
            if admin['password'] == hash_password(password):
                # Login bem-sucedido como administrador
                conn.execute('UPDATE admins SET last_login = CURRENT_TIMESTAMP WHERE id = ?', (admin['id'],))
                conn.commit()
                
                # Guardar ID do admin na sessão
                session['user_id'] = f"admin_{admin['id']}"
                session['username'] = admin['username']
                session['is_admin'] = True
                session.permanent = True
                
                flash(f'Bem-vindo, {admin["name"]}!', 'success')
                return redirect(url_for('admin_dashboard'))
            else:
                print(f"Senha incorreta para admin: {player_code}")
                flash('Credenciais inválidas. Tente novamente.', 'error')
        else:
            # 2. Se não for admin, verificar se é jogador
            player_code_upper = player_code.upper()  # Converter para uppercase para busca de jogador
            
            player = conn.execute('''
                SELECT * FROM players 
                WHERE player_code = ? AND active = 1
            ''', (player_code_upper,)).fetchone()
            
            if player and player['password'] == hash_password(password):
                # Login bem-sucedido como jogador
                conn.execute('UPDATE players SET last_login = CURRENT_TIMESTAMP WHERE id = ?', (player['id'],))
                conn.commit()
                
                # Guardar ID do jogador na sessão
                session['user_id'] = player['id']
                session['player_code'] = player['player_code']
                session['is_admin'] = False
                session.permanent = True
                
                flash(f'Bem-vindo, {player["name"]}!', 'success')
                return redirect(url_for('dashboard'))
            else:
                flash('Credenciais inválidas. Tente novamente.', 'error')
        
        conn.close()
    
    return render_template('login.html')





# Rota de logout
@app.route('/logout')
def logout():
    session.clear()
    flash('Você foi desconectado com sucesso.', 'success')
    return redirect(url_for('login'))

# Rota para troca de senha
@app.route('/change_password', methods=['GET', 'POST'])
@login_required
def change_password():
    if request.method == 'POST':
        old_password = request.form.get('old_password', '')
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        # Validar dados do formulário
        if not old_password or not new_password or not confirm_password:
            flash('Todos os campos são obrigatórios.', 'error')
            return redirect(url_for('change_password'))
        
        if new_password != confirm_password:
            flash('A nova senha e a confirmação não coincidem.', 'error')
            return redirect(url_for('change_password'))
        
        if len(new_password) < 4:
            flash('A nova senha deve ter pelo menos 4 caracteres.', 'error')
            return redirect(url_for('change_password'))
        
        # Verificar se a senha antiga está correta
        conn = get_db_connection()
        
        # Verificar se é um admin ou um jogador
        if session.get('is_admin', False):
            admin_id = session['user_id'].split('_')[1]
            user = conn.execute('SELECT * FROM admins WHERE id = ?', (admin_id,)).fetchone()
            user_type = 'admin'
        else:
            user = conn.execute('SELECT * FROM players WHERE id = ?', (session['user_id'],)).fetchone()
            user_type = 'player'
        
        if not user or user['password'] != hash_password(old_password):
            conn.close()
            flash('Senha atual incorreta.', 'error')
            return redirect(url_for('change_password'))
        
        # Atualizar a senha
        hashed_password = hash_password(new_password)
        
        if user_type == 'admin':
            conn.execute('UPDATE admins SET password = ? WHERE id = ?', 
                       (hashed_password, admin_id))
        else:
            conn.execute('UPDATE players SET password = ? WHERE id = ?', 
                       (hashed_password, session['user_id']))
        
        conn.commit()
        conn.close()
        
        flash('Senha alterada com sucesso!', 'success')
        
        # Redirecionar para o dashboard apropriado
        if user_type == 'admin':
            return redirect(url_for('admin_dashboard'))
        else:
            return redirect(url_for('dashboard'))
    
    return render_template('change_password.html')

# Rota para solicitar reset de senha
@app.route('/forgot_password', methods=['GET', 'POST'])
def forgot_password():
    if request.method == 'POST':
        player_code = request.form.get('player_code', '').strip().upper()
        
        if not player_code:
            flash('Por favor, informe seu código de jogador.', 'error')
            return redirect(url_for('forgot_password'))
        
        conn = get_db_connection()
        
        # Verificar se é um administrador pelo formato do código (admin ou admin_xyz)
        is_admin_code = player_code.lower() == 'admin' or player_code.lower().startswith('admin_')
        
        if is_admin_code:
            # Extrai o username do admin
            admin_username = player_code.split('_')[1] if '_' in player_code else 'admin'
            
            # Buscar admin pelo username
            admin = conn.execute('SELECT * FROM admins WHERE username = ?', (admin_username,)).fetchone()
            
            if admin:
                # Resetar a senha do admin para o próprio username
                new_password = admin_username
                hashed_password = hash_password(new_password)
                
                conn.execute('UPDATE admins SET password = ? WHERE id = ?', 
                           (hashed_password, admin['id']))
                
                conn.commit()
                conn.close()
                
                flash(f'Senha de administrador redefinida com sucesso. A nova senha é igual ao nome de usuário. Por favor, faça login e altere sua senha.', 'success')
                return redirect(url_for('login'))
            else:
                conn.close()
                flash('Administrador não encontrado.', 'error')
                return redirect(url_for('forgot_password'))
        else:
            # Buscar jogador pelo player_code
            player = conn.execute('''
                SELECT * FROM players 
                WHERE player_code = ? AND active = 1
            ''', (player_code,)).fetchone()
            
            if not player:
                conn.close()
                flash('Jogador não encontrado.', 'error')
                return redirect(url_for('forgot_password'))
            
            # Para fins de simplicidade, vamos resetar a senha para as 3 primeiras letras do nome
            default_password = player['name'].strip().lower()[:3]
            hashed_password = hash_password(default_password)
            
            # Atualizar a senha no banco de dados
            conn.execute('UPDATE players SET password = ? WHERE id = ?', 
                        (hashed_password, player['id']))
            
            conn.commit()
            conn.close()
            
            flash(f'A senha foi redefinida para as 3 primeiras letras do seu nome em minúsculas. Por favor, faça login e altere sua senha.', 'success')
            return redirect(url_for('login'))
    
    # Mostrar a página de "esqueci minha senha"
    return render_template('forgot_password.html')




# Rota para redefinir senha com token
@app.route('/reset_password/<token>', methods=['GET', 'POST'])
def reset_password(token):
    # Verificar se o token é válido
    conn = get_db_connection()
    player = conn.execute('''
        SELECT * FROM players 
        WHERE reset_token = ? AND datetime(reset_token_expiry) > datetime('now')
    ''', (token,)).fetchone()
    
    if not player:
        conn.close()
        flash('Link de redefinição de senha inválido ou expirado.', 'error')
        return redirect(url_for('login'))
    
    if request.method == 'POST':
        new_password = request.form.get('new_password', '')
        confirm_password = request.form.get('confirm_password', '')
        
        if not new_password or not confirm_password:
            flash('Todos os campos são obrigatórios.', 'error')
            return redirect(url_for('reset_password', token=token))
        
        if new_password != confirm_password:
            flash('A nova senha e a confirmação não coincidem.', 'error')
            return redirect(url_for('reset_password', token=token))
        
        if len(new_password) < 3:
            flash('A nova senha deve ter pelo menos 3 caracteres.', 'error')
            return redirect(url_for('reset_password', token=token))
        
        # Atualizar a senha e limpar o token
        hashed_password = hash_password(new_password)
        conn.execute('''
            UPDATE players 
            SET password = ?, reset_token = NULL, reset_token_expiry = NULL 
            WHERE id = ?
        ''', (hashed_password, player['id']))
        
        conn.commit()
        conn.close()
        
        flash('Senha redefinida com sucesso! Faça login com sua nova senha.', 'success')
        return redirect(url_for('login'))
    
    conn.close()
    return render_template('reset_password.html', token=token)

# ============================================================
# ROTA DASHBOARD - CORRIGIDA COM ALERTAS DE 2 DIAS
# ============================================================

@app.route('/dashboard')
@login_required
def dashboard():
    # Se for admin, redirecionar para o dashboard de admin
    if session.get('is_admin', False):
        return redirect(url_for('admin_dashboard'))
    
    conn = get_db_connection()
    
    # Buscar informações do jogador
    player = conn.execute('SELECT * FROM players WHERE id = ?', (session['user_id'],)).fetchone()
    
    if not player:
        session.clear()
        conn.close()
        flash('Sua conta não foi encontrada. Por favor, faça login novamente.', 'error')
        return redirect(url_for('login'))
    
    # Buscar desafios pendentes como desafiante
    challenges_as_challenger = conn.execute('''
        SELECT c.*, p.name as opponent_name, p.position as opponent_position
        FROM challenges c
        JOIN players p ON c.challenged_id = p.id
        WHERE c.challenger_id = ? AND c.status IN ('pending', 'accepted')
        ORDER BY c.scheduled_date
    ''', (session['user_id'],)).fetchall()
    
    # Buscar desafios onde o usuário é o desafiado e verificar prazos
    challenges_as_challenged = conn.execute('''
        SELECT c.*, p.name as opponent_name, p.position as opponent_position, c.response_deadline
        FROM challenges c
        JOIN players p ON c.challenger_id = p.id
        WHERE c.challenged_id = ? AND c.status IN ('pending', 'accepted')
        ORDER BY c.scheduled_date
    ''', (session['user_id'],)).fetchall()
    
    # Calcular dias restantes para cada desafio
    challenges_as_challenged_list = []
    for challenge in challenges_as_challenged:
        challenge_dict = dict(challenge)
        if challenge_dict['status'] == 'pending' and challenge_dict['response_deadline']:
            try:
                deadline_obj = datetime.strptime(challenge_dict['response_deadline'], '%Y-%m-%d %H:%M:%S')
                deadline_date = deadline_obj.date()
                today_date = datetime.now().date()
                
                days_remaining = (deadline_date - today_date).days
                
                challenge_dict['days_remaining'] = days_remaining
                challenge_dict['deadline_date'] = deadline_date.strftime('%Y-%m-%d')
            except Exception as e:
                print(f"Erro ao processar prazo de resposta: {str(e)}")
                challenge_dict['days_remaining'] = None
        challenges_as_challenged_list.append(challenge_dict)
    
    # Próximos 10 jogadores acima e abaixo na classificação
    players_above = conn.execute('''
        SELECT * FROM players 
        WHERE position < ? AND active = 1
        ORDER BY position DESC
        LIMIT 10
    ''', (player['position'],)).fetchall()
    
    players_below = conn.execute('''
        SELECT * FROM players 
        WHERE position > ? AND active = 1
        ORDER BY position
        LIMIT 10
    ''', (player['position'],)).fetchall()
    
    # Buscar jogadores que podem ser desafiados
    potential_challenges = []
    if player['active'] == 1 and player['position']:
        tier = player['tier']
        prev_tier = chr(ord(tier) - 1) if ord(tier) > ord('A') else tier
        min_position = max(1, player['position'] - 8)

        potential_challenges = conn.execute('''
            SELECT p.*
            FROM players p
            WHERE p.position < ? 
            AND p.position >= ?
            AND p.active = 1
            AND p.id NOT IN (
                SELECT challenged_id FROM challenges 
                WHERE challenger_id = ? AND status IN ('pending', 'accepted')
            )
            AND p.id NOT IN (
                SELECT challenger_id FROM challenges 
                WHERE challenged_id = ? AND status IN ('pending', 'accepted')
            )
            ORDER BY p.position DESC
        ''', (player['position'], min_position, player['id'], player['id'])).fetchall()
    
    conn.close()
    
    # ============================================================
    # ALERTAS DE DESAFIOS PENDENTES - PRAZO DE 2 DIAS
    # ============================================================
    for challenge in challenges_as_challenged_list:
        if challenge['status'] == 'pending' and 'days_remaining' in challenge:
            days_remaining = challenge['days_remaining']
            if days_remaining is not None:
                link = url_for("challenge_detail", challenge_id=challenge["id"])
                
                if days_remaining < 0:
                    # Prazo expirado
                    flash(f'⚠️ ATENÇÃO: Você foi desafiado por {challenge["opponent_name"]}. O prazo para responder EXPIROU há {abs(days_remaining)} dia(s)! <a href="{link}">Responder agora</a>.', 'danger')
                
                elif days_remaining == 0:
                    # Vence hoje
                    flash(f'⏰ URGENTE: Você foi desafiado por {challenge["opponent_name"]}! O prazo para responder vence HOJE. <a href="{link}">Responder agora</a>.', 'warning')
                
                elif days_remaining == 1:
                    # Vence amanhã
                    flash(f'⏳ ATENÇÃO: Você foi desafiado por {challenge["opponent_name"]}! Você tem apenas 1 dia para responder. <a href="{link}">Responder agora</a>.', 'warning')
                
                else:
                    # Ainda no prazo (2 dias ou mais - não deveria acontecer com prazo de 2 dias)
                    flash(f'📩 Você foi desafiado por {challenge["opponent_name"]}! Você tem {days_remaining} dias para aceitar ou rejeitar. <a href="{link}">Ver desafio</a>.', 'info')
    
    return render_template('dashboard.html', 
                          player=player,
                          challenges_as_challenger=challenges_as_challenger,
                          challenges_as_challenged=challenges_as_challenged_list,
                          players_above=players_above,
                          players_below=players_below,
                          potential_challenges=potential_challenges)


# Dashboard do administrador
@app.route('/admin')
@login_required
def admin_dashboard():
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    # Estatísticas gerais
    stats = {}
    
    # Total de jogadores ativos
    active_players = conn.execute('SELECT COUNT(*) as count FROM players WHERE active = 1').fetchone()
    stats['active_players'] = active_players['count']
    
    # Total de desafios pendentes
    pending_challenges = conn.execute('SELECT COUNT(*) as count FROM challenges WHERE status IN ("pending", "accepted")').fetchone()
    stats['pending_challenges'] = pending_challenges['count']
    
    # Verificar se a tabela system_settings existe
    table_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='system_settings'").fetchone()
    challenges_locked = False
    
    if table_exists:
        # Verificar se os desafios estão bloqueados
        setting = conn.execute('SELECT value FROM system_settings WHERE key = ?', ('challenges_locked',)).fetchone()
        challenges_locked = setting and setting['value'] == 'true'
    else:
        # Criar a tabela se não existir
        conn.execute('''
        CREATE TABLE IF NOT EXISTS system_settings (
            key TEXT PRIMARY KEY,
            value TEXT,
            updated_at DATETIME DEFAULT CURRENT_TIMESTAMP
        )
        ''')
        
        # Inserir configuração padrão para desafios
        conn.execute('''
        INSERT OR IGNORE INTO system_settings (key, value)
        VALUES ('challenges_locked', 'false')
        ''')
        conn.commit()
    
    # Desafios recentes
    recent_challenges = conn.execute('''
        SELECT c.*, 
               p1.name as challenger_name, 
               p2.name as challenged_name
        FROM challenges c
        JOIN players p1 ON c.challenger_id = p1.id
        JOIN players p2 ON c.challenged_id = p2.id
        ORDER BY c.created_at DESC
        LIMIT 10
    ''').fetchall()
    
    # Jogadores que nunca fizeram login
    never_logged = conn.execute('''
        SELECT * FROM players
        WHERE last_login IS NULL AND active = 1
        ORDER BY name
    ''').fetchall()
    
    conn.close()
    
    return render_template('admin_dashboard.html', 
                          stats=stats,
                          recent_challenges=recent_challenges,
                          never_logged=never_logged,
                          challenges_locked=challenges_locked)




# Inicialização da aplicação
if __name__ == '__main__':
    # Verificar se o banco de dados existe, caso contrário, importar dados
    if not os.path.exists(DATABASE):
        print("Banco de dados não encontrado. Executando script de importação...")
        import import_data
        import_data.create_database()
        import_data.import_players_data(import_data.cursor)
    
    # Criar tabelas de autenticação
    create_authentication_tables()

# 1. NOVA ESTRUTURA ESTENDIDA DA PIRÂMIDE
PYRAMID_STRUCTURE = {
    'A': [1, 2, 3, 4, 5],                                                    # 5 posições (1-5)
    'B': [6, 7, 8, 9, 10, 11, 12],                                           # 7 posições (6-12)
    'C': [13, 14, 15, 16, 17, 18, 19, 20, 21],                               # 9 posições (13-21)
    'D': [22, 23, 24, 25, 26, 27, 28, 29, 30, 31, 32],                       # 11 posições (22-32)
    'E': [33, 34, 35, 36, 37, 38, 39, 40, 41, 42, 43, 44, 45],               # 13 posições (33-45)
    'F': [46, 47, 48, 49, 50, 51, 52, 53, 54, 55, 56, 57, 58, 59, 60],       # 15 posições (46-60)
    'G': [61, 62, 63, 64, 65, 66, 67, 68, 69, 70, 71, 72, 73, 74, 75, 76, 77], # 17 posições (61-77)
    'H': [78, 79, 80, 81, 82, 83, 84, 85, 86, 87, 88, 89, 90, 91, 92, 93, 94, 95, 96], # 19 posições (78-96)
    'I': [97, 98, 99, 100, 101, 102, 103, 104, 105, 106, 107, 108, 109, 110, 111, 112, 113, 114, 115, 116, 117], # 21 posições (97-117)
    'J': [118, 119, 120, 121, 122, 123, 124, 125, 126, 127, 128, 129, 130, 131, 132, 133, 134, 135, 136, 137, 138, 139, 140], # 23 posições (118-140)
}


def get_db_connection():
    conn = sqlite3.connect(DATABASE)
    conn.row_factory = sqlite3.Row
    return conn

# Função para criar a tabela de histórico diário
def create_daily_history_table():
    conn = get_db_connection()
    conn.execute('''
    CREATE TABLE IF NOT EXISTS daily_ranking_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        player_id INTEGER NOT NULL,
        position INTEGER NOT NULL,
        tier TEXT NOT NULL,
        date_recorded DATE NOT NULL,
        FOREIGN KEY (player_id) REFERENCES players(id)
    )
    ''')


def create_business_table():
    try:
        conn = get_db_connection()
        
        # Verificar se a tabela já existe
        table_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='businesses'").fetchone()
        
        if not table_exists:
            conn.execute('''
            CREATE TABLE IF NOT EXISTS businesses (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                player_id INTEGER NOT NULL,
                name TEXT NOT NULL,
                category TEXT NOT NULL,
                description TEXT NOT NULL,
                image_path TEXT,
                contact_info TEXT,
                active INTEGER DEFAULT 1,
                created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (player_id) REFERENCES players(id)
            )
            ''')
            
            # Criar pasta para imagens de negócios
            business_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'business_images')
            os.makedirs(business_upload_folder, exist_ok=True)
            
            print("Tabela de negócios criada com sucesso e pasta de imagens verificada.")
        else:
            print("Tabela de negócios já existe.")
        
        conn.commit()
        conn.close()
        return True
    except Exception as e:
        print(f"ERRO ao criar tabela de negócios: {str(e)}")
        return False


# Função para criar tabela de histórico de handicap
def create_hcp_history_table():
    conn = get_db_connection()
    conn.execute('''
    CREATE TABLE IF NOT EXISTS hcp_history (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        player_id INTEGER NOT NULL,
        old_hcp REAL,
        new_hcp REAL NOT NULL,
        modified_by TEXT NOT NULL,
        notes TEXT,
        change_date DATETIME DEFAULT CURRENT_TIMESTAMP,
        FOREIGN KEY (player_id) REFERENCES players(id)
    )
    ''')
    
    # Criar um índice para melhorar a performance das consultas
    conn.execute('''
    CREATE INDEX IF NOT EXISTS idx_hcp_history_player_date 
    ON hcp_history (player_id, change_date)
    ''')
    
    conn.commit()
    conn.close()
    print("Tabela de histórico de handicap criada com sucesso.")

# Função para registrar alterações de handicap
def record_hcp_change(player_id, old_hcp, new_hcp, modified_by, notes=None):
    """
    Registra alterações no handicap de um jogador.
    
    Args:
        player_id: ID do jogador
        old_hcp: Handicap anterior (pode ser None)
        new_hcp: Novo handicap
        modified_by: Quem modificou ('admin', 'player', etc)
        notes: Observações sobre a alteração (opcional)
    """
    conn = get_db_connection()
    
    try:
        conn.execute('''
            INSERT INTO hcp_history 
            (player_id, old_hcp, new_hcp, modified_by, notes)
            VALUES (?, ?, ?, ?, ?)
        ''', (player_id, old_hcp, new_hcp, modified_by, notes))
        
        conn.commit()
    except Exception as e:
        conn.rollback()
        print(f"Erro ao registrar alteração de handicap: {str(e)}")
    finally:
        conn.close()



"""
Copie esta função corrigida para substituir a existente no seu arquivo app.py
"""

@app.route('/player/<int:player_id>/hcp_history')
def player_hcp_history(player_id):
    """
    Exibe o histórico de handicap de um jogador específico.
    Versão corrigida para tratamento de erros.
    """
    try:
        conn = get_db_connection()
        
        # Verificar se o jogador existe
        player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
        
        if not player:
            conn.close()
            flash('Jogador não encontrado!', 'error')
            return redirect(url_for('index'))
        
        # Verificar se a tabela hcp_history existe
        table_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='hcp_history'").fetchone()
        
        if not table_exists:
            # Criar a tabela se não existir
            conn.execute('''
            CREATE TABLE IF NOT EXISTS hcp_history (
                id INTEGER PRIMARY KEY AUTOINCREMENT,
                player_id INTEGER NOT NULL,
                old_hcp REAL,
                new_hcp REAL NOT NULL,
                modified_by TEXT NOT NULL,
                notes TEXT,
                change_date DATETIME DEFAULT CURRENT_TIMESTAMP,
                FOREIGN KEY (player_id) REFERENCES players(id)
            )
            ''')
            conn.commit()
        
        # Buscar histórico de handicap do jogador
        history = conn.execute('''
            SELECT hcp_history.*, players.name as player_name
            FROM hcp_history 
            JOIN players ON hcp_history.player_id = players.id
            WHERE player_id = ?
            ORDER BY change_date DESC
        ''', (player_id,)).fetchall()
        
        conn.close()
        
        # Converter os dados de Row para dict para evitar problemas
        history_list = []
        for item in history:
            # Copia os valores para um dicionário, garantindo tratamento adequado
            item_dict = {}
            for key in item.keys():
                item_dict[key] = item[key]
            history_list.append(item_dict)
        
        return render_template('player_hcp_history.html', 
                              player=player,
                              history=history_list)  # Enviar a lista convertida
    
    except Exception as e:
        # Tratar qualquer exceção para exibir uma mensagem útil ao usuário
        import traceback
        error_details = traceback.format_exc()
        
        # Registrar o erro para debug
        print(f"Erro ao acessar histórico de HCP: {str(e)}")
        print(error_details)
        
        # Mostrar mensagem amigável ao usuário
        flash(f'Erro ao carregar o histórico de handicap: {str(e)}', 'error')
        return redirect(url_for('player_detail', player_id=player_id))


# MODIFICAÇÃO: Melhoria na função record_daily_rankings para permitir sobrescrever registros
def record_daily_rankings(force_update=False):
    """
    Registra as posições diárias de todos os jogadores.
    Se force_update=True, registros existentes serão substituídos.
    """
    conn = get_db_connection()
    today = datetime.now().date()
    
    # Verificar se já temos registros para hoje
    existing = conn.execute(
        'SELECT COUNT(*) as count FROM daily_ranking_history WHERE date_recorded = ?', 
        (today.strftime('%Y-%m-%d'),)
    ).fetchone()
    
    if existing and existing['count'] > 0 and not force_update:
        print(f"Já existem registros para {today}. Pulando...")
        conn.close()
        return False
    
    # Se existem registros e force_update=True, remover registros existentes
    if existing and existing['count'] > 0 and force_update:
        conn.execute('DELETE FROM daily_ranking_history WHERE date_recorded = ?', 
                    (today.strftime('%Y-%m-%d'),))
        print(f"Removidos registros existentes de {today} para atualização forçada")
    
    try:
        # Obter todos os jogadores ativos
        players = conn.execute('SELECT id, position, tier FROM players WHERE active = 1 ORDER BY position').fetchall()
        
        # Registrar posição atual de cada jogador
        for player in players:
            conn.execute('''
                INSERT INTO daily_ranking_history 
                (player_id, position, tier, date_recorded)
                VALUES (?, ?, ?, ?)
            ''', (player['id'], player['position'], player['tier'], today.strftime('%Y-%m-%d')))
        
        conn.commit()
        print(f"Registrados {len(players)} jogadores no histórico diário para {today}")
        return True
    except Exception as e:
        conn.rollback()
        print(f"Erro ao registrar histórico diário: {str(e)}")
        return False
    finally:
        conn.close()


def sync_ranking_history_tables(conn=None, specific_date=None):
    """
    Sincroniza as tabelas ranking_history e daily_ranking_history.
    Se uma data específica for fornecida, sincroniza apenas para essa data.
    Caso contrário, sincroniza para a data atual.
    
    Args:
        conn: Conexão com o banco de dados (opcional)
        specific_date: Data específica para sincronização (opcional)
    """
    # Determinar se precisamos criar e fechar a conexão
    connection_provided = conn is not None
    if not connection_provided:
        conn = get_db_connection()
    
    try:
        # Determinar a data para sincronização
        if specific_date:
            target_date = specific_date
        else:
            target_date = datetime.now().date()
        
        target_date_str = target_date.strftime('%Y-%m-%d')
        
        # Verificar se existem registros no daily_ranking_history para a data alvo
        existing = conn.execute(
            'SELECT COUNT(*) as count FROM daily_ranking_history WHERE date_recorded = ?', 
            (target_date_str,)
        ).fetchone()
        
        # Obter todas as alterações de ranking para a data alvo
        ranking_changes = conn.execute('''
            SELECT player_id, new_position, new_tier, change_date 
            FROM ranking_history 
            WHERE date(change_date) = ? 
            ORDER BY change_date DESC
        ''', (target_date_str,)).fetchall()
        
        # Se existem alterações para hoje, vamos usar as informações mais recentes
        # para atualizar ou criar o registro diário
        if ranking_changes:
            # Remover registros existentes para a data alvo
            conn.execute('DELETE FROM daily_ranking_history WHERE date_recorded = ?', 
                       (target_date_str,))
            
            # Mapear as posições mais recentes para cada jogador alterado hoje
            player_latest_positions = {}
            for change in ranking_changes:
                player_id = change['player_id']
                if player_id not in player_latest_positions:
                    player_latest_positions[player_id] = {
                        'position': change['new_position'],
                        'tier': change['new_tier']
                    }
            
            # Obter todos os jogadores ativos
            all_players = conn.execute('SELECT id, position, tier FROM players WHERE active = 1').fetchall()
            
            # Inserir registros diários atualizados
            for player in all_players:
                player_id = player['id']
                
                # Se o jogador teve alteração hoje, use a posição da alteração
                if player_id in player_latest_positions:
                    position = player_latest_positions[player_id]['position']
                    tier = player_latest_positions[player_id]['tier']
                # Caso contrário, use a posição atual
                else:
                    position = player['position']
                    tier = player['tier']
                
                # Inserir registro diário
                conn.execute('''
                    INSERT INTO daily_ranking_history 
                    (player_id, position, tier, date_recorded)
                    VALUES (?, ?, ?, ?)
                ''', (player_id, position, tier, target_date_str))
            
            print(f"Sincronizado histórico diário para {target_date_str} com base em {len(ranking_changes)} alterações")
        # Se não existem alterações para a data alvo e não existem registros diários
        elif not existing or existing['count'] == 0:
            # Registrar snapshot das posições atuais
            record_daily_rankings(force_update=True)
            print(f"Criado novo snapshot para {target_date_str} por não existirem alterações ou registros")
        
        # Se não chegamos aqui, é porque já existem registros diários e não há alterações
        # para a data alvo, então não precisamos fazer nada
        
        if not connection_provided:
            conn.commit()
        
    except Exception as e:
        print(f"Erro ao sincronizar histórico: {str(e)}")
        if not connection_provided:
            conn.rollback()
    finally:
        if not connection_provided:
            conn.close()


@app.route('/force_record_daily', methods=['GET'])
def force_record_daily():
    conn = get_db_connection()
    today = datetime.now().date()
    
    try:
        # Remover registros existentes para hoje
        conn.execute(
            'DELETE FROM daily_ranking_history WHERE date_recorded = ?', 
            (today.strftime('%Y-%m-%d'),)
        )
        
        # Obter todos os jogadores ativos
        players = conn.execute('SELECT id, position, tier FROM players WHERE active = 1 ORDER BY position').fetchall()
        
        # Registrar posição atual de cada jogador
        for player in players:
            conn.execute('''
                INSERT INTO daily_ranking_history 
                (player_id, position, tier, date_recorded)
                VALUES (?, ?, ?, ?)
            ''', (player['id'], player['position'], player['tier'], today.strftime('%Y-%m-%d')))
        
        conn.commit()
        flash(f'Posições atualizadas com sucesso no histórico para {today.strftime("%d/%m/%Y")}!', 'success')
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar histórico: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('index'))

# Rota para diagnosticar e corrigir o histórico
@app.route('/fix_history', methods=['GET'])
def fix_history():
    """
    Verifica e corrige problemas no histórico diário:
    - Remove posições duplicadas para a mesma data
    - Garante que não há lacunas nas posições para cada data
    """
    conn = get_db_connection()
    
    try:
        # Buscar todas as datas distintas no histórico
        dates = conn.execute(
            'SELECT DISTINCT date_recorded FROM daily_ranking_history ORDER BY date_recorded'
        ).fetchall()
        
        total_fixed = 0
        
        for date_record in dates:
            date = date_record['date_recorded']
            
            # Verificar posições duplicadas na mesma data
            duplicates = conn.execute('''
                SELECT position, COUNT(*) as count
                FROM daily_ranking_history
                WHERE date_recorded = ?
                GROUP BY position
                HAVING COUNT(*) > 1
            ''', (date,)).fetchall()
            
            # Se encontrar duplicatas, corrigir
            if duplicates:
                for dup in duplicates:
                    position = dup['position']
                    
                    # Buscar jogadores com esta posição duplicada
                    players_with_dup = conn.execute('''
                        SELECT h.id, h.player_id, p.name
                        FROM daily_ranking_history h
                        JOIN players p ON h.player_id = p.id
                        WHERE h.date_recorded = ? AND h.position = ?
                        ORDER BY h.id
                    ''', (date, position)).fetchall()
                    
                    # Manter apenas o primeiro registro (o mais antigo) e remover os outros
                    if len(players_with_dup) > 1:
                        for player in players_with_dup[1:]:
                            conn.execute('DELETE FROM daily_ranking_history WHERE id = ?', (player['id'],))
                            total_fixed += 1
            
            # Verificar se há lacunas nas posições sequenciais para esta data
            positions = conn.execute('''
                SELECT position 
                FROM daily_ranking_history
                WHERE date_recorded = ?
                ORDER BY position
            ''', (date,)).fetchall()
            
            positions_list = [p['position'] for p in positions]
            expected_positions = list(range(1, len(positions_list) + 1))
            
            if positions_list != expected_positions:
                # Há uma discrepância - recalcular posições
                records = conn.execute('''
                    SELECT id, player_id
                    FROM daily_ranking_history
                    WHERE date_recorded = ?
                    ORDER BY position
                ''', (date,)).fetchall()
                
                # Atualizar posições para serem sequenciais
                for i, record in enumerate(records, 1):
                    conn.execute('''
                        UPDATE daily_ranking_history
                        SET position = ?
                        WHERE id = ?
                    ''', (i, record['id']))
                    
                    # Também atualizar o tier com base na nova posição
                    tier = get_tier_from_position(i)
                    conn.execute('''
                        UPDATE daily_ranking_history
                        SET tier = ?
                        WHERE id = ?
                    ''', (tier, record['id']))
                    
                    total_fixed += 1
        
        conn.commit()
        
        if total_fixed > 0:
            flash(f'Histórico corrigido: {total_fixed} problemas resolvidos.', 'success')
        else:
            flash('Nenhum problema encontrado no histórico.', 'info')
            
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao corrigir o histórico: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('ranking_history'))


# 2. FUNÇÃO CORRIGIDA DE CÁLCULO DE TIER
def get_tier_from_position(position):
    """
    Determina o nível (tier) com base na posição na pirâmide.
    Estrutura: A:5, B:7, C:9, D:11... (+2 a cada tier)
    """
    # Verificar em cada tier definido na estrutura
    for tier, positions in PYRAMID_STRUCTURE.items():
        if position in positions:
            return tier
    
    # Para posições que excederam a estrutura definida
    last_tier_letter = list(PYRAMID_STRUCTURE.keys())[-1]
    last_tier_positions = PYRAMID_STRUCTURE[last_tier_letter]
    last_tier_end = max(last_tier_positions)
    
    # Se a posição está além da estrutura definida
    if position > last_tier_end:
        remaining_position = position - last_tier_end
        current_tier_letter = last_tier_letter
        current_tier_size = len(PYRAMID_STRUCTURE[last_tier_letter])
        position_counter = 0
        
        while position_counter < remaining_position:
            current_tier_letter = chr(ord(current_tier_letter) + 1)
            current_tier_size += 2
            
            if position_counter + current_tier_size >= remaining_position:
                return current_tier_letter
            
            position_counter += current_tier_size
        
        return current_tier_letter
    
    return 'A'



# Função para atualizar todos os tiers baseado nas posições atuais
def update_all_tiers(conn):
    """
    Atualiza o tier de todos os jogadores com base em suas posições atuais e na estrutura fixa da pirâmide.
    """
    # Buscar todos os jogadores ordenados por posição
    players = conn.execute('SELECT id, position FROM players WHERE active = 1 AND position IS NOT NULL ORDER BY position').fetchall()
    
    # Atualizar o tier de cada jogador com base em sua posição
    for player in players:
        position = player['position']
        correct_tier = get_tier_from_position(position)
        
        # Atualizar o tier no banco de dados
        conn.execute('UPDATE players SET tier = ? WHERE id = ?', (correct_tier, player['id']))
    
    conn.commit()
    print("Todos os tiers atualizados com base nas posições fixas da pirâmide.")

# Função para verificar a estrutura da pirâmide
def verify_pyramid_structure(conn):
    """
    Verifica se todos os jogadores estão no tier correto de acordo com suas posições.
    Retorna uma lista de jogadores com tiers incorretos.
    """
    players = conn.execute('SELECT id, name, position, tier FROM players WHERE active = 1 AND position IS NOT NULL ORDER BY position').fetchall()
    incorrect_players = []
    
    for player in players:
        correct_tier = get_tier_from_position(player['position'])
        if player['tier'] != correct_tier:
            incorrect_players.append({
                'id': player['id'],
                'name': player['name'],
                'position': player['position'],
                'current_tier': player['tier'],
                'correct_tier': correct_tier
            })
    
    return incorrect_players

# Nova função para verificar e corrigir lacunas nas posições
def fix_position_gaps(conn):
    """
    Verifica se há lacunas nas posições dos jogadores e as corrige, 
    garantindo que as posições sejam sequenciais (1, 2, 3, ...).
    """
    # Buscar todos os jogadores ordenados por posição atual
    players = conn.execute('SELECT id, position FROM players WHERE active = 1 AND position IS NOT NULL ORDER BY position').fetchall()
    
    # Verificar e corrigir lacunas
    expected_position = 1
    for player in players:
        if player['position'] != expected_position:
            # Corrigir a posição se não estiver na sequência esperada
            conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                       (expected_position, player['id']))
            print(f"Corrigida posição do jogador ID {player['id']}: {player['position']} -> {expected_position}")
        expected_position += 1
    
    # Não é necessário commitar aqui, pois essa função é chamada dentro de outra
    # que já tem seu próprio commit



# Função aprimorada para ajustar a pirâmide quando ocorrem mudanças de posição
def rebalance_positions_after_challenge(conn, winner_id, loser_id, winner_new_pos, loser_new_pos):
    """
    Ajusta as posições de todos os jogadores após um desafio, mantendo a sequência correta.
    Versão melhorada que lida corretamente com todos os cenários de movimentação.
    """
    # Buscar posições atuais
    winner_data = conn.execute('SELECT position FROM players WHERE id = ?', (winner_id,)).fetchone()
    loser_data = conn.execute('SELECT position FROM players WHERE id = ?', (loser_id,)).fetchone()
    
    if not winner_data or not loser_data:
        print("Erro: Jogador não encontrado")
        return
        
    winner_old_pos = winner_data['position']
    loser_old_pos = loser_data['position']
    
    # Caso 1: Se o vencedor está subindo (posição menor numericamente é melhor)
    if winner_new_pos < winner_old_pos:
        # Primeiro, atualizar todos os jogadores entre as posições (ajustar uma posição para baixo)
        conn.execute('''
            UPDATE players 
            SET position = position + 1 
            WHERE position >= ? AND position < ?
            AND id != ? AND id != ?
            AND active = 1
        ''', (winner_new_pos, winner_old_pos, winner_id, loser_id))
        
        # Em seguida, definir as novas posições para vencedor e perdedor
        conn.execute('UPDATE players SET position = ? WHERE id = ?', (winner_new_pos, winner_id))
        
        # O perdedor só muda de posição se ele for o jogador diretamente desafiado
        if loser_old_pos == winner_new_pos:
            conn.execute('UPDATE players SET position = ? WHERE id = ?', (loser_new_pos, loser_id))
    
    # Caso 2: Caso especial ou ajuste direto de posições
    else:
        # Definir as novas posições diretamente
        conn.execute('UPDATE players SET position = ? WHERE id = ?', (winner_new_pos, winner_id))
        conn.execute('UPDATE players SET position = ? WHERE id = ?', (loser_new_pos, loser_id))
    
    # Verificar se há lacunas nas posições e corrigi-las
    fix_position_gaps(conn)
    
    # Atualizar todos os tiers com base nas novas posições
    update_all_tiers(conn)
    
    conn.commit()
    print("Posições e tiers rebalanceados após o desafio.")

# Função aprimorada para processar o resultado de um desafio
# Adição de código para função existente process_challenge_result

# ============================================================
# FUNÇÃO ATUALIZADA - process_challenge_result
# 
# INSTRUÇÕES: Substitua a função existente no app.py por esta versão
#
# REGRAS DE W.O. IMPLEMENTADAS:
# 1. W.O para o desafiado (wo_challenger - desafiado não compareceu):
#    - Desafiante ganha 1 posição (permuta com quem está acima dele)
#    - Desafiado assume a posição antiga do desafiante
#
# 2. W.O para o desafiante (wo_challenged - desafiante não compareceu):
#    - Desafiante perde 4 posições no ranking
#    - Desafiado não muda de posição
# ============================================================

def process_challenge_result(conn, challenge_id, status, result):
    """
    Processa o resultado de um desafio, atualizando posições conforme as regras:
    
    REGRAS NORMAIS:
    - Desafiante vence: assume posição do desafiado, desafiado desce 1 posição
    - Desafiado vence: desafiado sobe 1 posição (permuta com quem está acima), desafiante NÃO muda
    
    REGRAS DE W.O.:
    - wo_challenger (desafiado não compareceu - desafiante vence por WO):
      * Desafiante ganha 1 posição (permuta com quem está acima dele)
      * Desafiado assume a posição antiga do desafiante
    
    - wo_challenged (desafiante não compareceu - desafiado vence por WO):
      * Desafiante perde 4 posições no ranking
      * Desafiado não muda de posição
    """
    # Buscar o result_type do desafio
    challenge_data = conn.execute('SELECT result_type FROM challenges WHERE id = ?', (challenge_id,)).fetchone()
    result_type = challenge_data['result_type'] if challenge_data and challenge_data['result_type'] else 'normal'
    
    # Atualizar o status e resultado do desafio
    conn.execute('UPDATE challenges SET status = ?, result = ? WHERE id = ?', 
                (status, result, challenge_id))
    
    # Se for "Concluído (com pendência)", apenas registramos o resultado sem alterar posições
    if status == 'completed_pending':
        conn.commit()
        return
    
    if status == 'completed' and result:
        # Buscar informações detalhadas do desafio
        challenge = conn.execute('''
            SELECT c.*, 
                   p1.id as challenger_id, p1.position as challenger_position, p1.tier as challenger_tier, p1.sexo as challenger_sexo,
                   p2.id as challenged_id, p2.position as challenged_position, p2.tier as challenged_tier, p2.sexo as challenged_sexo
            FROM challenges c
            JOIN players p1 ON c.challenger_id = p1.id
            JOIN players p2 ON c.challenged_id = p2.id
            WHERE c.id = ?
        ''', (challenge_id,)).fetchone()
        
        if not challenge:
            print(f"Erro: Desafio ID {challenge_id} não encontrado")
            conn.rollback()
            return
        
        # Guardar posições e tiers antigos para histórico
        challenger_id = challenge['challenger_id']
        challenger_old_pos = challenge['challenger_position']
        challenger_old_tier = challenge['challenger_tier']
        challenged_id = challenge['challenged_id']
        challenged_old_pos = challenge['challenged_position']
        challenged_old_tier = challenge['challenged_tier']
        player_sexo = challenge['challenger_sexo'] or 'masculino'
        
        try:
            # =====================================================
            # W.O. - DESAFIADO NÃO COMPARECEU (wo_challenger)
            # Desafiante vence por WO
            # Desafiante ganha 1 posição, desafiado vai para posição do desafiante
            # =====================================================
            if result_type == 'wo_challenger' and result == 'challenger_win':
                print(f"🔴 Processando W.O. - DESAFIADO não compareceu")
                print(f"   Posições antes: Desafiante={challenger_old_pos}, Desafiado={challenged_old_pos}")
                
                # O desafiante sobe 1 posição (permuta com quem está imediatamente acima dele)
                if challenger_old_pos > 1:
                    # Buscar o jogador que está 1 posição acima do desafiante
                    player_above = conn.execute('''
                        SELECT id, position, tier FROM players 
                        WHERE position = ? AND active = 1
                        AND (sexo = ? OR (sexo IS NULL AND ? != 'feminino'))
                    ''', (challenger_old_pos - 1, player_sexo, player_sexo)).fetchone()
                    
                    if player_above:
                        player_above_id = player_above['id']
                        player_above_old_pos = player_above['position']
                        player_above_old_tier = player_above['tier']
                        
                        new_challenger_pos = challenger_old_pos - 1  # Desafiante sobe 1
                        new_challenged_pos = challenger_old_pos  # Desafiado vai para posição antiga do desafiante
                        
                        # Se o jogador acima for o próprio desafiado
                        if player_above_id == challenged_id:
                            # Permuta direta: desafiante sobe, desafiado desce
                            conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                                       (new_challenger_pos, challenger_id))
                            conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                                       (new_challenged_pos, challenged_id))
                        else:
                            # Há um jogador diferente acima do desafiante
                            # Desafiante permuta com ele
                            # Desafiado vai para a posição antiga do desafiante
                            
                            # Jogador que estava acima do desafiante desce para posição do desafiante
                            conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                                       (challenger_old_pos, player_above_id))
                            
                            # Desafiante sobe 1
                            conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                                       (new_challenger_pos, challenger_id))
                            
                            # Desafiado vai para posição após o jogador que desceu (se aplicável)
                            # Precisamos recalcular a posição do desafiado
                            if challenged_old_pos < challenger_old_pos:
                                # Desafiado estava acima do desafiante, vai para posição do desafiante
                                new_challenged_pos = challenger_old_pos
                            else:
                                # Desafiado estava abaixo ou na mesma posição (não deveria acontecer)
                                new_challenged_pos = challenger_old_pos
                            
                            conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                                       (new_challenged_pos, challenged_id))
                            
                            # Registrar no histórico - Jogador que foi deslocado
                            conn.execute('''
                                INSERT INTO ranking_history 
                                (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                                VALUES (?, ?, ?, ?, ?, ?, ?)
                            ''', (player_above_id, player_above_old_pos, challenger_old_pos, 
                                 player_above_old_tier, get_tier_from_position(challenger_old_pos), 
                                 'displaced_by_wo', challenge_id))
                        
                        # Registrar no histórico - Desafiante
                        conn.execute('''
                            INSERT INTO ranking_history 
                            (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (challenger_id, challenger_old_pos, new_challenger_pos, 
                             challenger_old_tier, get_tier_from_position(new_challenger_pos), 
                             'wo_win_promoted', challenge_id))
                        
                        # Registrar no histórico - Desafiado
                        conn.execute('''
                            INSERT INTO ranking_history 
                            (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (challenged_id, challenged_old_pos, new_challenged_pos, 
                             challenged_old_tier, get_tier_from_position(new_challenged_pos), 
                             'wo_loss_demoted', challenge_id))
                        
                        print(f"✅ W.O. Desafiado não compareceu:")
                        print(f"   Desafiante {challenger_id}: {challenger_old_pos} → {new_challenger_pos}")
                        print(f"   Desafiado {challenged_id}: {challenged_old_pos} → {new_challenged_pos}")
                    else:
                        # Não há ninguém acima do desafiante
                        # Apenas o desafiado vai para posição do desafiante + 1
                        new_challenged_pos = challenger_old_pos + 1
                        
                        conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                                   (new_challenged_pos, challenged_id))
                        
                        conn.execute('''
                            INSERT INTO ranking_history 
                            (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                            VALUES (?, ?, ?, ?, ?, ?, ?)
                        ''', (challenged_id, challenged_old_pos, new_challenged_pos, 
                             challenged_old_tier, get_tier_from_position(new_challenged_pos), 
                             'wo_loss_demoted', challenge_id))
                        
                        print(f"✅ W.O. (sem jogador acima): Desafiado {challenged_id} ({challenged_old_pos} → {new_challenged_pos})")
                else:
                    # Desafiante já está na posição 1
                    # Apenas o desafiado vai para posição do desafiante + 1
                    new_challenged_pos = challenger_old_pos + 1
                    
                    conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                               (new_challenged_pos, challenged_id))
                    
                    conn.execute('''
                        INSERT INTO ranking_history 
                        (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (challenged_id, challenged_old_pos, new_challenged_pos, 
                         challenged_old_tier, get_tier_from_position(new_challenged_pos), 
                         'wo_loss_demoted', challenge_id))
                    
                    print(f"✅ W.O. (desafiante já no topo): Desafiado {challenged_id} ({challenged_old_pos} → {new_challenged_pos})")
            
            # =====================================================
            # W.O. - DESAFIANTE NÃO COMPARECEU (wo_challenged)
            # Desafiado vence por WO
            # Desafiante perde 4 posições
            # =====================================================
            elif result_type == 'wo_challenged' and result == 'challenged_win':
                print(f"🔴 Processando W.O. - DESAFIANTE não compareceu")
                print(f"   Posição do desafiante antes: {challenger_old_pos}")
                
                # Calcular nova posição do desafiante (desce 4 posições)
                # Buscar quantos jogadores ativos existem do mesmo sexo
                max_pos_result = conn.execute('''
                    SELECT MAX(position) as max_pos FROM players 
                    WHERE active = 1 AND (sexo = ? OR (sexo IS NULL AND ? != 'feminino'))
                ''', (player_sexo, player_sexo)).fetchone()
                
                max_position = max_pos_result['max_pos'] if max_pos_result and max_pos_result['max_pos'] else challenger_old_pos
                
                # Nova posição = atual + 4, limitado ao máximo
                new_challenger_pos = min(challenger_old_pos + 4, max_position)
                
                if new_challenger_pos != challenger_old_pos:
                    # Puxar jogadores entre as posições para cima (ocupar o espaço deixado)
                    conn.execute('''
                        UPDATE players 
                        SET position = position - 1 
                        WHERE position > ? AND position <= ?
                        AND id != ?
                        AND active = 1
                        AND (sexo = ? OR (sexo IS NULL AND ? != 'feminino'))
                    ''', (challenger_old_pos, new_challenger_pos, challenger_id, player_sexo, player_sexo))
                    
                    # Atualizar posição do desafiante
                    conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                               (new_challenger_pos, challenger_id))
                    
                    # Registrar no histórico - Desafiante (penalizado)
                    conn.execute('''
                        INSERT INTO ranking_history 
                        (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (challenger_id, challenger_old_pos, new_challenger_pos, 
                         challenger_old_tier, get_tier_from_position(new_challenger_pos), 
                         'wo_penalty_4_positions', challenge_id))
                    
                    print(f"✅ W.O. Desafiante não compareceu:")
                    print(f"   Desafiante {challenger_id} PENALIZADO: {challenger_old_pos} → {new_challenger_pos}")
                    print(f"   Desafiado {challenged_id} não muda (posição {challenged_old_pos})")
                else:
                    print(f"ℹ️ W.O. Desafiante: Desafiante já está na última posição, sem mudança.")
                
                # IMPORTANTE: Desafiado NÃO muda de posição
            
            # =====================================================
            # RESULTADO NORMAL - DESAFIANTE VENCE
            # =====================================================
            elif result == 'challenger_win':
                new_challenger_pos = challenged_old_pos  # Desafiante vai para posição do desafiado
                new_challenged_pos = challenged_old_pos + 1  # Desafiado desce 1
                
                # Empurrar todos os jogadores entre as posições para baixo
                conn.execute('''
                    UPDATE players 
                    SET position = position + 1 
                    WHERE position >= ? AND position < ?
                    AND id != ? AND id != ?
                    AND active = 1
                    AND (sexo = ? OR (sexo IS NULL AND ? != 'feminino'))
                ''', (new_challenger_pos, challenger_old_pos, challenger_id, challenged_id, player_sexo, player_sexo))
                
                # Atualizar posição do desafiante (vencedor)
                conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                           (new_challenger_pos, challenger_id))
                
                # Atualizar posição do desafiado (perdedor)
                conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                           (new_challenged_pos, challenged_id))
                
                # Registrar no histórico - Desafiante
                conn.execute('''
                    INSERT INTO ranking_history 
                    (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (challenger_id, challenger_old_pos, new_challenger_pos, 
                     challenger_old_tier, get_tier_from_position(new_challenger_pos), 
                     'challenge_win', challenge_id))
                
                # Registrar no histórico - Desafiado
                conn.execute('''
                    INSERT INTO ranking_history 
                    (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                    VALUES (?, ?, ?, ?, ?, ?, ?)
                ''', (challenged_id, challenged_old_pos, new_challenged_pos, 
                     challenged_old_tier, get_tier_from_position(new_challenged_pos), 
                     'challenge_loss', challenge_id))
                
                print(f"✅ Desafiante venceu: {challenger_id} ({challenger_old_pos} → {new_challenger_pos}), "
                      f"Desafiado: {challenged_id} ({challenged_old_pos} → {new_challenged_pos})")
                
            # =====================================================
            # RESULTADO NORMAL - DESAFIADO VENCE
            # =====================================================
            elif result == 'challenged_win':
                # Verificar se existe alguém uma posição acima do desafiado
                player_above = conn.execute('''
                    SELECT id, position, tier FROM players 
                    WHERE position = ? AND active = 1
                    AND (sexo = ? OR (sexo IS NULL AND ? != 'feminino'))
                ''', (challenged_old_pos - 1, player_sexo, player_sexo)).fetchone()
                
                if player_above and challenged_old_pos > 1:
                    # Existe alguém acima - fazer a permuta
                    player_above_id = player_above['id']
                    player_above_old_pos = player_above['position']
                    player_above_old_tier = player_above['tier']
                    
                    new_challenged_pos = challenged_old_pos - 1  # Desafiado sobe 1
                    new_above_pos = challenged_old_pos  # Jogador acima desce 1
                    
                    # Atualizar posição do desafiado (vencedor - sobe)
                    conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                               (new_challenged_pos, challenged_id))
                    
                    # Atualizar posição do jogador que estava acima (desce)
                    conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                               (new_above_pos, player_above_id))
                    
                    # Registrar no histórico - Desafiado (vencedor)
                    conn.execute('''
                        INSERT INTO ranking_history 
                        (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (challenged_id, challenged_old_pos, new_challenged_pos, 
                         challenged_old_tier, get_tier_from_position(new_challenged_pos), 
                         'challenge_defense_win_promotion', challenge_id))
                    
                    # Registrar no histórico - Jogador que foi ultrapassado
                    conn.execute('''
                        INSERT INTO ranking_history 
                        (player_id, old_position, new_position, old_tier, new_tier, reason, challenge_id)
                        VALUES (?, ?, ?, ?, ?, ?, ?)
                    ''', (player_above_id, player_above_old_pos, new_above_pos, 
                         player_above_old_tier, get_tier_from_position(new_above_pos), 
                         'displaced_by_challenge_winner', challenge_id))
                    
                    print(f"✅ Desafiado venceu e subiu: {challenged_id} ({challenged_old_pos} → {new_challenged_pos}), "
                          f"Permuta com: {player_above_id} ({player_above_old_pos} → {new_above_pos}), "
                          f"Desafiante não muda: {challenger_id} (posição {challenger_old_pos})")
                else:
                    # Desafiado já está na posição 1 ou não há ninguém acima
                    print(f"ℹ️ Desafiado venceu mas já está na posição mais alta possível. "
                          f"Nenhuma mudança de posição.")
                
                # IMPORTANTE: Desafiante NÃO muda de posição quando perde
                
            else:
                print(f"Erro: Resultado inválido: {result}")
                conn.rollback()
                return
            
            # =====================================================
            # NORMALIZAÇÃO DO RANKING
            # =====================================================
            print("🔧 Executando normalização do ranking...")
            
            # Buscar jogadores do mesmo sexo ordenados pela posição atual
            players_to_normalize = conn.execute('''
                SELECT id, name, position, tier
                FROM players 
                WHERE active = 1 AND (sexo = ? OR (sexo IS NULL AND ? != 'feminino'))
                ORDER BY position, name
            ''', (player_sexo, player_sexo)).fetchall()
            
            # Reassignar posições sequenciais e recalcular tiers
            for i, player in enumerate(players_to_normalize, 1):
                new_position = i
                new_tier = get_tier_from_position(new_position)
                
                if player['position'] != new_position or player['tier'] != new_tier:
                    conn.execute('''
                        UPDATE players 
                        SET position = ?, tier = ? 
                        WHERE id = ?
                    ''', (new_position, new_tier, player['id']))
            
            print(f"✅ Ranking normalizado: {len(players_to_normalize)} jogadores")
            
            # Sincronizar as tabelas de histórico
            sync_ranking_history_tables(conn)
            
        except Exception as e:
            print(f"Erro ao processar resultado do desafio: {e}")
            conn.rollback()
            raise
    
    # Auto-corrigir ranking feminino se necessário
    auto_fix_female_ranking(conn)
    
    conn.commit()
    print("✅ Resultado do desafio processado com sucesso!")


def revert_challenge_result(conn, challenge_id):
    """
    Reverte as alterações feitas por um desafio no ranking.
    Restaura as posições anteriores dos jogadores, remove os registros de histórico
    e atualiza o histórico diário.
    """
    # Buscar registros de histórico para este desafio
    history_records = conn.execute('''
        SELECT * FROM ranking_history 
        WHERE challenge_id = ? 
        ORDER BY change_date DESC
    ''', (challenge_id,)).fetchall()
    
    # Para cada registro, restaurar a posição anterior
    for record in history_records:
        player_id = record['player_id']
        old_position = record['old_position']
        old_tier = record['old_tier']
        
        # Restaurar a posição e tier anteriores
        conn.execute('''
            UPDATE players 
            SET position = ?, tier = ? 
            WHERE id = ?
        ''', (old_position, old_tier, player_id))
    
    # Rebalancear todas as posições para garantir que não haja lacunas
    fix_position_gaps(conn)
    update_all_tiers(conn)
    
    # Remover os registros de histórico relacionados a este desafio
    conn.execute('DELETE FROM ranking_history WHERE challenge_id = ?', (challenge_id,))
    
    # Atualizar o desafio para remover o resultado
    conn.execute('UPDATE challenges SET result = NULL WHERE id = ?', (challenge_id,))
    
    # NOVA ADIÇÃO: Sincronizar as tabelas de histórico após reverter um desafio
    sync_ranking_history_tables(conn)
    
    conn.commit()
    print(f"Alterações do desafio ID {challenge_id} foram revertidas com sucesso.")

# Rota para registrar posições diárias manualmente
@app.route('/record_daily_rankings', methods=['GET', 'POST'])
def record_daily_rankings_route():
    if request.method == 'POST':
        # Verificação de admin (senha hardcoded removida)

        if not session.get('is_admin', False):

            flash('Acesso negado. Apenas administradores podem executar esta ação.', 'error')

            return redirect(url_for('dashboard'))
        
        result = record_daily_rankings()
        
        if result:
            flash('Posições registradas com sucesso no histórico diário!', 'success')
        else:
            flash('As posições de hoje já foram registradas anteriormente.', 'info')
        
        return redirect(url_for('index'))
    
    # Para método GET, mostrar o formulário
    return render_template('record_daily_rankings.html')

# Rota para visualizar o histórico de posições de um jogador
@app.route('/player/<int:player_id>/ranking_history')
def player_ranking_history(player_id):
    conn = get_db_connection()
    
    # Verificar se o jogador existe
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Obter o período desejado (padrão: últimos 30 dias)
    days = request.args.get('days', 30, type=int)
    
    # Calcular a data limite
    limit_date = (datetime.now() - timedelta(days=days)).date()
    
    # Buscar o histórico diário do jogador
    daily_history = conn.execute('''
        SELECT date_recorded, position, tier
        FROM daily_ranking_history
        WHERE player_id = ? AND date_recorded >= ?
        ORDER BY date_recorded
    ''', (player_id, limit_date.strftime('%Y-%m-%d'))).fetchall()
    
    # Buscar eventos específicos do ranking_history
    specific_changes = conn.execute('''
        SELECT date(change_date) as event_date, old_position, new_position, old_tier, new_tier, reason
        FROM ranking_history
        WHERE player_id = ? AND date(change_date) >= ?
        ORDER BY change_date
    ''', (player_id, limit_date.strftime('%Y-%m-%d'))).fetchall()
    
    # Combinar dados para visualização
    dates = []
    positions = []
    tiers = []
    events = []
    
    # Converter daily_history para um dicionário para fácil acesso
    daily_dict = {item['date_recorded']: {'position': item['position'], 'tier': item['tier']} for item in daily_history}
    
    # Converter specific_changes para um dicionário
    changes_dict = {}
    for change in specific_changes:
        if change['event_date'] not in changes_dict:
            changes_dict[change['event_date']] = []
        changes_dict[change['event_date']].append(change)
    
    # Criar série temporal contínua
    current_date = limit_date
    end_date = datetime.now().date()
    
    while current_date <= end_date:
        current_date_str = current_date.strftime('%Y-%m-%d')
        
        # Adicionar data
        dates.append(current_date_str)
        
        # Verificar se temos um registro diário para esta data
        if current_date_str in daily_dict:
            positions.append(daily_dict[current_date_str]['position'])
            tiers.append(daily_dict[current_date_str]['tier'])
        else:
            # Se não temos registro para esta data, usar valor anterior ou None
            if positions:
                positions.append(positions[-1])
                tiers.append(tiers[-1])
            else:
                positions.append(None)
                tiers.append(None)
        
        # Verificar se temos eventos específicos para esta data
        if current_date_str in changes_dict:
            # Usar o último evento do dia para esta posição
            latest_change = changes_dict[current_date_str][-1]
            positions[-1] = latest_change['new_position']
            tiers[-1] = latest_change['new_tier']
            events.append({
                'date': current_date_str,
                'reason': latest_change['reason'],
                'old_position': latest_change['old_position'],
                'new_position': latest_change['new_position']
            })
        
        current_date += timedelta(days=1)
    
    conn.close()
    
    return render_template('player_ranking_history.html', 
                          player=player, 
                          dates=dates, 
                          positions=positions,
                          tiers=tiers,
                          events=events,
                          days=days)


# API para obter dados filtrados para o gráfico
@app.route('/api/player/<int:player_id>/ranking_history')
def api_player_ranking_history(player_id):
    conn = get_db_connection()
    
    # Obter o período desejado
    days = request.args.get('days', 30, type=int)
    
    # Calcular a data limite
    limit_date = (datetime.now() - timedelta(days=days)).date()
    
    # Buscar o histórico do jogador
    history = conn.execute('''
        SELECT date_recorded, position, tier
        FROM daily_ranking_history
        WHERE player_id = ? AND date_recorded >= ?
        ORDER BY date_recorded
    ''', (player_id, limit_date.strftime('%Y-%m-%d'))).fetchall()
    
    # Prepara os dados para o gráfico
    dates = [item['date_recorded'] for item in history]
    positions = [item['position'] for item in history]
    tiers = [item['tier'] for item in history]
    
    conn.close()
    
    return jsonify({
        'dates': dates,
        'positions': positions,
        'tiers': tiers
    })

@app.route('/deactivate_player/<int:player_id>', methods=['GET', 'POST'])
def deactivate_player(player_id):
    """
    Inativa um jogador e oferece opções para reorganizar ou não o ranking.
    GET: Mostra formulário de confirmação
    POST: Processa a inativação
    """
    conn = get_db_connection()
    
    # Buscar jogador
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Para requisição GET, mostrar tela de confirmação
    if request.method == 'GET':
        conn.close()
        return render_template('deactivate_player.html', player=player)
    
    # Para requisição POST, processar a inativação
    senha = request.form.get('senha', '')
    
    if not session.get('is_admin', False):
        conn.close()
        flash('Senha incorreta! Operação não autorizada.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    # ✅ CORREÇÃO: Obter o valor de rerank do formulário
    rerank = request.form.get('rerank', 'no') == 'yes'
    
    try:
        current_position = player['position']
        current_tier = player['tier']
        player_sexo = player['sexo'] if player['sexo'] else 'masculino'
        
        # Se rerank=True, inativa e reorganiza ranking
        if rerank:
            # 1. Marcar o jogador como inativo - CORRIGIDO: não definir position/tier como NULL
            conn.execute('''
                UPDATE players
                SET active = 0, 
                    notes = ?
                WHERE id = ?
            ''', (f"Inativado em {datetime.now().strftime('%d/%m/%Y')}. Posição anterior: {current_position} (Tier {current_tier})", 
                  player_id))
            
            # 2. Registrar a inativação no histórico
            conn.execute('''
                INSERT INTO ranking_history 
                (player_id, old_position, new_position, old_tier, new_tier, reason)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (player_id, current_position, current_position, current_tier, current_tier, 'player_inactivated'))
            
            # 3. Ajustar posições de todos os jogadores abaixo
            conn.execute('''
                UPDATE players
                SET position = position - 1
                WHERE position > ? AND active = 1
            ''', (current_position,))
            
            flash_message = 'Jogador inativado com sucesso e ranking reorganizado.'
        else:
            # Apenas inativa o jogador sem reorganizar
            conn.execute('''
                UPDATE players
                SET active = 0, 
                    notes = ?
                WHERE id = ?
            ''', (f"Inativado em {datetime.now().strftime('%d/%m/%Y')}. Mantida posição: {current_position} (Tier {current_tier})", 
                  player_id))
            
            # Registrar a inativação no histórico sem ajuste de posição
            conn.execute('''
                INSERT INTO ranking_history 
                (player_id, old_position, new_position, old_tier, new_tier, reason)
                VALUES (?, ?, ?, ?, ?, ?)
            ''', (player_id, current_position, current_position, current_tier, current_tier, 'player_inactivated_nochange'))
            
            flash_message = 'Jogador inativado com sucesso. Posição no ranking mantida (jogador ficará invisível nas visualizações).'
        
        # 5. Cancelar quaisquer desafios pendentes
        conn.execute('''
            UPDATE challenges
            SET status = 'cancelled', result = 'player_inactive'
            WHERE (challenger_id = ? OR challenged_id = ?) AND status IN ('pending', 'accepted')
        ''', (player_id, player_id))
        
        # 6. Atualizar tiers após a reorganização de posições
        update_all_tiers(conn)
        
        conn.commit()
        
        # ✨ NOVA ADIÇÃO: Auto-corrigir ranking feminino se uma jogadora foi desativada
        if player_sexo == 'feminino':
            auto_fix_female_ranking(conn)
            conn.commit()
        
        flash(flash_message, 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao processar operação: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('index'))


@app.route('/reactivate_player/<int:player_id>', methods=['GET', 'POST'])
def reactivate_player(player_id):
    """
    Reativa um jogador inativo, colocando-o na última posição do ranking
    """
    conn = get_db_connection()
    
    # Buscar jogador
    player = conn.execute('SELECT * FROM players WHERE id = ? AND active = 0', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado ou já está ativo!', 'error')
        return redirect(url_for('index'))
    
    # Para requisição GET, mostrar tela de confirmação
    if request.method == 'GET':
        conn.close()
        return render_template('reactivate_player.html', player=player)
    
    # Para requisição POST, processar a reativação
    # Verificação de admin (senha hardcoded removida)
    if not session.get('is_admin', False):
        conn.close()
        flash('Acesso negado. Apenas administradores podem executar esta ação! Operação não autorizada.', 'error')
        return redirect(url_for('index'))
    
    try:
        player_sexo = player['sexo'] if player['sexo'] else 'masculino'
        
        # Determinar a última posição do ranking baseada no sexo
        if player_sexo == 'feminino':
            # Para mulheres: buscar última posição feminina
            last_pos = conn.execute('SELECT MAX(position) as max_pos FROM players WHERE active = 1 AND sexo = "feminino"').fetchone()
        else:
            # Para homens: buscar última posição masculina
            last_pos = conn.execute('SELECT MAX(position) as max_pos FROM players WHERE active = 1 AND (sexo != "feminino" OR sexo IS NULL)').fetchone()
        
        new_position = 1 if not last_pos['max_pos'] else last_pos['max_pos'] + 1
        new_tier = get_tier_from_position(new_position)
        
        # Reativar jogador na última posição do ranking
        conn.execute('''
            UPDATE players
            SET active = 1, 
                position = ?,
                tier = ?,
                notes = ?
            WHERE id = ?
        ''', (new_position, new_tier, 
              f"{player['notes'] or ''} | Reativado em {datetime.now().strftime('%d/%m/%Y')}",
              player_id))
        
        # Registrar a reativação no histórico
        conn.execute('''
            INSERT INTO ranking_history 
            (player_id, old_position, new_position, old_tier, new_tier, reason)
            VALUES (?, ?, ?, ?, ?, ?)
        ''', (player_id, player['position'], new_position, player['tier'], new_tier, 'player_reactivated'))
        
        conn.commit()
        
        # ✨ NOVA ADIÇÃO: Auto-corrigir ranking feminino se uma jogadora foi reativada
        if player_sexo == 'feminino':
            auto_fix_female_ranking(conn)
            conn.commit()
        
        flash(f'Jogador reativado com sucesso na posição {new_position}.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao reativar jogador: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))


@app.route('/delete_player/<int:player_id>', methods=['GET', 'POST'])
def delete_player(player_id):
    """
    Exclui permanentemente um jogador do sistema.
    O jogador deve estar inativo para ser excluído.
    GET: Mostra formulário de confirmação
    POST: Processa a exclusão
    """
    conn = get_db_connection()
    
    # Buscar jogador
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Verificar se o jogador está inativo (requisito para exclusão)
    if player['active'] == 1:
        conn.close()
        flash('O jogador deve estar inativo antes de ser excluído. Por favor, inative o jogador primeiro.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    # Para requisição GET, mostrar tela de confirmação
    if request.method == 'GET':
        # Verificar se existem desafios associados ao jogador
        challenges_count = conn.execute('''
            SELECT COUNT(*) AS count FROM challenges 
            WHERE challenger_id = ? OR challenged_id = ?
        ''', (player_id, player_id)).fetchone()['count']
        
        # Verificar se existem registros de histórico
        history_count = conn.execute('''
            SELECT COUNT(*) AS count FROM ranking_history 
            WHERE player_id = ?
        ''', (player_id,)).fetchone()['count']
        
        daily_history_count = conn.execute('''
            SELECT COUNT(*) AS count FROM daily_ranking_history 
            WHERE player_id = ?
        ''', (player_id,)).fetchone()['count']
        
        conn.close()
        
        return render_template('delete_player.html', 
                              player=player, 
                              challenges_count=challenges_count,
                              history_count=history_count + daily_history_count)
    
    # Para requisição POST, processar a exclusão
    senha = request.form.get('senha', '')
    
    if not session.get('is_admin', False):
        conn.close()
        flash('Senha incorreta! Operação não autorizada.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    if not confirm_delete:
        conn.close()
        flash('Você precisa confirmar a exclusão marcando a caixa de confirmação.', 'error')
        return redirect(url_for('delete_player', player_id=player_id))
    
    try:
        # 1. Excluir registros relacionados na tabela daily_ranking_history
        conn.execute('DELETE FROM daily_ranking_history WHERE player_id = ?', (player_id,))
        
        # 2. Excluir registros relacionados na tabela ranking_history
        conn.execute('DELETE FROM ranking_history WHERE player_id = ?', (player_id,))
        
        # 3. Excluir ou atualizar desafios relacionados
        # Como os desafios possuem foreign keys, podemos definir a estratégia:
        # Opção 1: Excluir todos os desafios relacionados
        conn.execute('''
            DELETE FROM challenges 
            WHERE challenger_id = ? OR challenged_id = ?
        ''', (player_id, player_id))
        
        # 4. Finalmente, excluir o jogador
        conn.execute('DELETE FROM players WHERE id = ?', (player_id,))
        
        conn.commit()
        flash(f'Jogador "{player["name"]}" foi excluído permanentemente.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao excluir jogador: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('index'))


@app.route('/update_player_name/<int:player_id>', methods=['POST'])
def update_player_name(player_id):
    """
    Atualiza o nome de um jogador
    """
    conn = get_db_connection()
    
    # Verificar se o jogador existe
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Verificar senha
    # Verificação de admin (senha hardcoded removida)
    if not session.get('is_admin', False):
        conn.close()
        flash('Acesso negado. Apenas administradores podem executar esta ação! Operação não autorizada.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    # Obter novo nome
    new_name = request.form.get('new_name', '').strip()
    old_name = player['name']
    
    # Validar novo nome
    if not new_name:
        conn.close()
        flash('O nome não pode estar vazio!', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    # Se o nome não mudou, não fazer nada
    if new_name == old_name:
        conn.close()
        flash('Nenhuma alteração foi realizada.', 'info')
        return redirect(url_for('player_detail', player_id=player_id))
    
    try:
        # Atualizar o nome do jogador
        conn.execute('UPDATE players SET name = ? WHERE id = ?', (new_name, player_id))
        
        # Opcional: Registrar alteração no histórico
        notes = f"Nome alterado de '{old_name}' para '{new_name}' em {datetime.now().strftime('%d/%m/%Y')}"
        
        # Se o jogador já tem notas, adicionar à frente
        if player['notes']:
            notes = f"{player['notes']} | {notes}"
        
        # Atualizar as notas
        conn.execute('UPDATE players SET notes = ? WHERE id = ?', (notes, player_id))
        
        conn.commit()
        flash(f'Nome atualizado com sucesso de "{old_name}" para "{new_name}".', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar o nome: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))

@app.route('/update_player_country/<int:player_id>', methods=['POST'])
def update_player_country(player_id):
    """
    Atualiza o país do jogador
    """
    conn = get_db_connection()
    
    try:
        # Verificar se o jogador existe
        player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
        
        if not player:
            conn.close()
            flash('Jogador não encontrado!', 'error')
            return redirect(url_for('index'))
        
        # Verificar se é o próprio usuário editando seu perfil
        is_own_profile = False
        user_id = session.get('user_id')
        
        # Verificar se é um admin (o ID pode ser no formato 'admin_1')
        if isinstance(user_id, str) and user_id.startswith('admin_'):
            is_own_profile = False
        elif isinstance(user_id, int):
            is_own_profile = user_id == player_id
        elif isinstance(user_id, str) and user_id.isdigit():
            is_own_profile = int(user_id) == player_id
        
        # Verificar senha apenas para administradores
        if not is_own_profile:
            # Verificação de admin (senha hardcoded removida)
            if not session.get('is_admin', False):
                conn.close()
                flash('Acesso negado. Apenas administradores podem executar esta ação.', 'error')
                return redirect(url_for('dashboard'))
        
        # Obter novo país
        new_country = request.form.get('new_country', '').strip()
        
        # Verificar se a coluna 'country' existe no objeto player
        try:
            old_country = player['country']
        except (KeyError, TypeError):
            # Se a coluna não existe, considerar como None
            old_country = None
        
        # Se o país não mudou, não fazer nada
        if new_country == old_country:
            conn.close()
            flash('Nenhuma alteração foi realizada.', 'info')
            return redirect(url_for('player_detail', player_id=player_id))
        
        # Verificar se a coluna 'country' existe na tabela players
        columns_info = conn.execute('PRAGMA table_info(players)').fetchall()
        column_names = [col[1] for col in columns_info]
        
        if 'country' not in column_names:
            # Se a coluna não existe, criar ela
            conn.execute('ALTER TABLE players ADD COLUMN country TEXT DEFAULT NULL')
            conn.commit()
            print("Coluna 'country' adicionada à tabela players.")
        
        # Atualizar o país do jogador
        conn.execute('UPDATE players SET country = ? WHERE id = ?', (new_country, player_id))
        
        # Opcional: Registrar alteração nas notas
        old_country_display = old_country if old_country else 'não informado'
        new_country_display = new_country if new_country else 'não informado'
        
        notes = f"País alterado de '{old_country_display}' para '{new_country_display}' em {datetime.now().strftime('%d/%m/%Y')}"
        
        # Se o jogador já tem notas, adicionar à frente
        if player['notes']:
            notes = f"{player['notes']} | {notes}"
        
        # Atualizar as notas
        conn.execute('UPDATE players SET notes = ? WHERE id = ?', (notes, player_id))
        
        conn.commit()
        flash(f'País atualizado com sucesso para "{new_country_display}"', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar o país: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))


@app.route('/update_player_sexo/<int:player_id>', methods=['POST'])
def update_player_sexo(player_id):
    """
    Atualiza o sexo de um jogador
    """
    conn = get_db_connection()
    
    # Verificar se o jogador existe
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Verificar se é o próprio usuário editando seu perfil
    is_own_profile = False
    user_id = session.get('user_id')
    
    # Verificar se é um admin (o ID pode ser no formato 'admin_1')
    if isinstance(user_id, str) and user_id.startswith('admin_'):
        is_own_profile = False
    elif isinstance(user_id, int):
        is_own_profile = user_id == player_id
    elif isinstance(user_id, str) and user_id.isdigit():
        is_own_profile = int(user_id) == player_id
    
    # Verificar senha apenas para administradores
    if not is_own_profile:
        # Verificação de admin (senha hardcoded removida)
        if not session.get('is_admin', False):
            conn.close()
            flash('Acesso negado. Apenas administradores podem executar esta ação.', 'error')
            return redirect(url_for('dashboard'))
    
    # Obter novo sexo
    new_sexo = request.form.get('new_sexo', '').strip()
    
    # Validar valor
    if new_sexo not in ['masculino', 'feminino']:
        conn.close()
        flash('Valor inválido para sexo.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    try:
        # Atualizar o sexo do jogador
        conn.execute('UPDATE players SET sexo = ? WHERE id = ?', (new_sexo, player_id))
        
        # Opcional: Registrar alteração nas notas
        notes = f"Sexo alterado para '{new_sexo}' em {datetime.now().strftime('%d/%m/%Y')}"
        
        # Se o jogador já tem notas, adicionar à frente
        if player['notes']:
            notes = f"{player['notes']} | {notes}"
        
        # Atualizar as notas
        conn.execute('UPDATE players SET notes = ? WHERE id = ?', (notes, player_id))
        
        conn.commit()
        flash('Sexo atualizado com sucesso!', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar o sexo: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))


# ============================================
# ROTA INDEX COMPLETA - Substitua no app.py
# ============================================

@app.route('/')
@login_required
def index():
    conn = get_db_connection()
    
    # Jogadores masculinos ativos (exclui VIPs com position = 0)
    male_players = conn.execute('''
        SELECT * FROM players 
        WHERE active = 1 
        AND (tipo_membro = 'jogador' OR tipo_membro IS NULL OR tipo_membro = '')
        AND (sexo = 'masculino' OR sexo IS NULL OR sexo = '')
        AND position > 0
        ORDER BY position
    ''').fetchall()
    
    # Jogadoras femininas ativas (exclui VIPs com position = 0)
    female_players = conn.execute('''
        SELECT * FROM players 
        WHERE active = 1 
        AND (tipo_membro = 'jogador' OR tipo_membro IS NULL OR tipo_membro = '')
        AND sexo = 'feminino'
        AND position > 0
        ORDER BY position
    ''').fetchall()
    
    # ============================================
    # IMPORTANTE: Query para Membros VIP
    # ============================================
    vip_members = conn.execute('''
        SELECT * FROM players 
        WHERE active = 1 
        AND tipo_membro = 'vip'
        ORDER BY name
    ''').fetchall()
    
    # Jogadores inativos (todos os tipos)
    inactive_players = conn.execute('''
        SELECT * FROM players 
        WHERE active = 0
        ORDER BY name
    ''').fetchall()
    
    conn.close()
    
    # ============================================
    # IMPORTANTE: Passar vip_members para o template
    # ============================================
    return render_template('index.html', 
                          male_players=male_players,
                          female_players=female_players,
                          vip_members=vip_members,
                          inactive_players=inactive_players)


@app.route('/pyramid')
@login_required 
def pyramid_redirect():
    """Redireciona a rota antiga para a nova rota da pirâmide"""
    return redirect(url_for('pyramid_dynamic'))

@app.route('/pyramid_dynamic')
@login_required 
def pyramid_dynamic():
    conn = get_db_connection()
    
    # Buscar jogadores ativos
    players = conn.execute('SELECT * FROM players WHERE active = 1 ORDER BY position').fetchall()
    
    # Buscar jogadores com desafios
    challenges = conn.execute('''
        SELECT DISTINCT c.challenger_id, c.challenged_id, c.status, c.scheduled_date,
               p1.position as challenger_position, p2.position as challenged_position
        FROM challenges c
        JOIN players p1 ON c.challenger_id = p1.id
        JOIN players p2 ON c.challenged_id = p2.id
        WHERE c.status IN ('pending', 'accepted', 'completed_pending', 'awaiting_date_confirmation')
    ''').fetchall()
    
    # Converter listas para conjuntos para busca eficiente
    players_with_challenges = set()
    players_with_completed_pending = {}
    
    # Mapear desafios por jogador
    player_challenges = {}
    for player in players:
        player_id = player['id']
        player_challenges[player_id] = {
            'challenging_positions': [],
            'challenged_by_positions': []
        }
    
    # Processar desafios
    for challenge in challenges:
        challenger_id = challenge['challenger_id']
        challenged_id = challenge['challenged_id']
        
        # Marcar jogadores envolvidos em desafios
        if challenge['status'] == 'completed_pending':
            players_with_completed_pending[challenger_id] = 'completed_pending'
            players_with_completed_pending[challenged_id] = 'completed_pending'
        else:
            players_with_challenges.add(challenger_id)
            players_with_challenges.add(challenged_id)
        
        # Registrar posições envolvidas nos desafios
        if challenger_id in player_challenges:
            player_challenges[challenger_id]['challenging_positions'].append(challenge['challenged_position'])
        if challenged_id in player_challenges:
            player_challenges[challenged_id]['challenged_by_positions'].append(challenge['challenger_position'])
    
    # Organizar jogadores por tier
    tiers = {}
    for player in players:
        if player['tier'] not in tiers:
            tiers[player['tier']] = []
        
        # Adicionar informações sobre desafios
        player_dict = dict(player)
        player_dict['has_pending_challenge'] = player['id'] in players_with_challenges
        player_dict['challenge_status'] = players_with_completed_pending.get(player['id'], None)
        
        # Adicionar informações sobre as posições envolvidas nos desafios
        player_dict['challenging_positions'] = player_challenges[player['id']]['challenging_positions']
        player_dict['challenged_by_positions'] = player_challenges[player['id']]['challenged_by_positions']
        
        tiers[player['tier']].append(player_dict)
    
    # Ordenar tiers alfabeticamente
    sorted_tiers = sorted(tiers.items())
    
    # Buscar todos os desafios aceitos e pendentes para mostrar datas ou indicadores
    challenges_for_display = conn.execute('''
        SELECT id, challenger_id, challenged_id, status, scheduled_date
        FROM challenges
        WHERE status IN ('accepted', 'pending')
    ''').fetchall()
    
    conn.close()
    return render_template('pyramid_dynamic.html', 
                          tiers=sorted_tiers, 
                          challenges=challenges_for_display)

@app.route('/pyramid_print')
@login_required 
def pyramid_print():
    """Página de impressão da pirâmide em alta resolução"""
    conn = get_db_connection()
    
    # Buscar jogadores ativos
    players = conn.execute('SELECT * FROM players WHERE active = 1 ORDER BY position').fetchall()
    
    # Organizar jogadores por tier
    tiers = {}
    for player in players:
        if player['tier'] not in tiers:
            tiers[player['tier']] = []
        tiers[player['tier']].append(dict(player))
    
    # Ordenar tiers alfabeticamente
    sorted_tiers = sorted(tiers.items())
    
    conn.close()
    return render_template('pyramid_print.html', tiers=sorted_tiers)


# Rota original (mantida para compatibilidade ou redirecionamento)
# Altere estas rotas no seu arquivo app.py:

@app.route('/challenges')
@login_required 
def challenges():
    """Redireciona para a página de lista de desafios (nova interface principal)"""
    return redirect(url_for('challenges_list'))

# Rota para o calendário de desafios
@app.route('/challenges/calendar')
@login_required 
def challenges_calendar():
    conn = get_db_connection()
    # Obter todos os desafios com nomes dos jogadores
    challenges = conn.execute('''
        SELECT c.*, 
               p1.name as challenger_name, p1.id as challenger_id,
               p2.name as challenged_name, p2.id as challenged_id,
               p1.position as challenger_position,
               p2.position as challenged_position,
               p1.tier as challenger_tier,
               p2.tier as challenged_tier
        FROM challenges c
        JOIN players p1 ON c.challenger_id = p1.id
        JOIN players p2 ON c.challenged_id = p2.id
        ORDER BY c.created_at DESC
    ''').fetchall()
    conn.close()
    return render_template('calendar_challenges.html', challenges=challenges)

# Rota para a lista de desafios (agora a padrão)
@app.route('/challenges/list')
@login_required 
def challenges_list():
    conn = get_db_connection()
    
    # Preparar a consulta base
    query = '''
        SELECT c.*, 
               p1.name as challenger_name, p1.id as challenger_id,
               p2.name as challenged_name, p2.id as challenged_id,
               p1.position as challenger_position,
               p2.position as challenged_position,
               p1.tier as challenger_tier,
               p2.tier as challenged_tier
        FROM challenges c
        JOIN players p1 ON c.challenger_id = p1.id
        JOIN players p2 ON c.challenged_id = p2.id
    '''
    
    # Parâmetros para filtros
    params = []
    where_clauses = []
    
    # Aplicar filtros se fornecidos
    status = request.args.get('status')
    if status:
        where_clauses.append('c.status = ?')
        params.append(status)
    
    player = request.args.get('player')
    if player:
        where_clauses.append('(p1.name LIKE ? OR p2.name LIKE ?)')
        params.append(f'%{player}%')
        params.append(f'%{player}%')
    
    date_from = request.args.get('date_from')
    if date_from:
        where_clauses.append('c.scheduled_date >= ?')
        params.append(date_from)
    
    date_to = request.args.get('date_to')
    if date_to:
        where_clauses.append('c.scheduled_date <= ?')
        params.append(date_to)
    
    # Adicionar cláusulas WHERE se houver filtros
    if where_clauses:
        query += ' WHERE ' + ' AND '.join(where_clauses)
    
    # Adicionar ordenação
    query += ' ORDER BY c.created_at DESC'
    
    # Executar a consulta
    challenges = conn.execute(query, params).fetchall()
    conn.close()
    
    return render_template('challenges_list.html', challenges=challenges)

# 1. Modificação na rota new_challenge para validar a data do desafio (máximo 7 dias)
# Substitua a rota new_challenge existente por esta versão modificada

# ============================================================
# ROTA NEW_CHALLENGE - COMPLETA E CORRIGIDA
# ============================================================
# Prazos:
# - response_deadline (aceitar/rejeitar): 2 DIAS
# - scheduled_date (data do jogo): máximo 7 DIAS
# ============================================================

# ============================================================
# ROTA NEW_CHALLENGE - COMPLETA
# ============================================================
# PRAZOS:
# - response_deadline (aceitar/rejeitar): 2 DIAS
# - scheduled_date (data do jogo): máximo 7 DIAS
# ============================================================

@app.route('/new_challenge', methods=['GET', 'POST'])
@login_required
def new_challenge():
    conn = get_db_connection()
    
    # Verificar se os desafios estão bloqueados
    setting = conn.execute('SELECT value FROM system_settings WHERE key = ?', ('challenges_locked',)).fetchone()
    challenges_locked = setting and setting['value'] == 'true'
    
    is_admin = session.get('is_admin', False)
    is_main_admin = is_admin and session.get('username') == 'admin'
    
    if challenges_locked and not is_admin:
        conn.close()
        flash('A criação de desafios está temporariamente bloqueada pelo administrador.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        challenger_id = request.form['challenger_id']
        challenged_id = request.form['challenged_id']
        scheduled_date = request.form['scheduled_date']
        
        # ============================================================
        # VALIDAÇÃO DA DATA DO JOGO: máximo 7 dias
        # ============================================================
        try:
            scheduled_date_obj = datetime.strptime(scheduled_date, '%Y-%m-%d').date()
            today_date = datetime.now().date()
            max_date = today_date + timedelta(days=7)
            
            if scheduled_date_obj > max_date:
                conn.close()
                flash(f'A data do desafio não pode ser superior a 7 dias a partir de hoje. Data máxima permitida: {max_date.strftime("%d/%m/%Y")}', 'error')
                return redirect(url_for('new_challenge', challenger_id=challenger_id))
            
            if scheduled_date_obj < today_date:
                conn.close()
                flash('A data do desafio não pode ser anterior à data atual.', 'error')
                return redirect(url_for('new_challenge', challenger_id=challenger_id))
        except ValueError:
            conn.close()
            flash('Formato de data inválido.', 'error')
            return redirect(url_for('new_challenge', challenger_id=challenger_id))
        
        # Verificar se ambos jogadores estão ativos
        challenger = conn.execute('SELECT * FROM players WHERE id = ? AND active = 1', (challenger_id,)).fetchone()
        challenged = conn.execute('SELECT * FROM players WHERE id = ? AND active = 1', (challenged_id,)).fetchone()
        
        if not challenger or not challenged:
            conn.close()
            flash('Um dos jogadores está inativo e não pode participar de desafios.', 'error')
            return redirect(url_for('new_challenge'))
        
        # Verificação de bloqueio
        if challenged['bloqueado'] == 1:
            motivo = challenged['bloqueio_motivo'] or 'indisponível'
            conn.close()
            flash(f'❌ {challenged["name"]} está temporariamente bloqueado ({motivo}).', 'error')
            return redirect(url_for('new_challenge', challenger_id=challenger_id))
        
        if challenger['bloqueado'] == 1:
            motivo = challenger['bloqueio_motivo'] or 'indisponível'
            conn.close()
            flash(f'❌ Você está temporariamente bloqueado ({motivo}) e não pode criar desafios.', 'error')
            return redirect(url_for('dashboard'))
        
        # Regras de validação
        error = None
        
        # Verificar desafios pendentes
        pending_challenges = conn.execute('''
            SELECT * FROM challenges 
            WHERE (challenger_id = ? OR challenged_id = ? OR challenger_id = ? OR challenged_id = ?)
            AND status IN ('pending', 'accepted')
        ''', (challenger_id, challenger_id, challenged_id, challenged_id)).fetchall()
        
        if pending_challenges:
            same_players_challenge = False
            for challenge in pending_challenges:
                if ((challenge['challenger_id'] == int(challenger_id) and challenge['challenged_id'] == int(challenged_id)) or
                    (challenge['challenger_id'] == int(challenged_id) and challenge['challenged_id'] == int(challenger_id))):
                    same_players_challenge = True
                    break
            
            if same_players_challenge:
                error = "Já existe um desafio pendente ou aceito entre estes jogadores."
            else:
                error = "Um dos jogadores já está envolvido em um desafio pendente ou aceito."
        
        # Regra de 8 posições
        if not error and not is_main_admin:
            challenger_position = challenger['position']
            challenged_position = challenged['position']
            challenger_sexo = challenger['sexo'] or 'masculino'
            
            if challenged_position >= challenger_position:
                error = "Você só pode desafiar jogadores em posições melhores que a sua."
            else:
                eligible_above = conn.execute('''
                    SELECT id, position FROM players 
                    WHERE active = 1 
                    AND position < ?
                    AND position > 0
                    AND (bloqueado = 0 OR bloqueado IS NULL)
                    AND (sexo = ? OR (sexo IS NULL AND ? = 'masculino'))
                    ORDER BY position DESC
                    LIMIT 8
                ''', (challenger_position, challenger_sexo, challenger_sexo)).fetchall()
                
                eligible_ids = [p['id'] for p in eligible_above]
                
                if int(challenged_id) not in eligible_ids:
                    error = "Você só pode desafiar os 8 jogadores elegíveis mais próximos acima da sua posição."
        
        if error:
            conn.close()
            flash(error, 'error')
            return redirect(url_for('new_challenge'))
        
        # ============================================================
        # CRIAR O DESAFIO
        # ============================================================
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        # ============================================================
        # PRAZO PARA RESPONDER: 2 DIAS
        # ============================================================
        response_deadline = (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S')
        
        conn.execute('''
            INSERT INTO challenges (
                challenger_id, 
                challenged_id, 
                status, 
                scheduled_date, 
                created_at, 
                response_deadline,
                challenger_position_at_creation,
                challenged_position_at_creation
            )
            VALUES (?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            challenger_id, 
            challenged_id, 
            'pending', 
            scheduled_date, 
            current_datetime, 
            response_deadline,
            challenger['position'],
            challenged['position']
        ))
        
        challenge_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        
        # Registrar log
        try:
            table_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='challenge_logs'").fetchone()
            
            if not table_exists:
                conn.execute('''
                    CREATE TABLE challenge_logs (
                        id INTEGER PRIMARY KEY AUTOINCREMENT,
                        challenge_id INTEGER,
                        user_id TEXT NOT NULL,
                        modified_by TEXT NOT NULL,
                        old_status TEXT,
                        new_status TEXT,
                        old_result TEXT,
                        new_result TEXT,
                        notes TEXT,
                        created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                    )
                ''')
            
            creator_type = "Admin Principal" if is_main_admin else "Admin" if is_admin else "Jogador"
            notes = f"Desafio criado. Jogo marcado para {scheduled_date}. Prazo para resposta: 2 dias."
            
            conn.execute('''
                INSERT INTO challenge_logs 
                (challenge_id, user_id, modified_by, old_status, new_status, old_result, new_result, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ''', (challenge_id, session.get('user_id', 'unknown'), creator_type, None, 'pending', None, None, notes, current_datetime))
            
        except Exception as e:
            print(f"Erro ao registrar log: {e}")
        
        conn.commit()

        # Notificação WhatsApp
        try:
            notificar_desafio_criado_whatsapp(challenge_id)
        except Exception as e:
            print(f"[WhatsApp] Erro ao notificar: {e}")

        conn.close()
        
        # ============================================================
        # MENSAGEM DE SUCESSO
        # ============================================================
        if is_main_admin:
            flash('Desafio criado com sucesso! O desafiado terá 2 dias para responder.', 'success')
        else:
            flash('Desafio criado com sucesso! O desafiado terá 2 dias para aceitar, rejeitar ou propor nova data.', 'success')
        
        return redirect(url_for('challenges_calendar'))
    
    # ============================================================
    # GET - Mostrar formulário
    # ============================================================
    preselected_challenger_id = None
    all_players = []
    eligible_challenged = []
    
    pending_challenges = conn.execute('''
        SELECT challenger_id, challenged_id 
        FROM challenges 
        WHERE status IN ('pending', 'accepted')
    ''').fetchall()
    
    players_with_challenges = set()
    for challenge in pending_challenges:
        players_with_challenges.add(challenge['challenger_id'])
        players_with_challenges.add(challenge['challenged_id'])
    
    blocked_players = conn.execute('SELECT id FROM players WHERE bloqueado = 1').fetchall()
    blocked_player_ids = {p['id'] for p in blocked_players}
    
    if is_main_admin:
        all_players = conn.execute('SELECT * FROM players WHERE active = 1 AND position > 0 ORDER BY position').fetchall()
        preselected_challenger_id = request.args.get('challenger_id')
        
        if preselected_challenger_id:
            try:
                preselected_challenger_id = int(preselected_challenger_id)
                challenger = conn.execute('SELECT * FROM players WHERE id = ? AND active = 1', 
                                         (preselected_challenger_id,)).fetchone()
                
                if challenger:
                    eligible_challenged = conn.execute('''
                        SELECT * FROM players 
                        WHERE active = 1 AND id != ? AND position > 0
                        ORDER BY position
                    ''', (preselected_challenger_id,)).fetchall()
                    
                    if preselected_challenger_id in players_with_challenges:
                        flash('Este jogador já está envolvido em um desafio pendente ou aceito.', 'warning')
                    
                    if preselected_challenger_id in blocked_player_ids:
                        flash('Este jogador está bloqueado e não pode criar desafios.', 'warning')
                    
                    eligible_challenged = [player for player in eligible_challenged 
                                          if player['id'] not in players_with_challenges
                                          and player['id'] not in blocked_player_ids]
            except (ValueError, TypeError):
                preselected_challenger_id = None
                
    elif is_admin:
        all_players = conn.execute('SELECT * FROM players WHERE active = 1 AND position > 0 ORDER BY position').fetchall()
        preselected_challenger_id = request.args.get('challenger_id')
        
        if preselected_challenger_id:
            try:
                preselected_challenger_id = int(preselected_challenger_id)
                challenger = conn.execute('SELECT * FROM players WHERE id = ? AND active = 1', 
                                         (preselected_challenger_id,)).fetchone()
                
                if challenger:
                    challenger_sexo = challenger['sexo'] or 'masculino'
                    eligible_challenged = conn.execute('''
                        SELECT * FROM players 
                        WHERE active = 1
                        AND position < ?
                        AND position > 0
                        AND (bloqueado = 0 OR bloqueado IS NULL)
                        AND (sexo = ? OR (sexo IS NULL AND ? = 'masculino'))
                        AND id != ?
                        ORDER BY position DESC
                        LIMIT 8
                    ''', (challenger['position'], challenger_sexo, challenger_sexo, preselected_challenger_id)).fetchall()
                    
                    if preselected_challenger_id in players_with_challenges:
                        flash('Este jogador já está envolvido em um desafio pendente ou aceito.', 'warning')
                    
                    if preselected_challenger_id in blocked_player_ids:
                        flash('Este jogador está bloqueado e não pode criar desafios.', 'warning')
                    
                    eligible_challenged = [player for player in eligible_challenged 
                                          if player['id'] not in players_with_challenges]
            except (ValueError, TypeError):
                preselected_challenger_id = None
    else:
        if 'user_id' in session and not is_admin:
            preselected_challenger_id = session['user_id']
        else:
            temp_id = request.args.get('challenger_id')
            if temp_id:
                try:
                    preselected_challenger_id = int(temp_id)
                except (ValueError, TypeError):
                    preselected_challenger_id = None
        
        if preselected_challenger_id:
            if preselected_challenger_id in blocked_player_ids:
                conn.close()
                flash('Você está temporariamente bloqueado e não pode criar desafios.', 'warning')
                return redirect(url_for('dashboard'))
            
            if preselected_challenger_id in players_with_challenges:
                conn.close()
                flash('Este jogador já está envolvido em um desafio pendente ou aceito.', 'warning')
                return redirect(url_for('challenges_calendar'))
            
            challenger = conn.execute('SELECT * FROM players WHERE id = ? AND active = 1', 
                                     (preselected_challenger_id,)).fetchone()
            
            if challenger:
                all_players = conn.execute('SELECT * FROM players WHERE active = 1 AND position > 0 ORDER BY position').fetchall()
                challenger_sexo = challenger['sexo'] or 'masculino'
                
                eligible_challenged = conn.execute('''
                    SELECT * FROM players 
                    WHERE active = 1
                    AND position < ?
                    AND position > 0
                    AND (bloqueado = 0 OR bloqueado IS NULL)
                    AND (sexo = ? OR (sexo IS NULL AND ? = 'masculino'))
                    AND id != ?
                    ORDER BY position DESC
                    LIMIT 8
                ''', (challenger['position'], challenger_sexo, challenger_sexo, preselected_challenger_id)).fetchall()
                
                eligible_challenged = [player for player in eligible_challenged 
                                      if player['id'] not in players_with_challenges]
        else:
            all_players = conn.execute('SELECT * FROM players WHERE active = 1 AND position > 0 ORDER BY position').fetchall()
            all_players = [player for player in all_players 
                          if player['id'] not in players_with_challenges
                          and player['id'] not in blocked_player_ids]
    
    today_date = datetime.now().strftime('%Y-%m-%d')
    
    challenger_info = None
    if preselected_challenger_id:
        challenger_info = conn.execute('SELECT * FROM players WHERE id = ? AND active = 1', 
                                      (preselected_challenger_id,)).fetchone()
    
    conn.close()
    
    return render_template('new_challenge.html', 
                          all_players=all_players, 
                          eligible_challenged=eligible_challenged,
                          preselected_challenger=preselected_challenger_id,
                          challenger_info=challenger_info,
                          today_date=today_date,
                          is_admin=is_admin,
                          is_main_admin=is_main_admin,
                          challenges_locked=challenges_locked)


@app.route('/admin/toggle_challenges', methods=['GET', 'POST'])
@login_required
def toggle_challenges():
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        action = request.form.get('action')
        # Verificação de admin (senha hardcoded removida)
        if not session.get('is_admin', False):
            conn.close()
            flash('Acesso negado. Apenas administradores podem executar esta ação! Operação não autorizada.', 'error')
            return redirect(url_for('toggle_challenges'))
        
        if action == 'lock':
            conn.execute('UPDATE system_settings SET value = ?, updated_at = CURRENT_TIMESTAMP WHERE key = ?', 
                       ('true', 'challenges_locked'))
            conn.commit()
            flash('Criação de desafios BLOQUEADA com sucesso!', 'success')
        elif action == 'unlock':
            conn.execute('UPDATE system_settings SET value = ?, updated_at = CURRENT_TIMESTAMP WHERE key = ?', 
                       ('false', 'challenges_locked'))
            conn.commit()
            flash('Criação de desafios LIBERADA com sucesso!', 'success')
    
    # Obter status atual
    setting = conn.execute('SELECT value, updated_at FROM system_settings WHERE key = ?', ('challenges_locked',)).fetchone()
    is_locked = setting and setting['value'] == 'true'
    updated_at = setting['updated_at'] if setting else None
    
    conn.close()
    
    return render_template('toggle_challenges.html', is_locked=is_locked, updated_at=updated_at)



# Alteração na rota delete_challenge
@app.route('/delete_challenge/<int:challenge_id>', methods=['POST'])
def delete_challenge(challenge_id):
    conn = get_db_connection()
    
    # Verificar se o desafio existe
    challenge = conn.execute('SELECT * FROM challenges WHERE id = ?', (challenge_id,)).fetchone()
    
    if not challenge:
        conn.close()
        flash('Desafio não encontrado!', 'error')
        return redirect(url_for('challenges_calendar'))
    
    # Verificar se é um admin
    is_admin = session.get('is_admin', False)
    if not is_admin:
        conn.close()
        flash('Apenas administradores podem excluir desafios.', 'error')
        return redirect(url_for('challenge_detail', challenge_id=challenge_id))
    
    # Obter o motivo da exclusão
    admin_delete_reason = request.form.get('admin_delete_reason', '')
    
    # Registrar a ação de exclusão em um log
    try:
        # Verificar se a tabela de logs existe
        table_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='challenge_logs'").fetchone()
        
        if not table_exists:
            # Criar a tabela se não existir
            conn.execute('''
                CREATE TABLE challenge_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    challenge_id INTEGER,
                    user_id TEXT NOT NULL,
                    modified_by TEXT NOT NULL,
                    old_status TEXT,
                    new_status TEXT,
                    old_result TEXT,
                    new_result TEXT,
                    notes TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
        
        # Inserir o log de exclusão
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        
        conn.execute('''
            INSERT INTO challenge_logs 
            (challenge_id, user_id, modified_by, old_status, new_status, old_result, new_result, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            challenge_id, 
            session.get('user_id', 'unknown'),
            "Admin",
            challenge['status'],
            "DELETED",
            challenge['result'],
            None,
            f"Desafio excluído. Motivo: {admin_delete_reason}",
            current_datetime
        ))
        
    except Exception as e:
        print(f"Erro ao registrar log de exclusão: {e}")
        # Continuar mesmo se o log falhar
    
    # Verificar se o desafio já afetou o ranking
    if challenge['status'] == 'completed' and challenge['result']:
        # Buscar histórico relacionado a este desafio
        history = conn.execute('SELECT * FROM ranking_history WHERE challenge_id = ?', (challenge_id,)).fetchall()
        
        if history:
            # Reverter as alterações no ranking
            try:
                revert_challenge_result(conn, challenge_id)
                flash('Alterações no ranking foram revertidas.', 'info')
            except Exception as e:
                conn.rollback()
                conn.close()
                flash(f'Erro ao reverter alterações no ranking: {str(e)}', 'error')
                return redirect(url_for('challenge_detail', challenge_id=challenge_id))
    
    # Excluir o desafio
    conn.execute('DELETE FROM challenges WHERE id = ?', (challenge_id,))
    conn.commit()
    conn.close()
    
    flash('Desafio excluído com sucesso!', 'success')
    return redirect(url_for('challenges_calendar'))

# Alteração na rota edit_challenge
@app.route('/edit_challenge/<int:challenge_id>', methods=['GET', 'POST'])
def edit_challenge(challenge_id):
    conn = get_db_connection()
    
    # Obter o desafio
    challenge = conn.execute('''
        SELECT c.*, 
               p1.name as challenger_name, p1.position as challenger_position,
               p2.name as challenged_name, p2.position as challenged_position
        FROM challenges c
        JOIN players p1 ON c.challenger_id = p1.id
        JOIN players p2 ON c.challenged_id = p2.id
        WHERE c.id = ?
    ''', (challenge_id,)).fetchone()
    
    if not challenge:
        conn.close()
        flash('Desafio não encontrado!', 'error')
        return redirect(url_for('challenges_calendar'))
    
    # Verificar se o desafio já afetou o ranking
    ranking_affected = False
    if challenge['status'] == 'completed' and challenge['result']:
        # Buscar histórico relacionado a este desafio
        history = conn.execute('SELECT * FROM ranking_history WHERE challenge_id = ?', (challenge_id,)).fetchone()
        if history:
            ranking_affected = True
    
    if request.method == 'POST':
        # Se o desafio está concluído (normal ou com pendência), verificar a senha
        if challenge['status'] == 'completed' or challenge['status'] == 'completed_pending':
            # Verificação de admin (senha hardcoded removida)
            if not session.get('is_admin', False):
                conn.close()
                flash('Acesso negado. Apenas administradores podem executar esta ação! Desafios concluídos só podem ser editados com a senha correta.', 'error')
                return redirect(url_for('challenge_detail', challenge_id=challenge_id))
        
        scheduled_date = request.form['scheduled_date']
        status = request.form.get('status', challenge['status'])
        result = request.form.get('result', challenge['result'])
        
        # Se estamos alterando um desafio que já afetou o ranking
        if ranking_affected and (status != 'completed' or result != challenge['result']):
            try:
                # Reverter as alterações no ranking
                revert_challenge_result(conn, challenge_id)
                flash('Alterações no ranking foram revertidas.', 'info')
                
                # Se o novo status for completed, processar o novo resultado
                if status == 'completed' and result:
                    process_challenge_result(conn, challenge_id, status, result)
                    flash('Ranking atualizado com o novo resultado.', 'success')
                # Se o novo status for completed_pending, processar sem alterar o ranking
                elif status == 'completed_pending' and result:
                    process_challenge_result(conn, challenge_id, status, result)
                    flash('Desafio marcado como Concluído (com pendência). O ranking não foi alterado.', 'success')
                else:
                    # Apenas atualizar o status e resultado
                    conn.execute('UPDATE challenges SET status = ?, result = ? WHERE id = ?', 
                               (status, result, challenge_id))
                    conn.commit()
            except Exception as e:
                conn.rollback()
                flash(f'Erro ao reverter alterações: {str(e)}', 'error')
                conn.close()
                return redirect(url_for('challenge_detail', challenge_id=challenge_id))
        else:
            # Atualizar o desafio normalmente
            conn.execute('''
                UPDATE challenges 
                SET scheduled_date = ?, status = ?, result = ?
                WHERE id = ?
            ''', (scheduled_date, status, result, challenge_id))
            conn.commit()
        
        conn.close()
        flash('Desafio atualizado com sucesso!', 'success')
        return redirect(url_for('challenge_detail', challenge_id=challenge_id))
    
    conn.close()
    return render_template('edit_challenge.html', challenge=challenge, ranking_affected=ranking_affected)

# Alteração na rota update_challenge (opcional, caso a atualização de status também deva ter restrição)
@app.route('/update_challenge/<int:challenge_id>', methods=['POST'])
def update_challenge(challenge_id):
    status = request.form['status']
    result = request.form.get('result', None)
    result_type = request.form.get('result_type', 'normal')  # NOVO: tipo de resultado
    admin_notes = request.form.get('admin_notes', '')
    is_admin_action = request.form.get('modified_by_admin') == 'true'
    
    conn = get_db_connection()
    
    # Verificar se o desafio existe
    challenge = conn.execute('SELECT * FROM challenges WHERE id = ?', (challenge_id,)).fetchone()
    
    if not challenge:
        conn.close()
        flash('Desafio não encontrado!', 'error')
        return redirect(url_for('challenges_calendar'))
    
    # Verificar permissões
    is_admin = session.get('is_admin', False)
    is_challenger = challenge['challenger_id'] == session.get('user_id')
    is_challenged = challenge['challenged_id'] == session.get('user_id')
    
    # Registrar quem fez a alteração para fins de auditoria
    modified_by = "Admin" if is_admin else "Desafiante" if is_challenger else "Desafiado" if is_challenged else "Desconhecido"
    current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
    
    # Verificar se a alteração é permitida
    if not (is_admin or is_challenger or is_challenged):
        conn.close()
        flash('Você não tem permissão para modificar este desafio.', 'error')
        return redirect(url_for('challenge_detail', challenge_id=challenge_id))
    
    # Verificações específicas por status
    if status == 'accepted' and not (is_admin or is_challenged):
        conn.close()
        flash('Apenas o desafiado ou um administrador pode aceitar um desafio.', 'error')
        return redirect(url_for('challenge_detail', challenge_id=challenge_id))
    
    # Para qualquer mudança de status para 'completed', apenas admin ou participantes podem fazer
    if status == 'completed':
        if not (is_admin or is_challenger or is_challenged):
            conn.close()
            flash('Apenas participantes do desafio ou administradores podem marcar um desafio como concluído.', 'error')
            return redirect(url_for('challenge_detail', challenge_id=challenge_id))
    
    # Criar log da alteração
    log_message = f"Status alterado de '{challenge['status']}' para '{status}'"
    if result:
        log_message += f", resultado: '{result}'"
    if result_type and result_type != 'normal':
        log_message += f", tipo: '{result_type}'"  # NOVO: registrar tipo no log
    if admin_notes:
        log_message += f", observações: '{admin_notes}'"
    
    # Armazenar log em uma tabela de histórico de alterações
    try:
        # Verificar se a tabela challenge_logs existe
        table_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='challenge_logs'").fetchone()
        
        if not table_exists:
            # Criar a tabela se não existir
            conn.execute('''
                CREATE TABLE challenge_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    challenge_id INTEGER NOT NULL,
                    user_id TEXT NOT NULL,
                    modified_by TEXT NOT NULL,
                    old_status TEXT,
                    new_status TEXT,
                    old_result TEXT,
                    new_result TEXT,
                    notes TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
                    FOREIGN KEY (challenge_id) REFERENCES challenges(id)
                )
            ''')
            print("Tabela challenge_logs criada com sucesso.")
        
        # Inserir o log
        conn.execute('''
            INSERT INTO challenge_logs 
            (challenge_id, user_id, modified_by, old_status, new_status, old_result, new_result, notes, created_at)
            VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
        ''', (
            challenge_id, 
            session.get('user_id', 'unknown'),
            modified_by,
            challenge['status'],
            status,
            challenge['result'],
            result,
            f"{admin_notes} [Tipo: {result_type}]" if result_type != 'normal' else admin_notes,
            current_datetime
        ))
        
    except Exception as e:
        print(f"Erro ao registrar log: {e}")
        # Continuar mesmo se o log falhar
    
    # Processar o desafio conforme o status
    if status == 'completed' and result:
        # NOVO: Atualizar também o result_type
        conn.execute('UPDATE challenges SET result_type = ? WHERE id = ?', (result_type, challenge_id))
        
        # Processar o resultado do desafio (alterando a pirâmide)
        process_challenge_result(conn, challenge_id, status, result)
        
        # Mensagem diferenciada para WO
        if result_type in ['wo_challenger', 'wo_challenged']:
            flash('Status do desafio atualizado para Concluído (WO) e ranking atualizado.', 'success')
        else:
            flash('Status do desafio atualizado para Concluído e ranking atualizado.', 'success')
    else:
        # Apenas atualizar o status
        conn.execute('UPDATE challenges SET status = ? WHERE id = ?', (status, challenge_id))
        conn.commit()
        flash('Status do desafio atualizado com sucesso!', 'success')
    
    conn.close()
    
    return redirect(url_for('challenge_detail', challenge_id=challenge_id))


# ============================================================
# PASSO 4: Criar filtro Jinja2 para exibir tipo de resultado
# ============================================================
# Adicione este filtro no app.py (perto dos outros filtros)

@app.template_filter('result_type_label')
def result_type_label_filter(result_type):
    """Converte o tipo de resultado para texto legível"""
    labels = {
        'normal': '',
        'wo_challenger': '(WO)',
        'wo_challenged': '(WO)',
        None: ''
    }
    return labels.get(result_type, '')


@app.template_filter('result_type_description')
def result_type_description_filter(result_type):
    """Descrição completa do tipo de resultado"""
    descriptions = {
        'normal': 'Jogo disputado normalmente',
        'wo_challenger': 'Vitória por WO - Desafiado não compareceu',
        'wo_challenged': 'Vitória por WO - Desafiante não compareceu',
        None: ''
    }
    return descriptions.get(result_type, '')




@app.route('/history')
@login_required 
def history():
    conn = get_db_connection()
    history = conn.execute('''
        SELECT rh.*, p.name as player_name
        FROM ranking_history rh
        JOIN players p ON rh.player_id = p.id
        ORDER BY rh.change_date DESC
    ''').fetchall()
    conn.close()
    return render_template('history.html', history=history)



# ============================================
# ROTA PLAYER_DETAIL COMPLETA
# ============================================
# Substitua no app.py - inclui queries de desafios
# ============================================

# ============================================
# ROTA PLAYER_DETAIL CORRIGIDA
# ============================================
# sqlite3.Row usa colchetes, não .get()
# ============================================

# ============================================================
# ROTA PLAYER_DETAIL - COMPLETA COM DESAFIOS
# ============================================================
# Substitua a rota player_detail no seu app.py por esta versão

@app.route('/jogador/<int:player_id>')
def player_detail(player_id):
    conn = get_db_connection()
    
    # Buscar jogador
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado.', 'danger')
        return redirect(url_for('index'))
    
    # Verificar se é o próprio perfil
    is_own_profile = False
    if 'user_id' in session and session['user_id'] == player_id:
        is_own_profile = True
    
    # Verificar se é admin
    is_admin = session.get('is_admin', False)
    
    # ============================================================
    # DESAFIOS COMO DESAFIANTE (challenger)
    # ============================================================
    challenges_as_challenger = conn.execute('''
        SELECT 
            c.*,
            p2.name as opponent_name,
            p2.position as opponent_position
        FROM challenges c
        JOIN players p2 ON c.challenged_id = p2.id
        WHERE c.challenger_id = ?
        ORDER BY 
            CASE c.status 
                WHEN 'pending' THEN 1 
                WHEN 'accepted' THEN 2 
                ELSE 3 
            END,
            c.scheduled_date DESC
    ''', (player_id,)).fetchall()
    
    # ============================================================
    # DESAFIOS COMO DESAFIADO (challenged)
    # ============================================================
    challenges_as_challenged = conn.execute('''
        SELECT 
            c.*,
            p1.name as opponent_name,
            p1.position as opponent_position
        FROM challenges c
        JOIN players p1 ON c.challenger_id = p1.id
        WHERE c.challenged_id = ?
        ORDER BY 
            CASE c.status 
                WHEN 'pending' THEN 1 
                WHEN 'accepted' THEN 2 
                ELSE 3 
            END,
            c.scheduled_date DESC
    ''', (player_id,)).fetchall()
    
    # ============================================================
    # JOGADORES DISPONÍVEIS PARA DESAFIAR (potential_challenges)
    # ============================================================
    potential_challenges = []
    
    # Verificar tipo_membro (sqlite3.Row não tem .get())
    try:
        tipo_membro = player['tipo_membro']
    except:
        tipo_membro = 'jogador'
    
    # Só busca se o jogador estiver ativo e não for VIP
    if player['active'] == 1 and tipo_membro != 'vip':
        player_position = player['position']
        player_sexo = player['sexo']
        
        # Buscar jogadores até 8 posições acima, mesmo sexo, ativos, não bloqueados
        potential_challenges = conn.execute('''
            SELECT id, name, position, tier
            FROM players
            WHERE active = 1
            AND sexo = ?
            AND position < ?
            AND position >= ?
            AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
            AND position > 0
            AND (bloqueado = 0 OR bloqueado IS NULL)
            AND id != ?
            ORDER BY position DESC
        ''', (player_sexo, player_position, max(1, player_position - 8), player_id)).fetchall()
    
    conn.close()
    
    return render_template('player_detail.html',
                          player=player,
                          is_own_profile=is_own_profile,
                          is_admin=is_admin,
                          challenges_as_challenger=challenges_as_challenger,
                          challenges_as_challenged=challenges_as_challenged,
                          potential_challenges=potential_challenges)


@app.route('/challenge_detail/<int:challenge_id>')
def challenge_detail(challenge_id):
    conn = get_db_connection()
    challenge = conn.execute('''
        SELECT c.*, 
               p1.name as challenger_name, p1.id as challenger_id, p1.position as challenger_position, p1.hcp_index as challenger_hcp,
               p2.name as challenged_name, p2.id as challenged_id, p2.position as challenged_position, p2.hcp_index as challenged_hcp
        FROM challenges c
        JOIN players p1 ON c.challenger_id = p1.id
        JOIN players p2 ON c.challenged_id = p2.id
        WHERE c.id = ?
    ''', (challenge_id,)).fetchone()
    
    # NOVA ADIÇÃO: Verificar se jogadores podem submeter resultados
    setting = conn.execute('SELECT value FROM system_settings WHERE key = ?', 
                          ('players_can_submit_results',)).fetchone()
    players_can_submit = setting and setting['value'] == 'true'
    
    conn.close()
    
    if not challenge:
        flash('Desafio não encontrado!', 'error')
        return redirect(url_for('challenges_calendar'))
    
    # Cálculo de dias restantes para resposta
    days_remaining = None
    expired = False
    
    if challenge['status'] == 'pending' and challenge['response_deadline']:
        try:
            deadline_obj = datetime.strptime(challenge['response_deadline'], '%Y-%m-%d %H:%M:%S')
            deadline_date = deadline_obj.date()
            today_date = datetime.now().date()
            delta = (deadline_date - today_date).days
            days_remaining = delta
            expired = days_remaining < 0
            if expired:
                days_remaining = abs(days_remaining)
        except Exception as e:
            print(f"Erro ao calcular dias restantes: {str(e)}")
            days_remaining = None
    
    return render_template('challenge_detail.html', 
                          challenge=challenge, 
                          days_remaining=days_remaining,
                          expired=expired,
                          players_can_submit=players_can_submit)  # NOVA VARIÁVEL




def add_result_type_column():
    """Adiciona coluna result_type na tabela challenges para registrar WO"""
    conn = get_db_connection()
    
    # Verificar se a coluna já existe
    columns_info = conn.execute('PRAGMA table_info(challenges)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    if 'result_type' not in column_names:
        conn.execute('ALTER TABLE challenges ADD COLUMN result_type TEXT DEFAULT "normal"')
        conn.commit()
        print("Coluna 'result_type' adicionada à tabela challenges.")
    else:
        print("Coluna 'result_type' já existe na tabela challenges.")
    
    conn.close()




# Rota aprimorada para verificar e corrigir completamente a estrutura da pirâmide
@app.route('/fix_pyramid', methods=['GET'])
def fix_pyramid():
    conn = get_db_connection()
    
    try:
        # 1. Corrigir ranking masculino
        male_players = conn.execute('SELECT id, position FROM players WHERE active = 1 AND (sexo != "feminino" OR sexo IS NULL) ORDER BY position').fetchall()
        fix_position_gaps(conn)
        
        # 2. NOVA ADIÇÃO: Corrigir ranking feminino
        female_players = conn.execute('''
            SELECT id, name, position FROM players 
            WHERE active = 1 AND sexo = 'feminino'
            ORDER BY position, name
        ''').fetchall()
        
        # Reorganizar posições das mulheres sequencialmente (1, 2, 3, 4...)
        for i, player in enumerate(female_players, 1):
            new_position = i
            new_tier = get_tier_from_position(new_position)
            
            conn.execute('''
                UPDATE players 
                SET position = ?, tier = ? 
                WHERE id = ? AND sexo = 'feminino'
            ''', (new_position, new_tier, player['id']))
        
        # 3. Atualizar todos os tiers
        update_all_tiers(conn)
        
        # 4. Verificação final
        incorrect_players = verify_pyramid_structure(conn)
        
        if len(male_players) > 0 or len(female_players) > 0:
            flash(f'Estrutura da pirâmide corrigida: {len(male_players)} homens e {len(female_players)} mulheres reorganizados.', 'success')
        else:
            flash('A estrutura da pirâmide já estava correta!', 'info')
        
        if incorrect_players:
            update_all_tiers(conn)
            flash(f'Tiers corrigidos automaticamente para {len(incorrect_players)} jogadores.', 'info')
        
        conn.commit()
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao corrigir a pirâmide: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('pyramid_dynamic'))

# Função para verificação da integridade da pirâmide (sem o scheduler)
def check_pyramid_integrity():
    """
    Executa uma verificação manual da integridade da pirâmide.
    Pode ser chamada a partir de rotas específicas quando necessário.
    """
    print("Executando verificação manual da integridade da pirâmide...")
    conn = get_db_connection()
    try:
        # Verificar lacunas nas posições
        players = conn.execute('SELECT id, position FROM players WHERE active = 1 AND position IS NOT NULL ORDER BY position').fetchall()
        expected_position = 1
        positions_fixed = 0
        
        for player in players:
            if player['position'] != expected_position:
                conn.execute('UPDATE players SET position = ? WHERE id = ?', 
                           (expected_position, player['id']))
                positions_fixed += 1
            expected_position += 1
        
        # Verificar tiers incorretos
        incorrect_players = verify_pyramid_structure(conn)
        
        if incorrect_players or positions_fixed > 0:
            # Corrigir tiers se necessário
            update_all_tiers(conn)
            print(f"Correção: {positions_fixed} posições e {len(incorrect_players)} tiers ajustados.")
            conn.commit()
        else:
            print("Verificação: Estrutura da pirâmide está correta.")
    
    except Exception as e:
        print(f"Erro na verificação da pirâmide: {e}")
        conn.rollback()
    finally:
        conn.close()


@app.route('/fix_only_ladies', methods=['GET'])
def fix_only_ladies():
    """
    Corrige APENAS o ranking feminino, sem tocar no masculino
    """
    conn = get_db_connection()
    try:
        # Buscar APENAS jogadoras femininas, ordenadas por posição atual
        female_players = conn.execute('''
            SELECT id, name, position FROM players 
            WHERE active = 1 AND sexo = 'feminino'
            ORDER BY position
        ''').fetchall()
        
        if not female_players:
            flash('Nenhuma jogadora feminina encontrada.', 'info')
            conn.close()
            return redirect(url_for('index'))
        
        print(f"Corrigindo {len(female_players)} jogadoras...")
        
        # Corrigir APENAS as jogadoras para posições 1, 2, 3, 4...
        for i, player in enumerate(female_players, 1):
            new_position = i
            new_tier = get_tier_from_position(new_position)
            
            print(f"Jogadora {player['name']}: posição {player['position']} → {new_position}, tier {new_tier}")
            
            conn.execute('''
                UPDATE players 
                SET position = ?, tier = ? 
                WHERE id = ? AND sexo = 'feminino'
            ''', (new_position, new_tier, player['id']))
        
        conn.commit()
        flash(f'✅ Ranking feminino corrigido! {len(female_players)} jogadoras: posições 1, 2, 3, 4...', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Erro: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('index'))


# Rota adicional para verificação manual da integridade da pirâmide
@app.route('/check_pyramid')
def check_pyramid_route():
    """Rota para executar a verificação da pirâmide sob demanda."""
    check_pyramid_integrity()
    flash('Verificação da integridade da pirâmide concluída.', 'info')
    return redirect(url_for('pyramid_dynamic'))

# Rota para verificar o status de um jogador (útil para diagnóstico)
@app.route('/check_player/<int:player_id>')
def check_player(player_id):
    conn = get_db_connection()
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    conn.close()
    
    if not player:
        return f"Jogador ID {player_id} não encontrado"
    
    return f"Jogador: {player['name']}, Active: {player['active']}, Position: {player['position']}, Notes: {player['notes']}"

# Rota para adicionar colunas (útil para atualização do banco de dados)
@app.route('/add_columns')
def add_columns():
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar se as colunas existem
    columns_info = cursor.execute('PRAGMA table_info(players)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    changes = []
    
    if 'active' not in column_names:
        cursor.execute('ALTER TABLE players ADD COLUMN active INTEGER DEFAULT 1')
        changes.append("Coluna 'active' adicionada")
    
    if 'notes' not in column_names:
        cursor.execute('ALTER TABLE players ADD COLUMN notes TEXT')
        changes.append("Coluna 'notes' adicionada")
    
    conn.commit()
    conn.close()
    
    if changes:
        return f"Alterações realizadas: {', '.join(changes)}"
    else:
        return "Nenhuma alteração necessária."

# Função para gerar automaticamente um novo player_code
def generate_player_code(conn):
    """
    Gera um novo código de jogador único no formato 'LOG' + número sequencial de 3 dígitos.
    
    Args:
        conn: Conexão com o banco de dados
    
    Returns:
        str: Novo código de jogador no formato 'LOG001', 'LOG002', etc.
    """
    # Buscar todos os códigos que seguem o padrão LOG + números
    result = conn.execute('''
        SELECT player_code FROM players 
        WHERE player_code GLOB 'LOG[0-9]*'
        ORDER BY LENGTH(player_code) DESC, player_code DESC
        LIMIT 1
    ''').fetchone()
    
    if result and result['player_code']:
        current_code = result['player_code']
        try:
            # Extrair apenas a parte numérica após 'LOG'
            numeric_part = current_code[3:]  # Remove 'LOG'
            current_number = int(numeric_part)
            new_number = current_number + 1
        except (ValueError, IndexError):
            # Se houver erro, começar do 1
            new_number = 1
    else:
        # Se não existir nenhum código, começar do 1
        new_number = 1
    
    # Gerar novo código com 3 dígitos
    new_code = f"LOG{new_number:03d}"
    
    # Verificar se o código já existe (segurança contra duplicatas)
    existing = conn.execute(
        'SELECT COUNT(*) as count FROM players WHERE player_code = ?', 
        (new_code,)
    ).fetchone()
    
    if existing and existing['count'] > 0:
        # Se já existe, buscar próximo disponível
        while True:
            new_number += 1
            test_code = f"LOG{new_number:03d}"
            check = conn.execute(
                'SELECT COUNT(*) as count FROM players WHERE player_code = ?', 
                (test_code,)
            ).fetchone()
            if not check or check['count'] == 0:
                return test_code
    
    return new_code


@app.route('/add_player', methods=['GET', 'POST'])
def add_player():
    if request.method == 'POST':
        name = request.form.get('name', '').strip()
        sexo = request.form.get('sexo', 'masculino').strip()
        tipo_membro = request.form.get('tipo_membro', 'jogador').strip()
        hcp_index = request.form.get('hcp_index', '').strip()
        email = request.form.get('email', '').strip()
        country = request.form.get('country', 'Brasil').strip()
        notes = request.form.get('notes', '').strip()
        
        if not name:
            flash('Nome é obrigatório!', 'error')
            return redirect(url_for('add_player'))
        
        if not session.get('is_admin', False):
            flash('Apenas administradores podem adicionar membros.', 'error')
            return redirect(url_for('add_player'))
        
        conn = get_db_connection()
        try:
            # Verificar se já existe
            existing = conn.execute('SELECT * FROM players WHERE name = ?', (name,)).fetchone()
            if existing:
                flash(f'Jogador "{name}" já existe!', 'error')
                conn.close()
                return redirect(url_for('add_player'))
            
            player_code = generate_player_code(conn)
            
            # MEMBRO VIP: não tem posição no ranking
            if tipo_membro == 'vip':
                new_position = None
                new_tier = None
            else:
                # Jogador normal: última posição do ranking
                if sexo == 'feminino':
                    last_pos_result = conn.execute(
                        '''SELECT MAX(position) as max_pos FROM players 
                           WHERE active = 1 AND sexo = 'feminino' 
                           AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)'''
                    ).fetchone()
                else:
                    last_pos_result = conn.execute(
                        '''SELECT MAX(position) as max_pos FROM players 
                           WHERE active = 1 AND (sexo != 'feminino' OR sexo IS NULL)
                           AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)'''
                    ).fetchone()
                
                new_position = (last_pos_result['max_pos'] or 0) + 1
                new_tier = get_tier_from_position(new_position)
            
            # Converter HCP
            hcp_value = None
            if hcp_index:
                try:
                    hcp_value = float(hcp_index.replace(',', '.'))
                except ValueError:
                    pass
            
            # Inserir
            conn.execute('''
                INSERT INTO players (name, sexo, tipo_membro, position, tier, player_code, 
                                     hcp_index, email, country, notes, active)
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, 1)
            ''', (name, sexo, tipo_membro, new_position, new_tier, player_code,
                  hcp_value, email, country, notes))
            
            player_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
            
            conn.commit()
            
            if tipo_membro == 'vip':
                flash(f'⭐ Membro VIP "{name}" criado com sucesso! Código: {player_code}', 'success')
            else:
                flash(f'🏌️ Jogador "{name}" adicionado na posição {new_position}! Código: {player_code}', 'success')
            
            return redirect(url_for('player_detail', player_id=player_id))
            
        except Exception as e:
            conn.rollback()
            flash(f'Erro: {str(e)}', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('index'))
    
    return render_template('add_player.html')


@app.route('/update_player_contact/<int:player_id>', methods=['POST'])
def update_player_contact(player_id):
    """
    Atualiza o contato (email/telefone) de um jogador
    """
    conn = get_db_connection()
    
    # Verificar se o jogador existe
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Verificar senha
    # Verificação de admin (senha hardcoded removida)
    if not session.get('is_admin', False):
        conn.close()
        flash('Acesso negado. Apenas administradores podem executar esta ação! Operação não autorizada.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    # Obter novo contato
    new_contact = request.form.get('new_contact', '').strip()
    old_contact = player['email']
    
    # Se o contato não mudou, não fazer nada
    if new_contact == old_contact:
        conn.close()
        flash('Nenhuma alteração foi realizada.', 'info')
        return redirect(url_for('player_detail', player_id=player_id))
    
    try:
        # Atualizar o contato do jogador
        conn.execute('UPDATE players SET email = ? WHERE id = ?', (new_contact, player_id))
        
        # Opcional: Registrar alteração nas notas
        notes = f"Contato alterado de '{old_contact or 'não informado'}' para '{new_contact or 'não informado'}' em {datetime.now().strftime('%d/%m/%Y')}"
        
        # Se o jogador já tem notas, adicionar à frente
        if player['notes']:
            notes = f"{player['notes']} | {notes}"
        
        # Atualizar as notas
        conn.execute('UPDATE players SET notes = ? WHERE id = ?', (notes, player_id))
        
        conn.commit()
        flash(f'Contato atualizado com sucesso.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar o contato: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))

@app.route('/update_player_hcp/<int:player_id>', methods=['POST'])
def update_player_hcp(player_id):
    """
    Atualiza o HCP Index de um jogador e calcula automaticamente o HCP para diferentes tees
    """
    conn = get_db_connection()
    
    # Verificar se o jogador existe
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Verificar senha
    # Verificação de admin (senha hardcoded removida)
    if not session.get('is_admin', False):
        conn.close()
        flash('Acesso negado. Apenas administradores podem executar esta ação! Operação não autorizada.', 'error')
        return redirect(url_for('player_detail', player_id=player_id))
    
    # Obter novo HCP
    new_hcp = request.form.get('new_hcp', '').strip()
    old_hcp = str(player['hcp_index']) if player['hcp_index'] is not None else ''
    
    # Se o HCP não mudou, não fazer nada
    if new_hcp == old_hcp:
        conn.close()
        flash('Nenhuma alteração foi realizada.', 'info')
        return redirect(url_for('player_detail', player_id=player_id))
    
    try:
        # Converter o novo HCP para float se não estiver vazio
        hcp_value = None
        if new_hcp:
            try:
                hcp_value = float(new_hcp.replace(',', '.'))
            except ValueError:
                conn.close()
                flash('Valor de HCP inválido. Use apenas números.', 'error')
                return redirect(url_for('player_detail', player_id=player_id))
        
        # Atualizar o HCP do jogador com a data atual
        conn.execute('UPDATE players SET hcp_index = ?, hcp_last_update = CURRENT_TIMESTAMP WHERE id = ?', (hcp_value, player_id))
        
        # Calcular e atualizar o HCP OGC Tee Branco se o HCP Index foi fornecido
        if hcp_value is not None:
            # Função para determinar o HCP OGC Tee Branco com base no HCP Index
            def get_hcp_ogc_white(hcp):
                # Para handicaps "plus" (valores negativos no banco)
                if hcp <= -0.3:
                    if -5.0 <= hcp <= -4.8: return '+7'
                    elif -4.7 <= hcp <= -3.9: return '+6'
                    elif -3.8 <= hcp <= -3.1: return '+5'
                    elif -3.0 <= hcp <= -2.2: return '+4'
                    elif -2.1 <= hcp <= -1.3: return '+3'
                    elif -1.2 <= hcp <= -0.4: return '+2'
                    elif -0.3 <= hcp <= 0.5: return '+1'
                
                # Para handicaps regulares (valores positivos no banco)
                if 0.6 <= hcp <= 1.4: return '0'
                elif 1.5 <= hcp <= 2.2: return '1'
                elif 2.3 <= hcp <= 3.1: return '2'
                elif 3.2 <= hcp <= 4.0: return '3'
                elif 4.1 <= hcp <= 4.9: return '4'
                elif 5.0 <= hcp <= 5.8: return '5'
                elif 5.9 <= hcp <= 6.7: return '6'
                elif 6.8 <= hcp <= 7.5: return '7'
                elif 7.6 <= hcp <= 8.4: return '8'
                elif 8.5 <= hcp <= 9.3: return '9'
                elif 9.4 <= hcp <= 10.2: return '10'
                elif 10.3 <= hcp <= 11.1: return '11'
                elif 11.2 <= hcp <= 12.0: return '12'
                elif 12.1 <= hcp <= 12.8: return '13'
                elif 12.9 <= hcp <= 13.7: return '14'
                elif 13.8 <= hcp <= 14.6: return '15'
                elif 14.7 <= hcp <= 15.5: return '16'
                elif 15.6 <= hcp <= 16.4: return '17'
                elif 16.5 <= hcp <= 17.3: return '18'
                elif 17.4 <= hcp <= 18.1: return '19'
                elif 18.2 <= hcp <= 19.0: return '20'
                elif 19.1 <= hcp <= 19.9: return '21'
                elif 20.0 <= hcp <= 20.8: return '22'
                elif 20.9 <= hcp <= 21.7: return '23'
                elif 21.8 <= hcp <= 22.5: return '24'
                elif 22.6 <= hcp <= 23.4: return '25'
                elif 23.5 <= hcp <= 24.3: return '26'
                elif 24.4 <= hcp <= 25.2: return '27'
                elif 25.3 <= hcp <= 26.1: return '28'
                elif 26.2 <= hcp <= 27.0: return '29'
                elif 27.1 <= hcp <= 27.8: return '30'
                elif 27.9 <= hcp <= 28.7: return '31'
                elif 28.8 <= hcp <= 29.6: return '32'
                elif 29.7 <= hcp <= 30.5: return '33'
                elif 30.6 <= hcp <= 31.4: return '34'
                elif 31.5 <= hcp <= 32.3: return '35'
                elif 32.4 <= hcp <= 33.1: return '36'
                elif 33.2 <= hcp <= 34.0: return '37'
                elif 34.1 <= hcp <= 34.9: return '38'
                elif 35.0 <= hcp <= 35.8: return '39'
                elif 35.9 <= hcp <= 36.7: return '40'
                elif 36.8 <= hcp <= 37.6: return '41'
                elif 37.7 <= hcp <= 38.4: return '42'
                elif 38.5 <= hcp <= 39.3: return '43'
                elif 39.4 <= hcp <= 40.2: return '44'
                elif 40.3 <= hcp <= 41.1: return '45'
                elif 41.2 <= hcp <= 42.0: return '46'
                elif 42.1 <= hcp <= 42.9: return '47'
                elif 43.0 <= hcp <= 43.7: return '48'
                elif 43.8 <= hcp <= 44.6: return '49'
                elif 44.7 <= hcp <= 45.5: return '50'
                elif 45.6 <= hcp <= 46.4: return '51'
                elif 46.5 <= hcp <= 47.3: return '52'
                elif 47.4 <= hcp <= 48.2: return '53'
                elif 48.3 <= hcp <= 49.0: return '54'
                elif 49.1 <= hcp <= 49.9: return '55'
                elif 50.0 <= hcp <= 50.8: return '56'
                elif 50.9 <= hcp <= 51.7: return '57'
                elif 51.8 <= hcp <= 52.6: return '58'
                elif 52.7 <= hcp <= 53.4: return '59'
                elif hcp >= 53.5: return '60'
                
                # Caso não encontre correspondência
                return 'N/A'
            
            # Calcular o HCP OGC Tee Branco
            hcp_ogc_white = get_hcp_ogc_white(hcp_value)
            
            # Atualizar o HCP OGC Tee Branco
            conn.execute('UPDATE players SET hcp_ogc_white = ? WHERE id = ?', (hcp_ogc_white, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Branco
            conn.execute('UPDATE players SET hcp_ogc_white = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Azul
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Azul: Course Rating = 71.4, Slope Rating = 131, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_azul = '+' + str(round(abs(hcp_value) * (131.0 / 113.0) + (71.4 - 71.0)))
            else:
                hcp_ogc_azul = str(round(hcp_value * (131.0 / 113.0) + (71.4 - 71.0)))
            
            # Atualizar o HCP OGC Tee Azul
            conn.execute('UPDATE players SET hcp_ogc_azul = ? WHERE id = ?', (hcp_ogc_azul, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Azul
            conn.execute('UPDATE players SET hcp_ogc_azul = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Preto
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Preto: Course Rating = 73.9, Slope Rating = 144, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_preto = '+' + str(round(abs(hcp_value) * (144.0 / 113.0) + (73.9 - 71.0)))
            else:
                hcp_ogc_preto = str(round(hcp_value * (144.0 / 113.0) + (73.9 - 71.0)))
            
            # Atualizar o HCP OGC Tee Preto
            conn.execute('UPDATE players SET hcp_ogc_preto = ? WHERE id = ?', (hcp_ogc_preto, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Preto
            conn.execute('UPDATE players SET hcp_ogc_preto = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Vermelho
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Vermelho: Course Rating = 68.1, Slope Rating = 125, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_vermelho = '+' + str(round(abs(hcp_value) * (125.0 / 113.0) + (68.1 - 71.0)))
            else:
                hcp_ogc_vermelho = str(round(hcp_value * (125.0 / 113.0) + (68.1 - 71.0)))
            
            # Atualizar o HCP OGC Tee Vermelho
            conn.execute('UPDATE players SET hcp_ogc_vermelho = ? WHERE id = ?', (hcp_ogc_vermelho, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Vermelho
            conn.execute('UPDATE players SET hcp_ogc_vermelho = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Amarelo
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Amarelo: Course Rating = 65.3, Slope Rating = 118, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_amarelo = '+' + str(round(abs(hcp_value) * (118.0 / 113.0) + (65.3 - 71.0)))
            else:
                hcp_ogc_amarelo = str(round(hcp_value * (118.0 / 113.0) + (65.3 - 71.0)))
            
            # Atualizar o HCP OGC Tee Amarelo
            conn.execute('UPDATE players SET hcp_ogc_amarelo = ? WHERE id = ?', (hcp_ogc_amarelo, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Amarelo
            conn.execute('UPDATE players SET hcp_ogc_amarelo = NULL WHERE id = ?', (player_id,))
        
        # Opcional: Registrar alteração nas notas
        notes = f"HCP Index alterado de '{old_hcp or 'não informado'}' para '{new_hcp or 'não informado'}' em {datetime.now().strftime('%d/%m/%Y')}"
        
        # Se o jogador já tem notas, adicionar à frente
        if player['notes']:
            notes = f"{player['notes']} | {notes}"
        
        # Atualizar as notas
        conn.execute('UPDATE players SET notes = ? WHERE id = ?', (notes, player_id))
        
        conn.commit()
        flash(f'HCP Index atualizado com sucesso.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar o HCP: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))


@app.route('/ranking_history')
@login_required 
def ranking_history():
    """Mostra o histórico de todas as posições em um gráfico"""
    conn = get_db_connection()
    
    # Verificar se foi fornecido um intervalo de datas personalizado
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Se intervalo personalizado não foi fornecido, usar período em dias
    if not (start_date and end_date):
        # Obter o período desejado (padrão: últimos 30 dias)
        days = request.args.get('days', 30, type=int)
        
        # Calcular a data limite
        limit_date = (datetime.now() - timedelta(days=days)).date()
        end_date = datetime.now().date().strftime('%Y-%m-%d')
    else:
        # Usar intervalo de datas personalizado
        limit_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        days = None  # Não usamos days quando temos start_date e end_date
    
    # Buscar as datas disponíveis no histórico
    dates = conn.execute('''
        SELECT DISTINCT date_recorded 
        FROM daily_ranking_history 
        WHERE date_recorded >= ?
        ORDER BY date_recorded
    ''', (limit_date.strftime('%Y-%m-%d'),)).fetchall()
    
    date_list = [d['date_recorded'] for d in dates]
    
    # Buscar os jogadores ativos
    players = conn.execute('''
        SELECT id, name, position 
        FROM players 
        WHERE active = 1 
        ORDER BY position
    ''').fetchall()
    
    # Criar uma lista combinada de players para o template
    players_list = [{'id': p['id'], 'name': p['name']} for p in players]
    
    conn.close()
    
    return render_template('ranking_history.html', 
                           days=days, 
                           dates=date_list,
                           players=players_list,
                           start_date=start_date,
                           end_date=end_date)


@app.route('/api/ranking_history_data')
def api_ranking_history_data():
    """API para obter os dados de histórico para o gráfico"""
    conn = get_db_connection()
    
    # Verificar se foi fornecido um intervalo de datas personalizado
    start_date = request.args.get('start_date')
    end_date = request.args.get('end_date')
    
    # Se intervalo personalizado não foi fornecido, usar período em dias
    if not (start_date and end_date):
        # Obter os parâmetros
        days = request.args.get('days', 30, type=int)
        
        # Calcular a data limite
        limit_date = (datetime.now() - timedelta(days=days)).date()
        end_date = datetime.now().date().strftime('%Y-%m-%d')
    else:
        # Usar intervalo de datas personalizado
        limit_date = datetime.strptime(start_date, '%Y-%m-%d').date()
        end_date = datetime.strptime(end_date, '%Y-%m-%d').date().strftime('%Y-%m-%d')
    
    # Converter limit_date para string no formato correto para SQL
    limit_date_str = limit_date.strftime('%Y-%m-%d')
    
    # Obter os IDs dos jogadores selecionados
    player_ids = request.args.getlist('player_ids[]', type=int)
    
    # Limitar o número de jogadores para performance
    if len(player_ids) > 30:
        player_ids = player_ids[:30]
    
    # Buscar as datas disponíveis no histórico
    date_query = '''
        SELECT DISTINCT date_recorded 
        FROM daily_ranking_history 
        WHERE date_recorded >= ? AND date_recorded <= ?
        ORDER BY date_recorded
    '''
    dates = conn.execute(date_query, (limit_date_str, end_date)).fetchall()
    
    date_list = [d['date_recorded'] for d in dates]
    
    # Preparar dados de cada jogador
    players_data = []
    
    for player_id in player_ids:
        # Buscar informações do jogador
        player = conn.execute('SELECT name FROM players WHERE id = ?', (player_id,)).fetchone()
        
        if not player:
            continue
        
        # Buscar histórico do jogador
        history_query = '''
            SELECT date_recorded, position 
            FROM daily_ranking_history 
            WHERE player_id = ? AND date_recorded >= ? AND date_recorded <= ?
            ORDER BY date_recorded
        '''
        history = conn.execute(history_query, (player_id, limit_date_str, end_date)).fetchall()
        
        # Criar um dicionário com as posições por data
        positions_by_date = {h['date_recorded']: h['position'] for h in history}
        
        # Montar a série temporal completa, mantendo a última posição conhecida para datas sem registro
        positions_series = []
        last_known_position = None
        
        for date in date_list:
            if date in positions_by_date:
                position = positions_by_date[date]
                last_known_position = position
            else:
                position = last_known_position
            
            if position is not None:
                positions_series.append(position)
            else:
                positions_series.append(None)  # Usar None para datas sem posição
        
        # Adicionar dados deste jogador
        players_data.append({
            'name': player['name'],
            'positions': positions_series
        })
    
    conn.close()
    
    return jsonify({
        'dates': date_list,
        'players': players_data
    })


@app.route('/sync_history', methods=['GET'])
def sync_history_route():
    """
    Sincroniza manualmente as tabelas de histórico para a data atual.
    """
    try:
        sync_ranking_history_tables()
        flash('Histórico sincronizado com sucesso!', 'success')
    except Exception as e:
        flash(f'Erro ao sincronizar histórico: {str(e)}', 'error')
    
    return redirect(url_for('ranking_history'))

# Atualização da parte relacionada à inicialização da aplicação


# Função para verificar e criar tabela de histórico diário
def create_daily_history_table():
    # código existente...
    print("Tabela de histórico diário criada com sucesso.")

# ============================================================
# FUNÇÃO add_response_deadline_column - CORRIGIDA
# ============================================================
# Prazo para RESPONDER: 2 dias
# ============================================================

# ============================================================
# FUNÇÃO add_response_deadline_column - CORRIGIDA
# ============================================================
# Prazo para RESPONDER: 2 dias
# ============================================================

def add_response_deadline_column():
    conn = get_db_connection()
    
    # Verificar se a coluna response_deadline existe
    columns_info = conn.execute('PRAGMA table_info(challenges)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    if 'response_deadline' not in column_names:
        # Adicionar coluna de prazo de resposta
        conn.execute('ALTER TABLE challenges ADD COLUMN response_deadline DATETIME')
        print("Coluna 'response_deadline' adicionada à tabela challenges.")
        
        # ============================================================
        # Definir prazo de resposta para desafios existentes: 2 DIAS
        # ============================================================
        conn.execute('''
            UPDATE challenges 
            SET response_deadline = datetime(created_at, '+2 days')
            WHERE status = 'pending' AND response_deadline IS NULL
        ''')
        print("Prazo de resposta (2 dias) definido para desafios pendentes existentes.")
    
    conn.commit()
    conn.close()


# ============================================================
# SQL PARA CORRIGIR DESAFIOS EXISTENTES
# ============================================================
# Execute este SQL diretamente no banco para corrigir
# desafios pendentes que foram criados com prazo de 7 dias:
#
# UPDATE challenges 
# SET response_deadline = datetime(created_at, '+2 days')
# WHERE status = 'pending';
#
# ============================================================
# Adicione esta nova rota ao seu arquivo app.py

@app.route('/admin/challenge_logs')
@login_required
def admin_challenge_logs():
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    # Consulta básica, sem JOINs para começo
    try:
        # Verificar se a tabela de logs existe
        table_exists = conn.execute("SELECT name FROM sqlite_master WHERE type='table' AND name='challenge_logs'").fetchone()
        
        if not table_exists:
            # Criar a tabela se não existir
            conn.execute('''
                CREATE TABLE challenge_logs (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    challenge_id INTEGER,
                    user_id TEXT NOT NULL,
                    modified_by TEXT NOT NULL,
                    old_status TEXT,
                    new_status TEXT,
                    old_result TEXT,
                    new_result TEXT,
                    notes TEXT,
                    created_at DATETIME DEFAULT CURRENT_TIMESTAMP
                )
            ''')
            conn.commit()
            
        # Consulta simplificada
        logs = conn.execute('SELECT * FROM challenge_logs ORDER BY created_at DESC LIMIT 100').fetchall()
        
    except Exception as e:
        logs = []
        flash(f'Erro ao carregar logs: {str(e)}', 'error')
    
    conn.close()
    
    return render_template('admin_challenge_logs.html', logs=logs, users=[])


@app.route('/privacy-policy')
def privacy_policy():
    """Página de Política de Privacidade e LGPD"""
    return render_template('privacy_policy.html')


@app.route('/data-export')
@login_required
def data_export():
    """Permite que o usuário baixe seus dados pessoais"""
    # Apenas usuário logado pode acessar seus próprios dados
    user_id = session.get('user_id')
    
    conn = get_db_connection()
    
    # Obter dados do jogador
    player_data = conn.execute('SELECT * FROM players WHERE id = ?', (user_id,)).fetchone()
    
    # Obter histórico de desafios
    challenges_as_challenger = conn.execute('''
        SELECT * FROM challenges WHERE challenger_id = ?
    ''', (user_id,)).fetchall()
    
    challenges_as_challenged = conn.execute('''
        SELECT * FROM challenges WHERE challenged_id = ?
    ''', (user_id,)).fetchall()
    
    # Obter histórico de rankings
    ranking_history = conn.execute('''
        SELECT * FROM ranking_history WHERE player_id = ?
    ''', (user_id,)).fetchall()
    
    conn.close()
    
    # Converter para formato JSON
    data = {
        'player_info': dict(player_data) if player_data else None,
        'challenges_as_challenger': [dict(row) for row in challenges_as_challenger],
        'challenges_as_challenged': [dict(row) for row in challenges_as_challenged],
        'ranking_history': [dict(row) for row in ranking_history]
    }
    
    # Criar resposta para download
    response = make_response(json.dumps(data, default=str, indent=4))
    response.headers["Content-Disposition"] = f"attachment; filename=data_export_{user_id}.json"
    response.headers["Content-Type"] = "application/json"
    
    return response

@app.route('/request-data-deletion', methods=['GET', 'POST'])
@login_required
def request_data_deletion():
    """Solicitar exclusão de dados pessoais"""
    if request.method == 'POST':
        # Implementação para lidar com a solicitação
        # Talvez envie um e-mail para o administrador ou marque o usuário
        # para exclusão futura
        
        flash('Sua solicitação de exclusão de dados foi recebida. Entraremos em contato em até 15 dias úteis.', 'success')
        return redirect(url_for('dashboard'))
    
    return render_template('request_data_deletion.html')



# Adicione esta nova rota ao seu arquivo app.py

# Adicione esta nova rota ao arquivo app.py

@app.route('/player/update_self_hcp', methods=['POST'])
@login_required
def update_self_hcp():
    """
    Permite que um jogador atualize seu próprio HCP Index sem necessidade de senha
    """
    # Obter o ID do jogador autenticado
    player_id = session.get('user_id')
    if not player_id or session.get('is_admin', False):
        flash('Acesso não autorizado.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    # Verificar se o jogador existe e está ativo
    player = conn.execute('SELECT * FROM players WHERE id = ? AND active = 1', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado ou inativo!', 'error')
        return redirect(url_for('dashboard'))
    
    # Obter novo HCP do formulário
    new_hcp = request.form.get('new_hcp', '').strip()
    old_hcp = str(player['hcp_index']) if player['hcp_index'] is not None else ''
    
    # Se o HCP não mudou, não fazer nada
    if new_hcp == old_hcp:
        conn.close()
        flash('Nenhuma alteração foi realizada.', 'info')
        return redirect(url_for('player_detail', player_id=player_id))
    
    try:
        # Converter o novo HCP para float se não estiver vazio
        hcp_value = None
        if new_hcp:
            try:
                hcp_value = float(new_hcp.replace(',', '.'))
            except ValueError:
                conn.close()
                flash('Valor de HCP inválido. Use apenas números.', 'error')
                return redirect(url_for('player_detail', player_id=player_id))
        
        # Atualizar o HCP do jogador com a data atual
        conn.execute('UPDATE players SET hcp_index = ?, hcp_last_update = CURRENT_TIMESTAMP WHERE id = ?', (hcp_value, player_id))
        
        # Calcular e atualizar o HCP OGC Tee Branco se o HCP Index foi fornecido
        if hcp_value is not None:
            # Função para determinar o HCP OGC Tee Branco com base no HCP Index
            def get_hcp_ogc_white(hcp):
                # Para handicaps "plus" (valores negativos no banco)
                if hcp <= -0.3:
                    if -5.0 <= hcp <= -4.8: return '+7'
                    elif -4.7 <= hcp <= -3.9: return '+6'
                    elif -3.8 <= hcp <= -3.1: return '+5'
                    elif -3.0 <= hcp <= -2.2: return '+4'
                    elif -2.1 <= hcp <= -1.3: return '+3'
                    elif -1.2 <= hcp <= -0.4: return '+2'
                    elif -0.3 <= hcp <= 0.5: return '+1'
                
                # Para handicaps regulares (valores positivos no banco)
                if 0.6 <= hcp <= 1.4: return '0'
                elif 1.5 <= hcp <= 2.2: return '1'
                elif 2.3 <= hcp <= 3.1: return '2'
                elif 3.2 <= hcp <= 4.0: return '3'
                elif 4.1 <= hcp <= 4.9: return '4'
                elif 5.0 <= hcp <= 5.8: return '5'
                elif 5.9 <= hcp <= 6.7: return '6'
                elif 6.8 <= hcp <= 7.5: return '7'
                elif 7.6 <= hcp <= 8.4: return '8'
                elif 8.5 <= hcp <= 9.3: return '9'
                elif 9.4 <= hcp <= 10.2: return '10'
                elif 10.3 <= hcp <= 11.1: return '11'
                elif 11.2 <= hcp <= 12.0: return '12'
                elif 12.1 <= hcp <= 12.8: return '13'
                elif 12.9 <= hcp <= 13.7: return '14'
                elif 13.8 <= hcp <= 14.6: return '15'
                elif 14.7 <= hcp <= 15.5: return '16'
                elif 15.6 <= hcp <= 16.4: return '17'
                elif 16.5 <= hcp <= 17.3: return '18'
                elif 17.4 <= hcp <= 18.1: return '19'
                elif 18.2 <= hcp <= 19.0: return '20'
                elif 19.1 <= hcp <= 19.9: return '21'
                elif 20.0 <= hcp <= 20.8: return '22'
                elif 20.9 <= hcp <= 21.7: return '23'
                elif 21.8 <= hcp <= 22.5: return '24'
                elif 22.6 <= hcp <= 23.4: return '25'
                elif 23.5 <= hcp <= 24.3: return '26'
                elif 24.4 <= hcp <= 25.2: return '27'
                elif 25.3 <= hcp <= 26.1: return '28'
                elif 26.2 <= hcp <= 27.0: return '29'
                elif 27.1 <= hcp <= 27.8: return '30'
                elif 27.9 <= hcp <= 28.7: return '31'
                elif 28.8 <= hcp <= 29.6: return '32'
                elif 29.7 <= hcp <= 30.5: return '33'
                elif 30.6 <= hcp <= 31.4: return '34'
                elif 31.5 <= hcp <= 32.3: return '35'
                elif 32.4 <= hcp <= 33.1: return '36'
                elif 33.2 <= hcp <= 34.0: return '37'
                elif 34.1 <= hcp <= 34.9: return '38'
                elif 35.0 <= hcp <= 35.8: return '39'
                elif 35.9 <= hcp <= 36.7: return '40'
                elif 36.8 <= hcp <= 37.6: return '41'
                elif 37.7 <= hcp <= 38.4: return '42'
                elif 38.5 <= hcp <= 39.3: return '43'
                elif 39.4 <= hcp <= 40.2: return '44'
                elif 40.3 <= hcp <= 41.1: return '45'
                elif 41.2 <= hcp <= 42.0: return '46'
                elif 42.1 <= hcp <= 42.9: return '47'
                elif 43.0 <= hcp <= 43.7: return '48'
                elif 43.8 <= hcp <= 44.6: return '49'
                elif 44.7 <= hcp <= 45.5: return '50'
                elif 45.6 <= hcp <= 46.4: return '51'
                elif 46.5 <= hcp <= 47.3: return '52'
                elif 47.4 <= hcp <= 48.2: return '53'
                elif 48.3 <= hcp <= 49.0: return '54'
                elif 49.1 <= hcp <= 49.9: return '55'
                elif 50.0 <= hcp <= 50.8: return '56'
                elif 50.9 <= hcp <= 51.7: return '57'
                elif 51.8 <= hcp <= 52.6: return '58'
                elif 52.7 <= hcp <= 53.4: return '59'
                elif hcp >= 53.5: return '60'
                
                # Caso não encontre correspondência
                return 'N/A'
            
            # Calcular o HCP OGC Tee Branco
            hcp_ogc_white = get_hcp_ogc_white(hcp_value)
            
            # Atualizar o HCP OGC Tee Branco
            conn.execute('UPDATE players SET hcp_ogc_white = ? WHERE id = ?', (hcp_ogc_white, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Branco
            conn.execute('UPDATE players SET hcp_ogc_white = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Azul
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Azul: Course Rating = 71.4, Slope Rating = 131, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_azul = '+' + str(round(abs(hcp_value) * (131.0 / 113.0) + (71.4 - 71.0)))
            else:
                hcp_ogc_azul = str(round(hcp_value * (131.0 / 113.0) + (71.4 - 71.0)))
            
            # Atualizar o HCP OGC Tee Azul
            conn.execute('UPDATE players SET hcp_ogc_azul = ? WHERE id = ?', (hcp_ogc_azul, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Azul
            conn.execute('UPDATE players SET hcp_ogc_azul = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Preto
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Preto: Course Rating = 73.9, Slope Rating = 144, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_preto = '+' + str(round(abs(hcp_value) * (144.0 / 113.0) + (73.9 - 71.0)))
            else:
                hcp_ogc_preto = str(round(hcp_value * (144.0 / 113.0) + (73.9 - 71.0)))
            
            # Atualizar o HCP OGC Tee Preto
            conn.execute('UPDATE players SET hcp_ogc_preto = ? WHERE id = ?', (hcp_ogc_preto, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Preto
            conn.execute('UPDATE players SET hcp_ogc_preto = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Vermelho
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Vermelho: Course Rating = 68.1, Slope Rating = 125, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_vermelho = '+' + str(round(abs(hcp_value) * (125.0 / 113.0) + (68.1 - 71.0)))
            else:
                hcp_ogc_vermelho = str(round(hcp_value * (125.0 / 113.0) + (68.1 - 71.0)))
            
            # Atualizar o HCP OGC Tee Vermelho
            conn.execute('UPDATE players SET hcp_ogc_vermelho = ? WHERE id = ?', (hcp_ogc_vermelho, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Vermelho
            conn.execute('UPDATE players SET hcp_ogc_vermelho = NULL WHERE id = ?', (player_id,))
        
        # Calcular o HCP OGC Tee Amarelo
        if hcp_value is not None:
            # Fórmula: Handicap Index × (Slope Rating ÷ 113) + (Course Rating - Par)
            # Para o Tee Amarelo: Course Rating = 65.3, Slope Rating = 118, Par = 71
            if hcp_value < 0:  # Handicap "plus"
                # Adiciona um sinal '+' para handicaps plus
                hcp_ogc_amarelo = '+' + str(round(abs(hcp_value) * (118.0 / 113.0) + (65.3 - 71.0)))
            else:
                hcp_ogc_amarelo = str(round(hcp_value * (118.0 / 113.0) + (65.3 - 71.0)))
            
            # Atualizar o HCP OGC Tee Amarelo
            conn.execute('UPDATE players SET hcp_ogc_amarelo = ? WHERE id = ?', (hcp_ogc_amarelo, player_id))
        else:
            # Se o HCP Index foi removido, remover também o HCP OGC Tee Amarelo
            conn.execute('UPDATE players SET hcp_ogc_amarelo = NULL WHERE id = ?', (player_id,))
        
        # Registrar a atualização no log de alterações
        now = datetime.now().strftime('%d/%m/%Y')
        log_message = f"HCP Index atualizado pelo próprio jogador de '{old_hcp or 'não informado'}' para '{new_hcp or 'não informado'}' em {now}"
        
        # Atualizar notas se a coluna existir
        if player['notes']:
            notes = f"{player['notes']} | {log_message}"
        else:
            notes = log_message
        
        conn.execute('UPDATE players SET notes = ? WHERE id = ?', (notes, player_id))
        
        # Registrar a alteração no histórico de HCP, se a função existir
        old_hcp_value = float(old_hcp.replace(',', '.')) if old_hcp and old_hcp.strip() else None
        try:
            # Verifique se a função record_hcp_change existe
            if 'record_hcp_change' in globals():
                record_hcp_change(player_id, old_hcp_value, hcp_value, 'player')
        except Exception as e:
            print(f"Erro ao registrar histórico de HCP: {e}")
        
        conn.commit()
        flash('Seu HCP Index foi atualizado com sucesso!', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao atualizar o HCP: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))


@app.route('/add_hcp_last_update_column')
def add_hcp_last_update_column():
    conn = get_db_connection()
    
    # Verificar se a coluna já existe
    columns_info = conn.execute('PRAGMA table_info(players)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    if 'hcp_last_update' not in column_names:
        conn.execute('ALTER TABLE players ADD COLUMN hcp_last_update DATETIME')
        conn.commit()
        result = "Coluna 'hcp_last_update' adicionada com sucesso."
    else:
        result = "Coluna 'hcp_last_update' já existe."
    
    conn.close()
    return result


@app.route('/reset_player_password/<int:player_id>', methods=['POST'])
@login_required
def reset_player_password(player_id):
    """
    Permite que um administrador resete a senha de um jogador para o padrão (3 primeiras letras do nome)
    """
    # Verificar se é administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    # Verificar senha de admin
    # Verificação de admin (senha hardcoded removida)

    if not session.get('is_admin', False):

        flash('Acesso negado. Apenas administradores podem executar esta ação.', 'error')

        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    # Buscar informações do jogador
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado!', 'error')
        return redirect(url_for('index'))
    
    # Definir nova senha como as 3 primeiras letras do nome em minúsculas
    default_password = player['name'].strip().lower()[:3]
    hashed_password = hash_password(default_password)
    
    # Atualizar a senha
    conn.execute('UPDATE players SET password = ? WHERE id = ?', 
                (hashed_password, player_id))
    
    conn.commit()
    conn.close()
    
    flash(f'Senha do jogador {player["name"]} resetada com sucesso! A nova senha é: {default_password}', 'success')
    return redirect(url_for('player_detail', player_id=player_id))




def add_country_column():
    conn = get_db_connection()
    
    # Verificar se a coluna já existe
    columns_info = conn.execute('PRAGMA table_info(players)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    if 'country' not in column_names:
        conn.execute('ALTER TABLE players ADD COLUMN country TEXT DEFAULT NULL')
        conn.commit()
        print("Coluna 'country' adicionada à tabela players com valor padrão NULL.")
    else:
        print("Coluna 'country' já existe na tabela players.")
    
    conn.close()



# Adicione este código ao app.py para criar um filtro personalizado

@app.template_filter('country_code')
def country_code_filter(country_name):
    """
    Converte o nome do país para o código ISO de 2 letras usado para exibir bandeiras.
    """
    # Mapeamento de nomes de países para códigos ISO de 2 letras
    country_mapping = {
        'Brasil': 'br',
        'Argentina': 'ar',
        'Portugal': 'pt',
        'Estados Unidos': 'us',
        'Espanha': 'es',
        'Itália': 'it',
        'França': 'fr',
        'Alemanha': 'de',
        'Reino Unido': 'gb',
        'Inglaterra': 'gb-eng',
        'Escócia': 'gb-sct',
        'País de Gales': 'gb-wls',
        'Irlanda do Norte': 'gb-nir',
        'Japão': 'jp',
        'Coreia do Sul': 'kr',
        'China': 'cn',
        'Austrália': 'au',
        'Canadá': 'ca',
        'México': 'mx',
        'Chile': 'cl',
        'Colômbia': 'co',
        'Uruguai': 'uy',
        'Paraguai': 'py',
        'Peru': 'pe',
        'Venezuela': 've',
        'África do Sul': 'za',
        'Suíça': 'ch',
        'Suécia': 'se',
        'Noruega': 'no',
        'Dinamarca': 'dk',
        'Holanda': 'nl',
        'Países Baixos': 'nl',
        'Bélgica': 'be',
        'Irlanda': 'ie',
        'Nova Zelândia': 'nz',
        'Índia': 'in',
        'Rússia': 'ru',
        'Polônia': 'pl',
        'Áustria': 'at',
        'Grécia': 'gr',
        'Turquia': 'tr'
    }
    
    # Retorna o código ISO ou o nome do país em minúsculas como fallback
    return country_mapping.get(country_name, country_name.lower())


@app.route('/regulamento')
def regulamento():
    conn = get_db_connection()
    regulamento = conn.execute('''
        SELECT * FROM regulamento ORDER BY data_upload DESC LIMIT 1
    ''').fetchone()
    conn.close()
    
    return render_template('regulamento.html', regulamento=regulamento)


@app.route('/relatorio')
def relatorio():
    return render_template('relatorio.html')


@app.route('/admin/create_admin', methods=['GET', 'POST'])
@login_required
def create_admin():
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Obter dados do formulário
        username = request.form.get('username', '').strip()
        password = request.form.get('password', '').strip()
        confirm_password = request.form.get('confirm_password', '').strip()
        name = request.form.get('name', '').strip()
        email = request.form.get('email', '').strip()
        admin_password = request.form.get('admin_password', '').strip()
        if not username or not password or not name:
            flash('Campos obrigatórios não preenchidos.', 'error')
            return redirect(url_for('create_admin'))
        
        # Verificar senha do admin atual
        if not session.get('is_admin', False):
            flash('Senha de administrador incorreta! Operação não autorizada.', 'error')
            return redirect(url_for('create_admin'))
        
        # Verificar se as senhas coincidem
        if password != confirm_password:
            flash('A senha e a confirmação não coincidem.', 'error')
            return redirect(url_for('create_admin'))
        
        conn = get_db_connection()
        
        try:
            # Verificar se o nome de usuário já existe
            existing_admin = conn.execute('SELECT * FROM admins WHERE username = ?', (username,)).fetchone()
            
            if existing_admin:
                conn.close()
                flash(f'O nome de usuário "{username}" já está em uso. Escolha outro.', 'error')
                return redirect(url_for('create_admin'))
            
            # Criar o hash da senha - verificar a implementação
            hashed_password = hash_password(password)
            
            # Imprimir para debug (remover em produção)
            print(f"Criando admin: {username}, Senha original: {password}, Hash: {hashed_password}")
            
            # Inserir o novo administrador
            conn.execute('''
                INSERT INTO admins (username, password, name, email, created_at)
                VALUES (?, ?, ?, ?, CURRENT_TIMESTAMP)
            ''', (username, hashed_password, name, email))
            
            # Registrar em log para debug (opcional)
            conn.execute('''
                INSERT INTO system_settings (key, value, updated_at)
                VALUES (?, ?, CURRENT_TIMESTAMP)
            ''', (f"admin_creation_{username}", f"Admin {name} criado em {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"))
            
            conn.commit()
            flash(f'Administrador "{name}" criado com sucesso! Use o nome de usuário "{username}" para login.', 'success')
            
            # Listar administradores para verificação
            admins = conn.execute('SELECT username, password FROM admins').fetchall()
            print("Lista de admins no banco:")
            for admin in admins:
                print(f"Username: {admin['username']}, Hash senha: {admin['password']}")
            
            return redirect(url_for('admin_dashboard'))
            
        except Exception as e:
            conn.rollback()
            flash(f'Erro ao criar administrador: {str(e)}', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('create_admin'))
    
    # Para requisição GET, mostrar formulário
    return render_template('create_admin.html')



# Rota para listar todos os administradores
@app.route('/admin/list_admins')
@login_required
def list_admins():
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    admins = conn.execute('SELECT id, username, name, email, created_at, last_login FROM admins ORDER BY name').fetchall()
    conn.close()
    
    return render_template('list_admins.html', admins=admins)


@app.route('/admin/fix_admin_passwords', methods=['GET', 'POST'])
@login_required
def fix_admin_passwords():
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        admin_password = request.form.get('admin_password', '')
        if not session.get('is_admin', False):
            conn.close()
            flash('Senha de administrador incorreta! Operação não autorizada.', 'error')
            return redirect(url_for('fix_admin_passwords'))
        
        # Obter todos os administradores
        admins = conn.execute('SELECT id, username FROM admins WHERE username != "admin"').fetchall()
        
        # Redefinir as senhas para os nomes de usuário
        for admin in admins:
            # A nova senha será o próprio nome de usuário
            new_password = admin['username']
            hashed_password = hash_password(new_password)
            
            # Atualizar a senha
            conn.execute('UPDATE admins SET password = ? WHERE id = ?', (hashed_password, admin['id']))
            
            # Registrar a alteração
            print(f"Redefinida senha do admin {admin['username']}: {new_password} -> {hashed_password}")
        
        conn.commit()
        flash('As senhas de todos os administradores foram redefinidas. A nova senha é igual ao nome de usuário.', 'success')
        return redirect(url_for('list_admins'))
    
    # Para requisição GET, mostrar formulário
    return render_template('fix_admin_passwords.html')


@app.route('/admin/reset_admin_password/<int:admin_id>', methods=['POST'])
@login_required
def reset_admin_password(admin_id):
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    admin_password = request.form.get('admin_password', '')
    if not session.get('is_admin', False):
        flash('Senha incorreta! Operação não autorizada.', 'error')
        return redirect(url_for('list_admins'))
    
    conn = get_db_connection()
    
    try:
        # Buscar o admin a ter a senha resetada
        admin = conn.execute('SELECT * FROM admins WHERE id = ?', (admin_id,)).fetchone()
        
        if not admin:
            conn.close()
            flash('Administrador não encontrado.', 'error')
            return redirect(url_for('list_admins'))
        
        # Não permitir resetar a senha do admin principal
        if admin['username'] == 'admin':
            conn.close()
            flash('Não é possível resetar a senha do administrador principal.', 'error')
            return redirect(url_for('list_admins'))
        
        # A nova senha será o próprio nome de usuário
        new_password = admin['username']
        hashed_password = hash_password(new_password)
        
        # Atualizar a senha
        conn.execute('UPDATE admins SET password = ? WHERE id = ?', (hashed_password, admin_id))
        conn.commit()
        
        flash(f'Senha de {admin["name"]} resetada com sucesso. A nova senha é: {new_password}', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao resetar senha: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('list_admins'))



@app.route('/admin/delete_admin/<int:admin_id>', methods=['GET', 'POST'])
@login_required
def delete_admin(admin_id):
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    # Buscar o admin a ser excluído
    admin = conn.execute('SELECT * FROM admins WHERE id = ?', (admin_id,)).fetchone()
    
    if not admin:
        conn.close()
        flash('Administrador não encontrado!', 'error')
        return redirect(url_for('list_admins'))
    
    # Verificar se é o admin principal (não pode ser excluído)
    if admin['username'] == 'admin':
        conn.close()
        flash('O administrador principal não pode ser excluído.', 'error')
        return redirect(url_for('list_admins'))
    
    # Verificar se é o próprio usuário tentando se excluir
    admin_current_id = session.get('user_id', '').split('_')[1] if isinstance(session.get('user_id', ''), str) else None
    if admin_current_id and int(admin_current_id) == admin_id:
        conn.close()
        flash('Você não pode excluir sua própria conta de administrador.', 'error')
        return redirect(url_for('list_admins'))
    
    # Para requisição GET, mostrar tela de confirmação
    if request.method == 'GET':
        conn.close()
        return render_template('delete_admin.html', admin=admin)
    
    # Para requisição POST, processar a exclusão
    senha = request.form.get('admin_password', '')
    confirm_delete = request.form.get('confirm_delete', 'no') == 'yes'
    
    if not session.get('is_admin', False):
        conn.close()
        flash('Senha incorreta! Operação não autorizada.', 'error')
        return redirect(url_for('delete_admin', admin_id=admin_id))
    
    if not confirm_delete:
        conn.close()
        flash('Você precisa confirmar a exclusão marcando a caixa de confirmação.', 'error')
        return redirect(url_for('delete_admin', admin_id=admin_id))
    
    try:
        # Registrar a ação de exclusão em um log
        conn.execute('''
            INSERT INTO system_settings (key, value, updated_at)
            VALUES (?, ?, CURRENT_TIMESTAMP)
        ''', (f"admin_deletion_{admin['username']}", f"Admin {admin['name']} excluído em {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} pelo administrador {session.get('username', 'desconhecido')}"))
        
        # Excluir o administrador
        conn.execute('DELETE FROM admins WHERE id = ?', (admin_id,))
        
        conn.commit()
        flash(f'Administrador "{admin["name"]}" (username: {admin["username"]}) foi excluído com sucesso.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao excluir administrador: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('list_admins'))

def add_profile_photo_column():
    conn = get_db_connection()
    
    # Verificar se a coluna já existe
    columns_info = conn.execute('PRAGMA table_info(players)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    if 'profile_photo' not in column_names:
        conn.execute('ALTER TABLE players ADD COLUMN profile_photo TEXT DEFAULT NULL')
        conn.commit()
        print("Coluna 'profile_photo' adicionada à tabela players com valor padrão NULL.")
    else:
        print("Coluna 'profile_photo' já existe na tabela players.")
    
    conn.close()


# Rota para a página Golf Business
# ============================================================
# ROTAS ATUALIZADAS PARA SUPORTAR AFFINITY CLUB
# Substitua as rotas correspondentes no seu app.py
# ============================================================


# Rota para a página Golf Business
# ============================================================
# ROTAS ATUALIZADAS PARA SUPORTAR AFFINITY CLUB
# Substitua as rotas correspondentes no seu app.py
# ============================================================


# Rota para a página Golf Business
@app.route('/golf-business')
def golf_business():
    conn = get_db_connection()
    # ALTERADO: LEFT JOIN para permitir player_id NULL (Affinity)
    businesses = conn.execute('''
        SELECT b.*, 
               COALESCE(p.name, 'LIGA OLÍMPICA DE GOLFE') as owner_name, 
               COALESCE(p.profile_photo, 'logo-liga.png') as owner_photo
        FROM businesses b
        LEFT JOIN players p ON b.player_id = p.id
        WHERE b.active = 1
        ORDER BY b.created_at DESC
    ''').fetchall()
    conn.close()
    
    return render_template('golf_business.html', businesses=businesses)


# Rota para processamento do formulário de adição de negócio
@app.route('/add-business', methods=['POST'])
@login_required
def add_business():
    # Verificar se é administrador
    if not session.get('is_admin', False):
        flash('Apenas administradores podem adicionar negócios.', 'error')
        return redirect(url_for('golf_business'))
    
    if request.method == 'POST':
        try:
            # Obter dados do formulário
            player_id = request.form.get('player_id')
            business_name = request.form.get('business_name')
            business_category = request.form.get('business_category')
            business_description = request.form.get('business_description')
            business_contact = request.form.get('business_contact')
            
            # NOVO: Para categoria 'affinity', player_id pode ser NULL
            if business_category == 'affinity':
                player_id = None  # Será divulgado pela LIGA OLÍMPICA
            elif not player_id:
                flash('Selecione um jogador para divulgar o negócio.', 'error')
                return redirect(url_for('admin_business'))
            
            # Validar campos obrigatórios (removido player_id da validação para affinity)
            if not business_name or not business_category or not business_description:
                flash('Todos os campos obrigatórios devem ser preenchidos.', 'error')
                return redirect(url_for('admin_business'))
                
            # Processar imagem
            if 'business_image' in request.files:
                file = request.files['business_image']
                if file and allowed_file(file.filename):
                    # Gerar nome de arquivo seguro
                    # ALTERADO: usar 'affinity' no nome se player_id for None
                    file_prefix = f"business_{player_id}" if player_id else "business_affinity"
                    filename = secure_filename(f"{file_prefix}_{int(datetime.now().timestamp())}.{file.filename.rsplit('.', 1)[1].lower()}")
                    
                    # Criar diretório se não existir
                    business_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'business_images')
                    os.makedirs(business_upload_folder, exist_ok=True)
                    
                    # Salvar arquivo
                    file_path = os.path.join(business_upload_folder, filename)
                    file.save(file_path)
                    
                    # Salvar no banco de dados
                    conn = get_db_connection()
                    conn.execute('''
                        INSERT INTO businesses 
                        (player_id, name, category, description, image_path, contact_info, created_at)
                        VALUES (?, ?, ?, ?, ?, ?, CURRENT_TIMESTAMP)
                    ''', (
                        player_id,  # Pode ser None para affinity
                        business_name,
                        business_category,
                        business_description,
                        filename,
                        business_contact
                    ))
                    
                    conn.commit()
                    conn.close()
                    
                    flash('Negócio cadastrado com sucesso!', 'success')
                    return redirect(url_for('admin_business'))
                else:
                    flash('Tipo de arquivo não permitido. Use apenas JPG, PNG ou GIF.', 'error')
            else:
                flash('Imagem é obrigatória para cadastro do negócio.', 'error')
        
        except Exception as e:
            flash(f'Erro ao cadastrar negócio: {str(e)}', 'error')
        
        return redirect(url_for('admin_business'))


@app.route('/api/businesses')
def api_businesses():
    filter_category = request.args.get('filter', 'all')
    
    conn = get_db_connection()
    
    # ALTERADO: LEFT JOIN e COALESCE para suportar player_id NULL
    query = '''
        SELECT b.*, 
               COALESCE(p.name, 'LIGA OLÍMPICA DE GOLFE') as owner_name, 
               COALESCE(p.profile_photo, 'logo-liga.png') as owner_photo
        FROM businesses b
        LEFT JOIN players p ON b.player_id = p.id
        WHERE b.active = 1
    '''
    
    # Aplicar filtro se não for "all"
    params = []
    if filter_category != 'all':
        query += ' AND b.category = ?'
        params.append(filter_category)
    
    query += ' ORDER BY b.created_at DESC'
    businesses = conn.execute(query, params).fetchall()
    
    # Converter para formato JSON
    business_list = []
    for b in businesses:
        # ALTERADO: Tratar owner_photo para casos sem jogador
        owner_photo = b['owner_photo']
        if owner_photo == 'logo-liga.png':
            owner_photo_url = "/static/images/logo-liga.png"
        elif owner_photo:
            owner_photo_url = f"/static/profile_photos/{owner_photo}"
        else:
            owner_photo_url = "/static/profile_photos/default.png"
        
        business_dict = {
            'id': b['id'],
            'name': b['name'],
            'description': b['description'],
            'category': b['category'],
            'image_path': f"/static/profile_photos/business_images/{b['image_path']}" if b['image_path'] else None,
            'contact_info': b['contact_info'],
            'owner_name': b['owner_name'],
            'owner_photo': owner_photo_url
        }
        business_list.append(business_dict)
    
    conn.close()
    
    return jsonify({'businesses': business_list})


@app.route('/admin/business')
@login_required
def admin_business():
    # Verificar permissão de administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    # ALTERADO: LEFT JOIN e COALESCE para suportar player_id NULL
    businesses = conn.execute('''
        SELECT b.*, 
               COALESCE(p.name, 'LIGA OLÍMPICA DE GOLFE') as owner_name
        FROM businesses b
        LEFT JOIN players p ON b.player_id = p.id
        ORDER BY b.created_at DESC
    ''').fetchall()
    
    # Buscar jogadores para o formulário
    players = conn.execute('SELECT id, name FROM players WHERE active = 1 ORDER BY name').fetchall()
    
    conn.close()
    
    return render_template('admin_business.html', businesses=businesses, players=players)


@app.route('/admin/edit-business/<int:business_id>', methods=['POST'])
@login_required
def edit_business(business_id):
    # Verificar permissão
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Obter dados do formulário
        player_id = request.form.get('player_id')
        business_name = request.form.get('business_name')
        business_category = request.form.get('business_category')
        business_description = request.form.get('business_description')
        business_contact = request.form.get('business_contact')
        
        # NOVO: Para categoria 'affinity', player_id pode ser NULL
        if business_category == 'affinity':
            player_id = None
        elif not player_id:
            flash('Selecione um jogador para divulgar o negócio.', 'error')
            return redirect(url_for('admin_business'))
        
        # Validar dados (removido player_id da validação para affinity)
        if not business_name or not business_category or not business_description:
            flash('Todos os campos obrigatórios devem ser preenchidos.', 'error')
            return redirect(url_for('admin_business'))
        
        conn = get_db_connection()
        
        # Obter informações do negócio atual
        current_business = conn.execute('SELECT * FROM businesses WHERE id = ?', (business_id,)).fetchone()
        
        if not current_business:
            conn.close()
            flash('Negócio não encontrado!', 'error')
            return redirect(url_for('admin_business'))
        
        # Processar atualização da imagem (se fornecida)
        if 'business_image' in request.files and request.files['business_image'].filename:
            file = request.files['business_image']
            
            if file and allowed_file(file.filename):
                # Gerar nome de arquivo seguro
                file_prefix = f"business_{player_id}" if player_id else "business_affinity"
                filename = secure_filename(f"{file_prefix}_{int(datetime.now().timestamp())}.{file.filename.rsplit('.', 1)[1].lower()}")
                
                # Criar diretório se não existir
                business_upload_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'business_images')
                os.makedirs(business_upload_folder, exist_ok=True)
                
                # Salvar o arquivo
                file_path = os.path.join(business_upload_folder, filename)
                file.save(file_path)
                
                # Remover a imagem antiga, se existir
                if current_business['image_path']:
                    try:
                        old_file_path = os.path.join(business_upload_folder, current_business['image_path'])
                        if os.path.exists(old_file_path):
                            os.remove(old_file_path)
                    except Exception as e:
                        print(f"Erro ao remover imagem antiga: {e}")
                
                # Atualizar com a nova imagem
                conn.execute('''
                    UPDATE businesses
                    SET player_id = ?, name = ?, category = ?, description = ?,
                        contact_info = ?, image_path = ?
                    WHERE id = ?
                ''', (
                    player_id, business_name, business_category,
                    business_description, business_contact, filename, business_id
                ))
            else:
                conn.close()
                flash('Tipo de arquivo não permitido. Use apenas JPG, PNG ou GIF.', 'error')
                return redirect(url_for('admin_business'))
        else:
            # Atualizar sem alterar a imagem
            conn.execute('''
                UPDATE businesses
                SET player_id = ?, name = ?, category = ?, description = ?,
                    contact_info = ?
                WHERE id = ?
            ''', (
                player_id, business_name, business_category,
                business_description, business_contact, business_id
            ))
        
        conn.commit()
        conn.close()
        
        flash('Negócio atualizado com sucesso!', 'success')
        
    except Exception as e:
        flash(f'Erro ao atualizar negócio: {str(e)}', 'error')
    
    return redirect(url_for('admin_business'))


@app.route('/admin/delete-business/<int:business_id>', methods=['POST'])
@login_required
def delete_business(business_id):
    # Verificar permissão
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    try:
        # Obter informações do negócio
        business = conn.execute('SELECT * FROM businesses WHERE id = ?', (business_id,)).fetchone()
        
        if not business:
            conn.close()
            flash('Negócio não encontrado!', 'error')
            return redirect(url_for('admin_business'))
        
        # Marcar como inativo (soft delete)
        conn.execute('UPDATE businesses SET active = 0 WHERE id = ?', (business_id,))
        
        # Opcional: Remover fisicamente a imagem
        if business['image_path']:
            try:
                image_path = os.path.join(app.config['UPLOAD_FOLDER'], 'business_images', business['image_path'])
                if os.path.exists(image_path):
                    os.remove(image_path)
            except Exception as e:
                print(f"Erro ao remover arquivo de imagem: {e}")
        
        conn.commit()
        flash('Negócio excluído com sucesso!', 'success')
    
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao excluir negócio: {str(e)}', 'error')
    
    finally:
        conn.close()
    
    return redirect(url_for('admin_business'))



@app.route('/fix_male_ranking_now')
@login_required
def fix_male_ranking_now():
    """Corrige o ranking masculino imediatamente"""
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    try:
        # Buscar jogadores masculinos ativos ordenados por posição atual
        male_players = conn.execute('''
            SELECT id FROM players 
            WHERE active = 1 AND (sexo != 'feminino' OR sexo IS NULL)
            ORDER BY position
        ''').fetchall()
        
        # Reatribuir posições sequenciais
        for i, player in enumerate(male_players, 1):
            new_tier = get_tier_from_position(i)
            conn.execute('UPDATE players SET position = ?, tier = ? WHERE id = ?', 
                        (i, new_tier, player['id']))
        
        conn.commit()
        flash(f'✅ Ranking masculino corrigido! {len(male_players)} jogadores reorganizados.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Erro: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('pyramid_dynamic'))



# Adicione esta rota para análise sistemática

@app.route('/analyze_tier_structure')
@login_required
def analyze_tier_structure():
    if not session.get('is_admin', False):
        return "Acesso negado"
    
    conn = get_db_connection()
    
    # 1. Verificar a estrutura PYRAMID_STRUCTURE definida
    analysis = []
    analysis.append("=== ANÁLISE DA ESTRUTURA DE TIERS ===\n")
    
    # Verificar a estrutura definida no código
    analysis.append("1. ESTRUTURA PYRAMID_STRUCTURE:")
    for tier, positions in PYRAMID_STRUCTURE.items():
        analysis.append(f"   Tier {tier}: {len(positions)} posições ({min(positions)}-{max(positions)})")
    
    analysis.append("\n2. VERIFICANDO CONTINUIDADE DAS POSIÇÕES:")
    all_positions = []
    for positions in PYRAMID_STRUCTURE.values():
        all_positions.extend(positions)
    all_positions.sort()
    
    # Verificar se há lacunas ou duplicatas na estrutura
    expected = list(range(1, len(all_positions) + 1))
    if all_positions != expected:
        analysis.append(f"   ❌ PROBLEMA: Posições esperadas {expected[:10]}...{expected[-10:]}")
        analysis.append(f"   ❌ PROBLEMA: Posições definidas {all_positions[:10]}...{all_positions[-10:]}")
        
        # Encontrar lacunas
        missing = set(expected) - set(all_positions)
        duplicates = [pos for pos in all_positions if all_positions.count(pos) > 1]
        
        if missing:
            analysis.append(f"   ❌ Posições faltando: {sorted(missing)}")
        if duplicates:
            analysis.append(f"   ❌ Posições duplicadas: {sorted(set(duplicates))}")
    else:
        analysis.append(f"   ✅ Estrutura correta: posições 1-{len(all_positions)} sem lacunas")
    
    # 3. Verificar jogadores reais no banco
    analysis.append("\n3. JOGADORES REAIS NO BANCO:")
    male_players = conn.execute('''
        SELECT position, tier, name 
        FROM players 
        WHERE active = 1 AND (sexo != 'feminino' OR sexo IS NULL)
        ORDER BY position
    ''').fetchall()
    
    analysis.append(f"   Total de jogadores masculinos ativos: {len(male_players)}")
    
    # Verificar posições dos jogadores
    actual_positions = [p['position'] for p in male_players]
    expected_positions = list(range(1, len(male_players) + 1))
    
    if actual_positions != expected_positions:
        analysis.append("   ❌ PROBLEMA: Posições dos jogadores não são sequenciais")
        analysis.append(f"   Posições reais: {actual_positions[:20]}{'...' if len(actual_positions) > 20 else ''}")
        analysis.append(f"   Posições esperadas: {expected_positions[:20]}{'...' if len(expected_positions) > 20 else ''}")
        
        # Encontrar problemas específicos
        missing_pos = set(expected_positions) - set(actual_positions)
        extra_pos = set(actual_positions) - set(expected_positions)
        
        if missing_pos:
            analysis.append(f"   Posições faltando: {sorted(missing_pos)}")
        if extra_pos:
            analysis.append(f"   Posições extras: {sorted(extra_pos)}")
    else:
        analysis.append("   ✅ Posições sequenciais corretas")
    
    # 4. Verificar cálculo de tier para cada jogador
    analysis.append("\n4. VERIFICANDO CÁLCULO DE TIERS:")
    tier_counts = {}
    incorrect_tiers = []
    
    for player in male_players:
        pos = player['position']
        current_tier = player['tier']
        calculated_tier = get_tier_from_position(pos)
        
        # Contar jogadores por tier
        if calculated_tier not in tier_counts:
            tier_counts[calculated_tier] = 0
        tier_counts[calculated_tier] += 1
        
        # Verificar se o tier está incorreto
        if current_tier != calculated_tier:
            incorrect_tiers.append({
                'name': player['name'],
                'position': pos,
                'current': current_tier,
                'calculated': calculated_tier
            })
    
    analysis.append("   Contagem por tier (baseado no cálculo correto):")
    for tier in sorted(tier_counts.keys()):
        expected_count = len(PYRAMID_STRUCTURE.get(tier, []))
        actual_count = tier_counts[tier]
        status = "✅" if actual_count == expected_count else "❌"
        analysis.append(f"   {tier}: {actual_count} jogadores (esperado: {expected_count}) {status}")
    
    if incorrect_tiers:
        analysis.append(f"\n   ❌ {len(incorrect_tiers)} jogadores com tier incorreto:")
        for player in incorrect_tiers[:10]:  # Mostrar apenas os primeiros 10
            analysis.append(f"   - {player['name']} (pos {player['position']}): {player['current']} → {player['calculated']}")
    else:
        analysis.append("\n   ✅ Todos os jogadores têm tier correto")
    
    # 5. Analisar especificamente o tier J
    analysis.append("\n5. ANÁLISE ESPECÍFICA DO TIER J:")
    tier_j_players = [p for p in male_players if get_tier_from_position(p['position']) == 'J']
    analysis.append(f"   Jogadores que DEVERIAM estar no tier J: {len(tier_j_players)}")
    analysis.append(f"   Posições do tier J na estrutura: {PYRAMID_STRUCTURE['J']}")
    analysis.append(f"   Primeira posição tier J: {min(PYRAMID_STRUCTURE['J'])}")
    analysis.append(f"   Última posição tier J: {max(PYRAMID_STRUCTURE['J'])}")
    
    if tier_j_players:
        analysis.append("   Jogadores no tier J:")
        for player in tier_j_players:
            analysis.append(f"   - {player['name']} (pos {player['position']}, tier no banco: {player['tier']})")
    
    conn.close()
    
    # Retornar análise formatada
    return "<pre>" + "\n".join(analysis) + "</pre>"


# 3. FUNÇÃO DE NORMALIZAÇÃO DE POSIÇÕES
def normalize_male_player_positions():
    """
    Normaliza as posições dos jogadores masculinos para serem sequenciais (1, 2, 3...)
    sem lacunas, mantendo a ordem relativa atual.
    """
    conn = get_db_connection()
    try:
        # Buscar jogadores masculinos ordenados pela posição atual
        male_players = conn.execute('''
            SELECT id, name, position, tier
            FROM players 
            WHERE active = 1 AND (sexo != 'feminino' OR sexo IS NULL)
            ORDER BY position, name
        ''').fetchall()
        
        print(f"Normalizando posições para {len(male_players)} jogadores masculinos...")
        
        changes_made = 0
        
        # Reassignar posições sequenciais
        for i, player in enumerate(male_players, 1):
            new_position = i
            new_tier = get_tier_from_position(new_position)
            
            # Só atualizar se houve mudança
            if player['position'] != new_position or player['tier'] != new_tier:
                print(f"  {player['name']}: pos {player['position']} → {new_position}, tier {player['tier']} → {new_tier}")
                
                conn.execute('''
                    UPDATE players 
                    SET position = ?, tier = ? 
                    WHERE id = ?
                ''', (new_position, new_tier, player['id']))
                
                # Registrar no histórico
                conn.execute('''
                    INSERT INTO ranking_history 
                    (player_id, old_position, new_position, old_tier, new_tier, reason)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (player['id'], player['position'], new_position, player['tier'], new_tier, 'position_normalization'))
                
                changes_made += 1
        
        conn.commit()
        print(f"✅ Normalização concluída: {changes_made} jogadores atualizados")
        print(f"   Posições agora: 1-{len(male_players)} (sequencial)")
        
        return {
            'total_players': len(male_players),
            'changes_made': changes_made,
            'final_range': f"1-{len(male_players)}"
        }
        
    except Exception as e:
        conn.rollback()
        print(f"❌ Erro ao normalizar posições: {e}")
        raise
    finally:
        conn.close()

# 4. FUNÇÃO DE VALIDAÇÃO FINAL
def validate_pyramid_structure():
    """
    Valida se a estrutura da pirâmide está correta após as correções.
    """
    conn = get_db_connection()
    try:
        # Verificar jogadores masculinos
        male_players = conn.execute('''
            SELECT position, tier, name 
            FROM players 
            WHERE active = 1 AND (sexo != 'feminino' OR sexo IS NULL)
            ORDER BY position
        ''').fetchall()
        
        # Contar jogadores por tier
        tier_counts = {}
        position_errors = []
        tier_errors = []
        
        for i, player in enumerate(male_players, 1):
            expected_position = i
            calculated_tier = get_tier_from_position(player['position'])
            
            # Verificar posição
            if player['position'] != expected_position:
                position_errors.append(f"{player['name']}: pos {player['position']} (esperado {expected_position})")
            
            # Verificar tier
            if player['tier'] != calculated_tier:
                tier_errors.append(f"{player['name']}: tier {player['tier']} (esperado {calculated_tier})")
            
            # Contar por tier
            tier = player['tier']
            tier_counts[tier] = tier_counts.get(tier, 0) + 1
        
        # Verificar contagens por tier
        tier_count_errors = []
        for tier, expected_positions in PYRAMID_STRUCTURE.items():
            expected_count = len(expected_positions)
            actual_count = tier_counts.get(tier, 0)
            
            if actual_count != expected_count:
                tier_count_errors.append(f"Tier {tier}: {actual_count} jogadores (esperado {expected_count})")
        
        # Montar resultado
        result = {
            'total_players': len(male_players),
            'position_errors': position_errors,
            'tier_errors': tier_errors,
            'tier_count_errors': tier_count_errors,
            'tier_counts': tier_counts,
            'is_valid': len(position_errors) == 0 and len(tier_errors) == 0 and len(tier_count_errors) == 0
        }
        
        return result
        
    except Exception as e:
        print(f"❌ Erro na validação: {e}")
        raise
    finally:
        conn.close()

# 5. ROTA PRINCIPAL - EXECUTA A SOLUÇÃO HÍBRIDA COMPLETA
@app.route('/fix_pyramid_hybrid')
@login_required
def fix_pyramid_hybrid():
    """
    Executa a solução híbrida completa:
    1. Normaliza posições sequenciais
    2. Recalcula tiers com estrutura estendida
    3. Valida resultado final
    """
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    try:
        # Passo 1: Normalizar posições
        print("=== INICIANDO SOLUÇÃO HÍBRIDA ===")
        normalization_result = normalize_male_player_positions()
        
        # Passo 2: Validar resultado
        print("\n=== VALIDANDO RESULTADO ===")
        validation_result = validate_pyramid_structure()
        
        # Passo 3: Auto-corrigir ranking feminino também
        print("\n=== CORRIGINDO RANKING FEMININO ===")
        auto_fix_female_ranking()
        
        # Passo 4: Sincronizar histórico
        print("\n=== SINCRONIZANDO HISTÓRICO ===")
        sync_ranking_history_tables()
        
        # Mostrar resultado
        if validation_result['is_valid']:
            message = f"""
            ✅ SOLUÇÃO HÍBRIDA CONCLUÍDA COM SUCESSO!
            
            📊 Resultados:
            • {normalization_result['total_players']} jogadores masculinos processados
            • {normalization_result['changes_made']} jogadores tiveram posições/tiers atualizados
            • Posições agora: {normalization_result['final_range']} (sequencial)
            • Tier J: {validation_result['tier_counts'].get('J', 0)} jogadores
            • Tier K: {validation_result['tier_counts'].get('K', 0)} jogadores
            
            🏆 A pirâmide agora está perfeitamente estruturada!
            """
            flash(message, 'success')
        else:
            error_details = []
            if validation_result['position_errors']:
                error_details.append(f"Posições incorretas: {len(validation_result['position_errors'])}")
            if validation_result['tier_errors']:
                error_details.append(f"Tiers incorretos: {len(validation_result['tier_errors'])}")
            if validation_result['tier_count_errors']:
                error_details.append(f"Contagens incorretas: {len(validation_result['tier_count_errors'])}")
            
            flash(f"⚠️ Correção parcial. Problemas restantes: {', '.join(error_details)}", 'warning')
        
        print("\n=== SOLUÇÃO HÍBRIDA CONCLUÍDA ===")
        
    except Exception as e:
        flash(f'❌ Erro na solução híbrida: {str(e)}', 'error')
        print(f"❌ ERRO: {e}")
    
    return redirect(url_for('pyramid_dynamic'))

# 6. ROTA DE VALIDAÇÃO (para verificar o resultado)
@app.route('/validate_pyramid')
@login_required
def validate_pyramid_route():
    """
    Valida a estrutura atual da pirâmide e mostra relatório detalhado.
    """
    if not session.get('is_admin', False):
        return "Acesso negado"
    
    try:
        result = validate_pyramid_structure()
        
        report = ["=== RELATÓRIO DE VALIDAÇÃO DA PIRÂMIDE ===\n"]
        
        report.append(f"Total de jogadores masculinos: {result['total_players']}")
        report.append(f"Status geral: {'✅ VÁLIDA' if result['is_valid'] else '❌ PROBLEMAS DETECTADOS'}\n")
        
        # Contagem por tier
        report.append("Contagem por tier:")
        for tier in sorted(result['tier_counts'].keys()):
            count = result['tier_counts'][tier]
            expected = len(PYRAMID_STRUCTURE.get(tier, []))
            status = "✅" if count == expected else "❌"
            report.append(f"  Tier {tier}: {count} jogadores (esperado: {expected}) {status}")
        
        # Erros de posição
        if result['position_errors']:
            report.append(f"\n❌ Erros de posição ({len(result['position_errors'])}):")
            for error in result['position_errors'][:10]:  # Mostrar apenas os primeiros 10
                report.append(f"  {error}")
        
        # Erros de tier
        if result['tier_errors']:
            report.append(f"\n❌ Erros de tier ({len(result['tier_errors'])}):")
            for error in result['tier_errors'][:10]:
                report.append(f"  {error}")
        
        # Erros de contagem
        if result['tier_count_errors']:
            report.append(f"\n❌ Erros de contagem por tier:")
            for error in result['tier_count_errors']:
                report.append(f"  {error}")
        
        if result['is_valid']:
            report.append("\n🎉 A pirâmide está perfeitamente estruturada!")
        
        return "<pre>" + "\n".join(report) + "</pre>"
        
    except Exception as e:
        return f"<pre>❌ Erro na validação: {str(e)}</pre>"


def create_player_result_setting():
    conn = get_db_connection()
    conn.execute('''
    INSERT OR IGNORE INTO system_settings (key, value)
    VALUES ('players_can_submit_results', 'true')
    ''')
    conn.commit()
    conn.close()
    print("Configuração de submissão de resultados por jogadores criada/verificada.")


@app.route('/admin/toggle_player_results', methods=['GET', 'POST'])
@login_required
def toggle_player_results():
    # Verificar se é um administrador
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    
    if request.method == 'POST':
        action = request.form.get('action')
        
        # Verificar senha do admin
        if not session.get('is_admin', False):
            conn.close()
            flash('Acesso negado. Apenas administradores podem executar esta ação!', 'error')
            return redirect(url_for('toggle_player_results'))
        
        if action == 'enable':
            conn.execute('UPDATE system_settings SET value = ?, updated_at = CURRENT_TIMESTAMP WHERE key = ?', 
                       ('true', 'players_can_submit_results'))
            conn.commit()
            flash('Jogadores agora PODEM submeter resultados de desafios!', 'success')
        elif action == 'disable':
            conn.execute('UPDATE system_settings SET value = ?, updated_at = CURRENT_TIMESTAMP WHERE key = ?', 
                       ('false', 'players_can_submit_results'))
            conn.commit()
            flash('Jogadores agora NÃO PODEM submeter resultados de desafios!', 'success')
    
    # Obter status atual
    setting = conn.execute('SELECT value, updated_at FROM system_settings WHERE key = ?', 
                          ('players_can_submit_results',)).fetchone()
    is_enabled = setting and setting['value'] == 'true'
    updated_at = setting['updated_at'] if setting else None
    
    conn.close()
    
    return render_template('toggle_player_results.html', is_enabled=is_enabled, updated_at=updated_at)


@app.route('/migrate_tiers_to_new_structure')
@login_required
def migrate_tiers_to_new_structure():
    """
    Migra todos os jogadores para a nova estrutura de tiers (A, B, C... ao invés de C, D, E...)
    """
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    conn = get_db_connection()
    try:
        # Buscar todos os jogadores ativos ordenados por posição
        players = conn.execute('''
            SELECT id, name, position, tier 
            FROM players 
            WHERE active = 1
            ORDER BY position
        ''').fetchall()
        
        updated_count = 0
        
        for player in players:
            # Calcular o tier correto baseado na posição
            correct_tier = get_tier_from_position(player['position'])
            
            # Atualizar se necessário
            if player['tier'] != correct_tier:
                conn.execute('''
                    UPDATE players SET tier = ? WHERE id = ?
                ''', (correct_tier, player['id']))
                updated_count += 1
                print(f"Atualizado: {player['name']} - Posição {player['position']}: {player['tier']} → {correct_tier}")
        
        conn.commit()
        flash(f'✅ Migração concluída! {updated_count} jogadores tiveram seus tiers atualizados.', 'success')
        
    except Exception as e:
        conn.rollback()
        flash(f'❌ Erro na migração: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('pyramid_dynamic'))


# ADICIONE ESTA ROTA NO app.py
# Certifique-se de ter o openpyxl instalado: pip install openpyxl

from io import BytesIO
from datetime import datetime


# ADICIONE ESTA ROTA NO app.py
# Certifique-se de ter o openpyxl instalado: pip install openpyxl

from io import BytesIO
from datetime import datetime

@app.route('/export_ranking_excel')
def export_ranking_excel():
    """Exporta o ranking de jogadores para Excel"""
    try:
        import openpyxl
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    except ImportError:
        flash('Erro: biblioteca openpyxl não instalada.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    
    # Buscar jogadores ativos ordenados por posição
    players = conn.execute('''
        SELECT name, position, player_code, sexo 
        FROM players 
        WHERE active = 1 
        ORDER BY position
    ''').fetchall()
    
    conn.close()
    
    # Criar workbook
    wb = openpyxl.Workbook()
    
    # Estilos
    header_font = Font(bold=True, color="FFFFFF", size=12)
    header_alignment = Alignment(horizontal="center", vertical="center")
    cell_alignment = Alignment(horizontal="center", vertical="center")
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )
    
    # ==================== ABA MASCULINO ====================
    ws_masc = wb.active
    ws_masc.title = "Ranking Masculino"
    
    header_fill_masc = PatternFill(start_color="002970", end_color="002970", fill_type="solid")
    
    # Título
    ws_masc.merge_cells('A1:C1')
    ws_masc['A1'] = f"Ranking Masculino - Liga Olímpica de Golfe - {datetime.now().strftime('%d/%m/%Y')}"
    ws_masc['A1'].font = Font(bold=True, size=14, color="002970")
    ws_masc['A1'].alignment = Alignment(horizontal="center")
    ws_masc.row_dimensions[1].height = 30
    
    # Cabeçalhos
    headers = ["Posição", "Código", "Nome"]
    for col, header in enumerate(headers, 1):
        cell = ws_masc.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_masc
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Dados masculinos
    row = 4
    position_counter = 0
    for player in players:
        if player['sexo'] != 'feminino':
            position_counter += 1
            ws_masc.cell(row=row, column=1, value=position_counter).alignment = cell_alignment
            ws_masc.cell(row=row, column=2, value=player['player_code']).alignment = cell_alignment
            ws_masc.cell(row=row, column=3, value=player['name']).alignment = Alignment(horizontal="left", vertical="center")
            
            # Aplicar bordas
            for col in range(1, 4):
                ws_masc.cell(row=row, column=col).border = thin_border
            
            row += 1
    
    # Ajustar largura das colunas
    ws_masc.column_dimensions['A'].width = 10  # Posição
    ws_masc.column_dimensions['B'].width = 12  # Código
    ws_masc.column_dimensions['C'].width = 35  # Nome
    
    # ==================== ABA FEMININO ====================
    ws_fem = wb.create_sheet("Ranking Ladies")
    
    header_fill_ladies = PatternFill(start_color="E91E63", end_color="E91E63", fill_type="solid")
    
    # Título
    ws_fem.merge_cells('A1:C1')
    ws_fem['A1'] = f"Ranking Ladies - Liga Olímpica de Golfe - {datetime.now().strftime('%d/%m/%Y')}"
    ws_fem['A1'].font = Font(bold=True, size=14, color="E91E63")
    ws_fem['A1'].alignment = Alignment(horizontal="center")
    ws_fem.row_dimensions[1].height = 30
    
    # Cabeçalhos
    for col, header in enumerate(headers, 1):
        cell = ws_fem.cell(row=3, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill_ladies
        cell.alignment = header_alignment
        cell.border = thin_border
    
    # Dados femininos
    row = 4
    position_counter = 0
    for player in players:
        if player['sexo'] == 'feminino':
            position_counter += 1
            ws_fem.cell(row=row, column=1, value=position_counter).alignment = cell_alignment
            ws_fem.cell(row=row, column=2, value=player['player_code']).alignment = cell_alignment
            ws_fem.cell(row=row, column=3, value=player['name']).alignment = Alignment(horizontal="left", vertical="center")
            
            # Aplicar bordas
            for col in range(1, 4):
                ws_fem.cell(row=row, column=col).border = thin_border
            
            row += 1
    
    # Ajustar largura das colunas
    ws_fem.column_dimensions['A'].width = 10
    ws_fem.column_dimensions['B'].width = 12
    ws_fem.column_dimensions['C'].width = 35
    
    # Salvar em memória
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    # Nome do arquivo com data
    filename = f"ranking_golf_{datetime.now().strftime('%Y%m%d_%H%M')}.xlsx"
    
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# ADICIONE ESTA ROTA NO app.py (área de rotas admin)

# ADICIONE ESTA ROTA NO app.py (área de rotas admin)

# ADICIONE ESTA ROTA NO app.py (área de rotas admin)

@app.route('/admin/reset_challenges', methods=['GET', 'POST'])
@login_required
def reset_challenges():
    """
    Reseta todos os desafios e histórico para iniciar uma nova etapa do ranking.
    Arquiva os desafios e histórico antigos e limpa as tabelas.
    """
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('dashboard'))
    
    if request.method == 'POST':
        # Confirmação de segurança
        confirm_text = request.form.get('confirm_text', '')
        if confirm_text != 'RESETAR':
            flash('❌ Texto de confirmação incorreto. Digite RESETAR para confirmar.', 'error')
            return redirect(url_for('reset_challenges'))
        
        conn = get_db_connection()
        try:
            # Contar desafios antes de resetar
            stats = conn.execute('''
                SELECT 
                    COUNT(*) as total,
                    SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending,
                    SUM(CASE WHEN status = 'accepted' THEN 1 ELSE 0 END) as accepted,
                    SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed,
                    SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled
                FROM challenges
            ''').fetchone()
            
            # Contar registros de histórico
            history_count = 0
            try:
                history_result = conn.execute('SELECT COUNT(*) as total FROM ranking_history').fetchone()
                history_count = history_result['total'] if history_result else 0
            except:
                pass  # Tabela pode não existir
            
            # Verificar colunas existentes na tabela challenges
            columns_info = conn.execute("PRAGMA table_info(challenges)").fetchall()
            existing_columns = [col[1] for col in columns_info]
            
            # Criar tabela de histórico de desafios se não existir
            conn.execute('''
                CREATE TABLE IF NOT EXISTS challenges_history (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    original_id INTEGER,
                    challenger_id INTEGER,
                    challenged_id INTEGER,
                    scheduled_date TEXT,
                    status TEXT,
                    winner_id INTEGER,
                    challenger_score TEXT,
                    challenged_score TEXT,
                    notes TEXT,
                    created_at TEXT,
                    archived_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    etapa TEXT
                )
            ''')
            
            # Criar tabela de histórico de ranking arquivado se não existir
            conn.execute('''
                CREATE TABLE IF NOT EXISTS ranking_history_archive (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    original_id INTEGER,
                    player_id INTEGER,
                    position INTEGER,
                    tier TEXT,
                    record_date TEXT,
                    archived_at TEXT DEFAULT CURRENT_TIMESTAMP,
                    etapa TEXT
                )
            ''')
            
            # Definir nome da etapa
            etapa_name = request.form.get('etapa_name', f"Etapa até {datetime.now().strftime('%d/%m/%Y')}")
            
            # ==================== ARQUIVAR DESAFIOS ====================
            desired_columns = [
                'id', 'challenger_id', 'challenged_id', 'scheduled_date',
                'status', 'winner_id', 'challenger_score', 'challenged_score', 
                'notes', 'created_at'
            ]
            
            columns_to_copy = [col for col in desired_columns if col in existing_columns]
            
            if columns_to_copy:
                dest_columns = ['original_id' if col == 'id' else col for col in columns_to_copy]
                dest_columns.append('etapa')
                
                source_columns = columns_to_copy.copy()
                
                insert_query = f'''
                    INSERT INTO challenges_history ({', '.join(dest_columns)})
                    SELECT {', '.join(source_columns)}, ?
                    FROM challenges
                '''
                
                conn.execute(insert_query, (etapa_name,))
            
            # Limpar tabela de desafios
            conn.execute('DELETE FROM challenges')
            
            # Resetar o autoincrement
            conn.execute("DELETE FROM sqlite_sequence WHERE name='challenges'")
            
            # ==================== ARQUIVAR HISTÓRICO DO RANKING ====================
            try:
                # Verificar se a tabela ranking_history existe
                table_exists = conn.execute('''
                    SELECT name FROM sqlite_master 
                    WHERE type='table' AND name='ranking_history'
                ''').fetchone()
                
                if table_exists:
                    # Verificar colunas da tabela ranking_history
                    history_columns_info = conn.execute("PRAGMA table_info(ranking_history)").fetchall()
                    history_existing_columns = [col[1] for col in history_columns_info]
                    
                    # Colunas desejadas para copiar
                    history_desired_columns = ['id', 'player_id', 'position', 'tier', 'record_date']
                    history_columns_to_copy = [col for col in history_desired_columns if col in history_existing_columns]
                    
                    if history_columns_to_copy:
                        hist_dest_columns = ['original_id' if col == 'id' else col for col in history_columns_to_copy]
                        hist_dest_columns.append('etapa')
                        
                        hist_source_columns = history_columns_to_copy.copy()
                        
                        hist_insert_query = f'''
                            INSERT INTO ranking_history_archive ({', '.join(hist_dest_columns)})
                            SELECT {', '.join(hist_source_columns)}, ?
                            FROM ranking_history
                        '''
                        
                        conn.execute(hist_insert_query, (etapa_name,))
                    
                    # Limpar tabela de histórico
                    conn.execute('DELETE FROM ranking_history')
                    
                    # Resetar o autoincrement
                    conn.execute("DELETE FROM sqlite_sequence WHERE name='ranking_history'")
            except Exception as e:
                print(f"Aviso ao processar ranking_history: {e}")
            
            conn.commit()
            
            total = stats['total'] or 0
            pending = stats['pending'] or 0
            accepted = stats['accepted'] or 0
            completed = stats['completed'] or 0
            cancelled = stats['cancelled'] or 0
            
            flash(f'✅ Nova etapa iniciada! {total} desafios e {history_count} registros de histórico foram arquivados. '
                  f'(Desafios - Pendentes: {pending}, Aceitos: {accepted}, '
                  f'Concluídos: {completed}, Cancelados: {cancelled})', 'success')
            
        except Exception as e:
            conn.rollback()
            flash(f'❌ Erro ao resetar desafios: {str(e)}', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('dashboard'))
    
    # GET - Mostrar página de confirmação
    conn = get_db_connection()
    stats = conn.execute('''
        SELECT 
            COUNT(*) as total,
            SUM(CASE WHEN status = 'pending' THEN 1 ELSE 0 END) as pending,
            SUM(CASE WHEN status = 'accepted' THEN 1 ELSE 0 END) as accepted,
            SUM(CASE WHEN status = 'completed' THEN 1 ELSE 0 END) as completed,
            SUM(CASE WHEN status = 'cancelled' THEN 1 ELSE 0 END) as cancelled
        FROM challenges
    ''').fetchone()
    
    # Contar histórico
    history_count = 0
    try:
        history_result = conn.execute('SELECT COUNT(*) as total FROM ranking_history').fetchone()
        history_count = history_result['total'] if history_result else 0
    except:
        pass
    
    conn.close()
    
    return render_template('admin_reset_challenges.html', stats=stats, history_count=history_count)


# IMPORTANTE: Adicione também o import no topo do app.py:
# from flask import send_file



# ============================================================
# FUNÇÃO CORRIGIDA - adjust_player_position
# 
# PROBLEMA: Quando sexo = NULL ou '', a query "sexo = ?" não 
# funciona corretamente, limitando a posição máxima incorretamente.
#
# SOLUÇÃO: Tratar masculino como (sexo = 'masculino' OR sexo IS NULL OR sexo = '')
#
# Substitua a função existente no app.py (aproximadamente linha 5765)
# ============================================================

@app.route('/admin/adjust_position/<int:player_id>', methods=['GET', 'POST'])
@login_required
def adjust_player_position(player_id):
    """
    Permite ao admin ajustar manualmente a posição de um jogador no ranking.
    CORRIGIDO: Trata corretamente jogadores com sexo NULL ou vazio.
    """
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    
    # Buscar jogador
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('❌ Jogador não encontrado.', 'error')
        return redirect(url_for('index'))
    
    if request.method == 'POST':
        try:
            new_position = int(request.form.get('new_position'))
            reason = request.form.get('reason', '').strip()
            
            if new_position < 1:
                flash('❌ A posição deve ser maior que zero.', 'error')
                return redirect(url_for('adjust_player_position', player_id=player_id))
            
            old_position = player['position']
            player_sexo = player['sexo']
            
            if new_position == old_position:
                flash('ℹ️ A nova posição é igual à posição atual. Nenhuma alteração feita.', 'info')
                return redirect(url_for('player_detail', player_id=player_id))
            
            # ============================================================
            # CORREÇÃO PRINCIPAL: Query que trata corretamente sexo NULL/vazio
            # ============================================================
            if player_sexo == 'feminino':
                # Jogadoras femininas
                players_same_gender = conn.execute('''
                    SELECT id, position FROM players 
                    WHERE active = 1 AND sexo = 'feminino' AND id != ?
                    ORDER BY position
                ''', (player_id,)).fetchall()
                gender_condition = "sexo = 'feminino'"
            else:
                # Jogadores masculinos (inclui NULL e string vazia)
                players_same_gender = conn.execute('''
                    SELECT id, position FROM players 
                    WHERE active = 1 
                    AND (sexo = 'masculino' OR sexo IS NULL OR sexo = '')
                    AND id != ?
                    ORDER BY position
                ''', (player_id,)).fetchall()
                gender_condition = "(sexo = 'masculino' OR sexo IS NULL OR sexo = '')"
            
            # Verificar posição máxima
            max_position = len(players_same_gender) + 1
            
            # DEBUG: Mostrar informações (remover após confirmar que funciona)
            print(f"DEBUG: Jogador {player['name']}, sexo='{player_sexo}'")
            print(f"DEBUG: Total de jogadores do mesmo gênero: {len(players_same_gender)}")
            print(f"DEBUG: Posição máxima permitida: {max_position}")
            print(f"DEBUG: Nova posição solicitada: {new_position}")
            
            if new_position > max_position:
                flash(f'⚠️ Posição {new_position} excede o máximo ({max_position}). Ajustando para {max_position}.', 'warning')
                new_position = max_position
            
            # Reorganizar posições usando a condição correta
            if new_position < old_position:
                # Jogador subiu no ranking - empurrar outros para baixo
                conn.execute(f'''
                    UPDATE players 
                    SET position = position + 1 
                    WHERE active = 1 AND {gender_condition} 
                    AND position >= ? AND position < ? AND id != ?
                ''', (new_position, old_position, player_id))
            else:
                # Jogador desceu no ranking - puxar outros para cima
                conn.execute(f'''
                    UPDATE players 
                    SET position = position - 1 
                    WHERE active = 1 AND {gender_condition}
                    AND position > ? AND position <= ? AND id != ?
                ''', (old_position, new_position, player_id))
            
            # Atualizar posição do jogador
            new_tier = get_tier_from_position(new_position)
            conn.execute('''
                UPDATE players SET position = ?, tier = ? WHERE id = ?
            ''', (new_position, new_tier, player_id))
            
            # Atualizar tiers de TODOS os jogadores afetados (usando condição correta)
            if player_sexo == 'feminino':
                affected_players = conn.execute('''
                    SELECT id, position FROM players 
                    WHERE active = 1 AND sexo = 'feminino'
                    ORDER BY position
                ''').fetchall()
            else:
                affected_players = conn.execute('''
                    SELECT id, position FROM players 
                    WHERE active = 1 AND (sexo = 'masculino' OR sexo IS NULL OR sexo = '')
                    ORDER BY position
                ''').fetchall()
            
            for p in affected_players:
                correct_tier = get_tier_from_position(p['position'])
                conn.execute('''
                    UPDATE players SET tier = ? WHERE id = ?
                ''', (correct_tier, p['id']))
            
            # Registrar no histórico
            try:
                conn.execute('''
                    INSERT INTO ranking_history 
                    (player_id, old_position, new_position, old_tier, new_tier, reason)
                    VALUES (?, ?, ?, ?, ?, ?)
                ''', (player_id, old_position, new_position, 
                      get_tier_from_position(old_position), new_tier, 
                      f'admin_manual_adjust: {reason}' if reason else 'admin_manual_adjust'))
            except Exception as e:
                print(f"Aviso: Não foi possível registrar no histórico: {e}")
            
            conn.commit()
            
            flash(f'✅ Posição de {player["name"]} alterada de #{old_position} para #{new_position}. Tiers atualizados.', 'success')
            
            return redirect(url_for('player_detail', player_id=player_id))
            
        except ValueError:
            flash('❌ Posição inválida. Digite um número.', 'error')
        except Exception as e:
            conn.rollback()
            flash(f'❌ Erro ao ajustar posição: {str(e)}', 'error')
        finally:
            conn.close()
        
        return redirect(url_for('adjust_player_position', player_id=player_id))
    
    # GET - Mostrar formulário
    # Buscar jogadores do mesmo sexo para mostrar contexto (usando query correta)
    if player['sexo'] == 'feminino':
        players_same_gender = conn.execute('''
            SELECT id, name, position, tier FROM players 
            WHERE active = 1 AND sexo = 'feminino'
            ORDER BY position
        ''').fetchall()
    else:
        players_same_gender = conn.execute('''
            SELECT id, name, position, tier FROM players 
            WHERE active = 1 AND (sexo = 'masculino' OR sexo IS NULL OR sexo = '')
            ORDER BY position
        ''').fetchall()
    
    conn.close()
    
    return render_template('admin_adjust_position.html', 
                           player=player, 
                           players_list=players_same_gender)



# ============================================================
# ROTA PARA ADMIN ATIVAR/DESATIVAR CARTEIRINHA
# ============================================================

@app.route('/admin/toggle-carteirinha/<int:player_id>', methods=['POST'])
@login_required
def toggle_carteirinha_forcada(player_id):
    """Admin pode forçar carteirinha como ativa/inativa"""
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    try:
        player = conn.execute('SELECT name, carteirinha_forcada FROM players WHERE id = ?', 
                              (player_id,)).fetchone()
        
        if not player:
            flash('❌ Jogador não encontrado.', 'error')
            conn.close()
            return redirect(url_for('index'))
        
        # Toggle: se está 1 vira 0, se está 0 ou NULL vira 1
        novo_valor = 0 if player['carteirinha_forcada'] == 1 else 1
        
        conn.execute('''
            UPDATE players SET carteirinha_forcada = ? WHERE id = ?
        ''', (novo_valor, player_id))
        conn.commit()
        
        if novo_valor == 1:
            flash(f'✅ Carteirinha de {player["name"]} FORÇADA COMO ATIVA pelo admin.', 'success')
        else:
            flash(f'ℹ️ Carteirinha de {player["name"]} voltou ao modo automático.', 'info')
        
    except Exception as e:
        flash(f'❌ Erro: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))



"""
============================================================
CARTEIRINHA DIGITAL - Implementação Completa
============================================================

Adicione este código ao seu app.py

============================================================
"""

import secrets
from datetime import datetime, timedelta

# ============================================================
# PASSO 1: Função para criar a tabela de tokens
# ============================================================

def create_verification_tokens_table():
    """Cria a tabela para armazenar tokens de verificação da carteirinha"""
    conn = get_db_connection()
    
    conn.execute('''
        CREATE TABLE IF NOT EXISTS verification_tokens (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            player_id INTEGER NOT NULL,
            token TEXT NOT NULL UNIQUE,
            created_at DATETIME DEFAULT CURRENT_TIMESTAMP,
            expires_at DATETIME NOT NULL,
            used_at DATETIME,
            used_by_business TEXT,
            FOREIGN KEY (player_id) REFERENCES players(id)
        )
    ''')
    
    conn.execute('CREATE INDEX IF NOT EXISTS idx_token ON verification_tokens(token)')
    
    conn.commit()
    conn.close()
    print("Tabela 'verification_tokens' verificada/criada.")


# ============================================================
# PASSO 2: Funções auxiliares para tokens
# ============================================================

def parse_datetime(dt_value):
    """Converte string para datetime se necessário"""
    if dt_value is None:
        return None
    if isinstance(dt_value, datetime):
        return dt_value
    if isinstance(dt_value, str):
        for fmt in ['%Y-%m-%d %H:%M:%S.%f', '%Y-%m-%d %H:%M:%S', '%Y-%m-%dT%H:%M:%S.%f', '%Y-%m-%dT%H:%M:%S']:
            try:
                return datetime.strptime(dt_value, fmt)
            except ValueError:
                continue
    return None


def generate_verification_token(player_id, validity_minutes=10):
    """Gera um token de verificação temporário para a carteirinha."""
    conn = get_db_connection()
    
    # Limpar tokens expirados do jogador
    conn.execute('''
        DELETE FROM verification_tokens 
        WHERE player_id = ? AND expires_at < ?
    ''', (player_id, datetime.now()))
    
    # Verificar se já existe um token válido
    existing = conn.execute('''
        SELECT token, expires_at FROM verification_tokens 
        WHERE player_id = ? AND expires_at > ? AND used_at IS NULL
        ORDER BY expires_at DESC LIMIT 1
    ''', (player_id, datetime.now())).fetchone()
    
    if existing:
        conn.close()
        expires = parse_datetime(existing['expires_at'])
        if expires is None:
            expires = datetime.now() + timedelta(minutes=validity_minutes)
        return {
            'token': existing['token'],
            'expires_at': expires
        }
    
    # Gerar novo token
    token = secrets.token_urlsafe(32)
    expires_at = datetime.now() + timedelta(minutes=validity_minutes)
    
    conn.execute('''
        INSERT INTO verification_tokens (player_id, token, expires_at)
        VALUES (?, ?, ?)
    ''', (player_id, token, expires_at))
    
    conn.commit()
    conn.close()
    
    return {
        'token': token,
        'expires_at': expires_at
    }


def validate_verification_token(token):
    """Valida um token de verificação."""
    conn = get_db_connection()
    
    result = conn.execute('''
        SELECT vt.*, 
               p.id as player_id, 
               p.name, 
               p.player_code,
               p.position, 
               p.tier, 
               p.country, 
               p.profile_photo, 
               p.active, 
               p.created_at as member_since
        FROM verification_tokens vt
        JOIN players p ON vt.player_id = p.id
        WHERE vt.token = ? AND vt.expires_at > ?
    ''', (token, datetime.now())).fetchone()
    
    conn.close()
    
    if result:
        return dict(result)
    return None


# ============================================================
# PASSO 3: Rota da Carteirinha (para o jogador)
# ============================================================

@app.route('/carteirinha')
@login_required
def carteirinha():
    """Exibe a carteirinha digital do jogador logado"""
    user_id = session.get('user_id')
    
    if not user_id:
        flash('Você precisa estar logado para acessar sua carteirinha.', 'error')
        return redirect(url_for('login'))
    
    conn = get_db_connection()
    player = conn.execute('SELECT * FROM players WHERE id = ?', (user_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado.', 'error')
        return redirect(url_for('dashboard'))
    
    if not player['active']:
        conn.close()
        flash('Sua carteirinha está inativa. Entre em contato com a administração.', 'warning')
        return redirect(url_for('dashboard'))
    
    # Verificar atividade (agora retorna 3 valores)
    is_card_active, is_forced, is_vip = check_player_activity(conn, user_id)
    
    # Buscar último jogo (não aplicável para VIP)
    last_game = None
    last_game_date = None
    if player['tipo_membro'] != 'vip':
        last_game = conn.execute('''
            SELECT scheduled_date FROM challenges 
            WHERE (challenger_id = ? OR challenged_id = ?)
            AND status = 'completed'
            ORDER BY scheduled_date DESC
            LIMIT 1
        ''', (user_id, user_id)).fetchone()
        last_game_date = last_game['scheduled_date'] if last_game else None
    
    # Gerar token de verificação
    token_data = generate_verification_token(user_id, validity_minutes=10)
    verification_url = url_for('verificar_carteirinha', token=token_data['token'], _external=True)
    
    expires_at = token_data['expires_at']
    if not isinstance(expires_at, datetime):
        expires_at = parse_datetime(expires_at)
    if expires_at is None:
        expires_at = datetime.now() + timedelta(minutes=10)
    
    seconds_remaining = int((expires_at - datetime.now()).total_seconds())
    seconds_remaining = max(0, seconds_remaining)
    
    conn.close()
    
    return render_template('carteirinha.html', 
                          player=player,
                          token=token_data['token'],
                          expires_at=expires_at,
                          seconds_remaining=seconds_remaining,
                          verification_url=verification_url,
                          is_card_active=is_card_active,
                          is_forced=is_forced,
                          is_vip=is_vip,
                          last_game_date=last_game_date)


@app.route('/carteirinha/renovar', methods=['POST'])
@login_required
def renovar_carteirinha():
    """Renova o token da carteirinha (gera um novo)"""
    user_id = session.get('user_id')
    
    if not user_id:
        return {'error': 'Não autorizado'}, 401
    
    conn = get_db_connection()
    conn.execute('DELETE FROM verification_tokens WHERE player_id = ?', (user_id,))
    conn.commit()
    conn.close()
    
    # Gerar novo token
    token_data = generate_verification_token(user_id, validity_minutes=10)
    verification_url = url_for('verificar_carteirinha', token=token_data['token'], _external=True)
    
    # Calcular segundos restantes
    expires_at = token_data['expires_at']
    if isinstance(expires_at, datetime):
        seconds_remaining = int((expires_at - datetime.now()).total_seconds())
    else:
        seconds_remaining = 600
    
    return {
        'token': token_data['token'],
        'expires_at': expires_at.isoformat() if isinstance(expires_at, datetime) else str(expires_at),
        'seconds_remaining': max(0, seconds_remaining),
        'verification_url': verification_url
    }


# ============================================================
# PASSO 4: Rota de Verificação (para o estabelecimento)
# ============================================================

@app.route('/verificar/<token>')
def verificar_carteirinha(token):
    """Página pública para verificação da carteirinha."""
    import traceback
    
    try:
        verified_at = datetime.now()
        result = validate_verification_token(token)
        
        if not result:
            return render_template('verificar_carteirinha.html', 
                                  valid=False,
                                  error='Token inválido ou expirado',
                                  verified_at=verified_at)
        
        conn = get_db_connection()
        
        # Buscar tipo de membro
        player_info = conn.execute('SELECT tipo_membro FROM players WHERE id = ?', 
                                   (result['player_id'],)).fetchone()
        is_vip = player_info and player_info['tipo_membro'] == 'vip'
        
        # Verificar atividade
        is_card_active, is_forced, _ = check_player_activity(conn, result['player_id'])
        
        # Buscar último jogo (apenas para jogadores normais)
        last_game_date = None
        if not is_vip:
            last_game = conn.execute('''
                SELECT scheduled_date FROM challenges 
                WHERE (challenger_id = ? OR challenged_id = ?)
                AND status = 'completed'
                ORDER BY scheduled_date DESC
                LIMIT 1
            ''', (result['player_id'], result['player_id'])).fetchone()
            last_game_date = last_game['scheduled_date'] if last_game else None
        
        conn.close()
        
        # Calcular tempo restante
        expires_at = parse_datetime(result.get('expires_at'))
        if expires_at:
            total_seconds = max(0, (expires_at - datetime.now()).total_seconds())
            minutes_remaining = int(total_seconds // 60)
            seconds_remaining = int(total_seconds % 60)
        else:
            minutes_remaining = 0
            seconds_remaining = 0
        
        return render_template('verificar_carteirinha.html',
                              valid=True,
                              player=result,
                              verified_at=verified_at,
                              minutes_remaining=minutes_remaining,
                              seconds_remaining=seconds_remaining,
                              is_card_active=is_card_active,
                              is_forced=is_forced,
                              is_vip=is_vip,
                              last_game_date=last_game_date)
    
    except Exception as e:
        error_details = traceback.format_exc()
        return f"<h1>Erro na verificação</h1><p>{str(e)}</p>", 500


# ============================================================
# SISTEMA DE CONTROLE MANUAL DA CARTEIRINHA
# 
# 1. Execute a rota /criar-campo-carteirinha-manual para criar o campo
# 2. Substitua a função check_player_activity
# 3. Adicione a rota de toggle
# 4. Adicione o botão no template player_detail.html
# ============================================================

# ============================================================
# PASSO 1: Rota para criar o campo na tabela (execute uma vez)
# ============================================================

@app.route('/criar-campo-carteirinha-manual')
@login_required
def criar_campo_carteirinha_manual():
    """Cria o campo para controle manual da carteirinha"""
    if not session.get('is_admin', False):
        flash('Acesso restrito.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    try:
        # Verificar se o campo já existe
        cursor = conn.execute("PRAGMA table_info(players)")
        columns = [row[1] for row in cursor.fetchall()]
        
        if 'carteirinha_forcada' not in columns:
            conn.execute('''
                ALTER TABLE players 
                ADD COLUMN carteirinha_forcada INTEGER DEFAULT 0
            ''')
            conn.commit()
            flash('✅ Campo carteirinha_forcada criado com sucesso!', 'success')
        else:
            flash('ℹ️ Campo já existe.', 'info')
    except Exception as e:
        flash(f'❌ Erro: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))


def check_player_activity(conn, player_id):
    """
    Verifica se o jogador tem carteirinha ativa.
    Retorna tupla: (is_active, is_forced, is_vip)
    """
    try:
        player = conn.execute('''
            SELECT carteirinha_forcada, tipo_membro FROM players WHERE id = ?
        ''', (player_id,)).fetchone()
        
        # Membro VIP sempre tem carteirinha ativa
        if player and player['tipo_membro'] == 'vip':
            return True, False, True
        
        # Admin forçou como ativa
        if player and player['carteirinha_forcada'] == 1:
            return True, True, False
    except:
        pass
    
    # Verificar atividade normal (jogou nos últimos 30 dias)
    result = conn.execute('''
        SELECT COUNT(*) as count FROM challenges 
        WHERE (challenger_id = ? OR challenged_id = ?)
        AND status = 'completed'
        AND scheduled_date >= date('now', '-30 days')
    ''', (player_id, player_id)).fetchone()
    
    is_active = result['count'] > 0
    return is_active, False, False



@app.route('/criar-coluna-carteirinha-forcada')
def criar_coluna_carteirinha_forcada():
    conn = get_db_connection()
    try:
        conn.execute('ALTER TABLE players ADD COLUMN carteirinha_forcada INTEGER DEFAULT 0')
        conn.commit()
        return "Coluna criada com sucesso!"
    except Exception as e:
        return f"Erro ou já existe: {e}"
    finally:
        conn.close()




@app.route('/admin/recalcular-posicoes')
@login_required
def recalcular_posicoes():
    """Recalcula as posições de todos os jogadores, separadamente por sexo"""
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    
    # Recalcular posições MASCULINAS
    male_players = conn.execute('''
        SELECT id FROM players 
        WHERE active = 1 AND (sexo = 'masculino' OR sexo IS NULL OR sexo = '')
        ORDER BY position
    ''').fetchall()
    
    for i, player in enumerate(male_players, start=1):
        new_tier = get_tier_from_position(i)
        conn.execute('UPDATE players SET position = ?, tier = ? WHERE id = ?', 
                    (i, new_tier, player['id']))
    
    # Recalcular posições FEMININAS
    female_players = conn.execute('''
        SELECT id FROM players 
        WHERE active = 1 AND sexo = 'feminino'
        ORDER BY position
    ''').fetchall()
    
    for i, player in enumerate(female_players, start=1):
        new_tier = get_tier_from_position(i)
        conn.execute('UPDATE players SET position = ?, tier = ? WHERE id = ?', 
                    (i, new_tier, player['id']))
    
    conn.commit()
    conn.close()
    
    flash(f'✅ Posições recalculadas: {len(male_players)} masculinos e {len(female_players)} femininas.', 'success')
    return redirect(url_for('index'))


def auto_fix_female_ranking(conn=None):
    """
    Detecta e corrige automaticamente o ranking feminino se estiver incorreto.
    Executa sempre que há mudanças que possam afetar posições.
    """
    # Determinar se precisamos criar e fechar a conexão
    connection_provided = conn is not None
    if not connection_provided:
        conn = get_db_connection()
    
    try:
        # Buscar jogadoras femininas ativas ordenadas por posição atual
        female_players = conn.execute('''
            SELECT id, name, position FROM players 
            WHERE active = 1 AND sexo = 'feminino'
            ORDER BY position
        ''').fetchall()
        
        if not female_players:
            return  # Nenhuma jogadora feminina, nada a fazer
        
        # Verificar se as posições estão sequenciais (1, 2, 3, 4...)
        needs_fix = False
        expected_positions = list(range(1, len(female_players) + 1))
        current_positions = [player['position'] for player in female_players]
        
        if current_positions != expected_positions:
            needs_fix = True
            print(f"🔧 Auto-correção detectada: Ranking feminino incorreto")
            print(f"   Posições atuais: {current_positions}")
            print(f"   Posições esperadas: {expected_positions}")
        
        # Corrigir automaticamente se necessário
        if needs_fix:
            for i, player in enumerate(female_players, 1):
                new_position = i
                new_tier = get_tier_from_position(new_position)
                
                conn.execute('''
                    UPDATE players 
                    SET position = ?, tier = ? 
                    WHERE id = ? AND sexo = 'feminino'
                ''', (new_position, new_tier, player['id']))
            
            if not connection_provided:
                conn.commit()
            
            print(f"✅ Auto-correção concluída: {len(female_players)} jogadoras reorganizadas")
            
            # Registrar no histórico se necessário
            for i, player in enumerate(female_players, 1):
                if current_positions[i-1] != i:  # Só registra se houve mudança
                    conn.execute('''
                        INSERT INTO ranking_history 
                        (player_id, old_position, new_position, old_tier, new_tier, reason)
                        VALUES (?, ?, ?, ?, ?, ?)
                    ''', (
                        player['id'], 
                        current_positions[i-1], 
                        i, 
                        get_tier_from_position(current_positions[i-1]), 
                        get_tier_from_position(i), 
                        'auto_fix_female_ranking'
                    ))
            
            if not connection_provided:
                conn.commit()
        
    except Exception as e:
        print(f"Erro na auto-correção do ranking feminino: {str(e)}")
        if not connection_provided:
            conn.rollback()
    finally:
        if not connection_provided:
            conn.close()

# ============================================================
# ROTAS DO REGULAMENTO - COPIE TUDO ABAIXO PARA O app.py
# ============================================================

@app.route('/admin/regulamento/upload', methods=['POST'])
@login_required
def upload_regulamento():
    if not session.get('is_admin', False):
        flash('Apenas administradores podem enviar o regulamento.', 'error')
        return redirect(url_for('regulamento'))
    
    if 'regulamento_pdf' not in request.files:
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('regulamento'))
    
    file = request.files['regulamento_pdf']
    
    if file.filename == '':
        flash('Nenhum arquivo selecionado.', 'error')
        return redirect(url_for('regulamento'))
    
    if not file.filename.lower().endswith('.pdf'):
        flash('Apenas arquivos PDF são permitidos.', 'error')
        return redirect(url_for('regulamento'))
    
    try:
        regulamento_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'regulamento')
        os.makedirs(regulamento_folder, exist_ok=True)
        
        timestamp = int(datetime.now().timestamp())
        nome_arquivo = f"regulamento_{timestamp}.pdf"
        file_path = os.path.join(regulamento_folder, nome_arquivo)
        
        file.save(file_path)
        
        conn = get_db_connection()
        
        old_regulamento = conn.execute('SELECT nome_arquivo FROM regulamento ORDER BY data_upload DESC LIMIT 1').fetchone()
        
        if old_regulamento:
            old_file_path = os.path.join(regulamento_folder, old_regulamento['nome_arquivo'])
            if os.path.exists(old_file_path):
                os.remove(old_file_path)
            conn.execute('DELETE FROM regulamento')
        
        conn.execute('''
            INSERT INTO regulamento (nome_arquivo, nome_original, data_upload)
            VALUES (?, ?, CURRENT_TIMESTAMP)
        ''', (nome_arquivo, file.filename))
        
        conn.commit()
        conn.close()
        
        flash('Regulamento atualizado com sucesso!', 'success')
        
    except Exception as e:
        flash(f'Erro ao enviar regulamento: {str(e)}', 'error')
    
    return redirect(url_for('regulamento'))


@app.route('/regulamento/download')
def download_regulamento():
    conn = get_db_connection()
    regulamento = conn.execute('''
        SELECT * FROM regulamento ORDER BY data_upload DESC LIMIT 1
    ''').fetchone()
    conn.close()
    
    if not regulamento:
        flash('Regulamento não disponível.', 'error')
        return redirect(url_for('regulamento'))
    
    regulamento_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'regulamento')
    file_path = os.path.join(regulamento_folder, regulamento['nome_arquivo'])
    
    if not os.path.exists(file_path):
        flash('Arquivo não encontrado.', 'error')
        return redirect(url_for('regulamento'))
    
    return send_file(
        file_path,
        as_attachment=True,
        download_name=regulamento['nome_original'],
        mimetype='application/pdf'
    )


@app.route('/admin/regulamento/delete', methods=['POST'])
@login_required
def delete_regulamento():
    if not session.get('is_admin', False):
        flash('Apenas administradores podem excluir o regulamento.', 'error')
        return redirect(url_for('regulamento'))
    
    try:
        conn = get_db_connection()
        
        regulamento = conn.execute('SELECT nome_arquivo FROM regulamento ORDER BY data_upload DESC LIMIT 1').fetchone()
        
        if regulamento:
            regulamento_folder = os.path.join(app.config['UPLOAD_FOLDER'], 'regulamento')
            file_path = os.path.join(regulamento_folder, regulamento['nome_arquivo'])
            
            if os.path.exists(file_path):
                os.remove(file_path)
            
            conn.execute('DELETE FROM regulamento')
            conn.commit()
            
            flash('Regulamento excluído com sucesso!', 'success')
        else:
            flash('Nenhum regulamento para excluir.', 'warning')
        
        conn.close()
        
    except Exception as e:
        flash(f'Erro ao excluir regulamento: {str(e)}', 'error')
    
    return redirect(url_for('regulamento'))



@app.route('/criar-tabela-regulamento')
def criar_tabela_regulamento():
    conn = get_db_connection()
    conn.execute('''
        CREATE TABLE IF NOT EXISTS regulamento (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            nome_arquivo TEXT NOT NULL,
            nome_original TEXT NOT NULL,
            data_upload TIMESTAMP DEFAULT CURRENT_TIMESTAMP
        )
    ''')
    conn.commit()
    conn.close()
    return "Tabela 'regulamento' criada com sucesso!"


# ============================================================
# SISTEMA DE BLOQUEIO DE JOGADOR (SAÚDE/VIAGEM)
# 
# 1. Execute a rota /criar-coluna-bloqueio uma vez
# 2. Adicione as rotas no app.py
# 3. Adicione o CSS e HTML na pirâmide
# 4. Adicione o controle no player_detail.html
# ============================================================

# ============================================================
# PASSO 1: Rota para criar coluna no banco (executar uma vez)
# ============================================================

@app.route('/criar-coluna-bloqueio')
@login_required
def criar_coluna_bloqueio():
    """Cria as colunas para bloqueio de jogador - executar uma vez"""
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    try:
        cursor = conn.execute("PRAGMA table_info(players)")
        columns = [col[1] for col in cursor.fetchall()]
        
        if 'bloqueado' not in columns:
            conn.execute('ALTER TABLE players ADD COLUMN bloqueado INTEGER DEFAULT 0')
            conn.commit()
            flash('✅ Coluna "bloqueado" criada!', 'success')
        
        if 'bloqueio_motivo' not in columns:
            conn.execute('ALTER TABLE players ADD COLUMN bloqueio_motivo TEXT')
            conn.commit()
            flash('✅ Coluna "bloqueio_motivo" criada!', 'success')
        
        if 'bloqueio_ate' not in columns:
            conn.execute('ALTER TABLE players ADD COLUMN bloqueio_ate DATE')
            conn.commit()
            flash('✅ Coluna "bloqueio_ate" criada!', 'success')
            
        flash('✅ Sistema de bloqueio configurado com sucesso!', 'success')
        
    except Exception as e:
        flash(f'❌ Erro: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))


# ============================================================
# PASSO 2: Rota para bloquear/desbloquear jogador
# ============================================================

@app.route('/admin/toggle-bloqueio/<int:player_id>', methods=['POST'])
@login_required
def toggle_bloqueio_jogador(player_id):
    """Admin pode bloquear/desbloquear jogador"""
    if not session.get('is_admin', False):
        flash('Acesso restrito a administradores.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    try:
        player = conn.execute('SELECT name, bloqueado FROM players WHERE id = ?', 
                              (player_id,)).fetchone()
        
        if not player:
            flash('❌ Jogador não encontrado.', 'error')
            conn.close()
            return redirect(url_for('index'))
        
        # Se está bloqueado, desbloqueia
        if player['bloqueado'] == 1:
            conn.execute('''
                UPDATE players 
                SET bloqueado = 0, bloqueio_motivo = NULL, bloqueio_ate = NULL 
                WHERE id = ?
            ''', (player_id,))
            conn.commit()
            flash(f'✅ {player["name"]} foi DESBLOQUEADO e pode receber desafios.', 'success')
        else:
            # Se não está bloqueado, pega os dados do formulário
            motivo = request.form.get('motivo', 'Não especificado')
            data_ate = request.form.get('data_ate', None)
            
            conn.execute('''
                UPDATE players 
                SET bloqueado = 1, bloqueio_motivo = ?, bloqueio_ate = ? 
                WHERE id = ?
            ''', (motivo, data_ate, player_id))
            conn.commit()
            
            msg_data = f" até {data_ate}" if data_ate else ""
            flash(f'🚫 {player["name"]} foi BLOQUEADO ({motivo}){msg_data}.', 'warning')
        
    except Exception as e:
        flash(f'❌ Erro: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))


@app.route('/criar-colunas-posicao-desafio')
def criar_colunas_posicao_desafio():
    """Cria as colunas para armazenar as posições no momento do desafio"""
    conn = get_db_connection()
    
    try:
        # Verificar se as colunas já existem
        columns = conn.execute("PRAGMA table_info(challenges)").fetchall()
        column_names = [col['name'] for col in columns]
        
        colunas_criadas = []
        
        # Adicionar coluna challenger_position_at_creation
        if 'challenger_position_at_creation' not in column_names:
            conn.execute('ALTER TABLE challenges ADD COLUMN challenger_position_at_creation INTEGER')
            colunas_criadas.append('challenger_position_at_creation')
        
        # Adicionar coluna challenged_position_at_creation
        if 'challenged_position_at_creation' not in column_names:
            conn.execute('ALTER TABLE challenges ADD COLUMN challenged_position_at_creation INTEGER')
            colunas_criadas.append('challenged_position_at_creation')
        
        conn.commit()
        conn.close()
        
        if colunas_criadas:
            return f"✅ Colunas criadas com sucesso: {', '.join(colunas_criadas)}"
        else:
            return "ℹ️ As colunas já existem na tabela challenges."
            
    except Exception as e:
        conn.close()
        return f"❌ Erro ao criar colunas: {str(e)}"

@app.route('/preencher-posicoes-historicas')
@login_required
def preencher_posicoes_historicas():
    """Tenta preencher as posições dos desafios antigos baseado no histórico"""
    if not session.get('is_admin'):
        flash('Acesso negado.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    
    try:
        # Buscar desafios sem posição preenchida
        desafios = conn.execute('''
            SELECT id, challenger_id, challenged_id, created_at
            FROM challenges
            WHERE challenger_position_at_creation IS NULL
               OR challenged_position_at_creation IS NULL
        ''').fetchall()
        
        atualizados = 0
        
        for desafio in desafios:
            # Tentar buscar a posição do challenger no momento do desafio
            challenger_pos = None
            challenged_pos = None
            
            # Buscar no ranking_history a posição mais próxima da data do desafio
            hist_challenger = conn.execute('''
                SELECT new_position 
                FROM ranking_history 
                WHERE player_id = ? 
                AND change_date <= ?
                ORDER BY change_date DESC
                LIMIT 1
            ''', (desafio['challenger_id'], desafio['created_at'])).fetchone()
            
            if hist_challenger:
                challenger_pos = hist_challenger['new_position']
            
            hist_challenged = conn.execute('''
                SELECT new_position 
                FROM ranking_history 
                WHERE player_id = ? 
                AND change_date <= ?
                ORDER BY change_date DESC
                LIMIT 1
            ''', (desafio['challenged_id'], desafio['created_at'])).fetchone()
            
            if hist_challenged:
                challenged_pos = hist_challenged['new_position']
            
            # Se encontrou ambas as posições, atualizar
            if challenger_pos and challenged_pos:
                conn.execute('''
                    UPDATE challenges 
                    SET challenger_position_at_creation = ?,
                        challenged_position_at_creation = ?
                    WHERE id = ?
                ''', (challenger_pos, challenged_pos, desafio['id']))
                atualizados += 1
        
        conn.commit()
        conn.close()
        
        return f"✅ {atualizados} desafios atualizados com posições históricas."
        
    except Exception as e:
        conn.close()
        return f"❌ Erro: {str(e)}"


@app.route('/criar-coluna-telefone')
def criar_coluna_telefone():
    conn = get_db_connection()
    try:
        columns = conn.execute("PRAGMA table_info(players)").fetchall()
        column_names = [col['name'] for col in columns]
        
        if 'telefone' not in column_names:
            conn.execute('ALTER TABLE players ADD COLUMN telefone TEXT')
            conn.commit()
            conn.close()
            return "✅ Coluna 'telefone' criada com sucesso!"
        else:
            conn.close()
            return "ℹ️ Coluna 'telefone' já existe."
    except Exception as e:
        conn.close()
        return f"❌ Erro: {str(e)}"



# ============================================================
# ROTA PARA ATUALIZAR WHATSAPP DO JOGADOR
# Adicionar ao app.py
# ============================================================

@app.route('/player/<int:player_id>/update-whatsapp', methods=['POST'])
@login_required
def update_player_whatsapp(player_id):
    """Atualiza o WhatsApp do jogador para notificações"""
    
    # Verificar permissão (próprio jogador ou admin)
    if session.get('user_id') != player_id and not session.get('is_admin'):
        flash('Você não tem permissão para editar este perfil.', 'danger')
        return redirect(url_for('player_detail', player_id=player_id))
    
    new_whatsapp = request.form.get('new_whatsapp', '').strip()
    
    # Remover caracteres não numéricos
    new_whatsapp = ''.join(filter(str.isdigit, new_whatsapp))
    
    # Se vazio, define como NULL
    if not new_whatsapp:
        new_whatsapp = None
    
    conn = get_db_connection()
    conn.execute(
        'UPDATE players SET telefone = ? WHERE id = ?',
        (new_whatsapp, player_id)
    )
    conn.commit()
    conn.close()
    
    if new_whatsapp:
        flash(f'WhatsApp atualizado para {new_whatsapp}. Você receberá notificações!', 'success')
    else:
        flash('WhatsApp removido. Você não receberá mais notificações.', 'info')
    
    return redirect(url_for('player_detail', player_id=player_id))


# ============================================================
# CHATBOT WHATSAPP - LIGA OLÍMPICA DE GOLFE
# ============================================================
# Adicione este código ao seu app.py
# ============================================================

import re
from datetime import datetime, timedelta
from flask import request, jsonify

# ============================================================
# CONFIGURAÇÃO DO WEBHOOK NA EVOLUTION API
# ============================================================
# Execute este comando no servidor para configurar o webhook:
#
# curl -X POST "http://159.89.35.66:8080/webhook/set/liga-golf" \
#   -H "apikey: liga-golf-api-key-2024" \
#   -H "Content-Type: application/json" \
#   -d '{
#     "url": "https://SEU_DOMINIO/webhook/whatsapp",
#     "webhookByEvents": true,
#     "events": ["MESSAGES_UPSERT"]
#   }'
#
# Substitua SEU_DOMINIO pelo domínio da sua aplicação Flask
# ============================================================


# ============================================================
# FUNÇÕES AUXILIARES
# ============================================================

# ============================================================
# CHATBOT WHATSAPP - Funções de Estado da Conversa
# ============================================================

def get_chat_state(telefone):
    """Retorna o estado atual da conversa do usuário (do banco de dados)"""
    telefone_norm = normalizar_telefone(telefone)
    
    conn = get_db_connection()
    state = conn.execute(
        'SELECT estado, dados, expira FROM chat_states WHERE telefone = ?',
        (telefone_norm,)
    ).fetchone()
    conn.close()
    
    if not state:
        return None
    
    # Verificar se expirou
    try:
        expira = datetime.strptime(state['expira'], '%Y-%m-%d %H:%M:%S')
        if datetime.now() > expira:
            clear_chat_state(telefone_norm)
            return None
    except:
        pass
    
    # Parsear dados JSON
    import json
    try:
        dados = json.loads(state['dados']) if state['dados'] else {}
    except:
        dados = {}
    
    return {
        'estado': state['estado'],
        'dados': dados
    }


def set_chat_state(telefone, estado, dados=None):
    """Define o estado da conversa do usuário (no banco de dados)"""
    import json
    telefone_norm = normalizar_telefone(telefone)
    
    dados_json = json.dumps(dados or {})
    expira = (datetime.now() + timedelta(minutes=30)).strftime('%Y-%m-%d %H:%M:%S')
    
    conn = get_db_connection()
    conn.execute('''
        INSERT OR REPLACE INTO chat_states (telefone, estado, dados, expira)
        VALUES (?, ?, ?, ?)
    ''', (telefone_norm, estado, dados_json, expira))
    conn.commit()
    conn.close()


def clear_chat_state(telefone):
    """Limpa o estado da conversa do usuário (do banco de dados)"""
    telefone_norm = normalizar_telefone(telefone)
    
    conn = get_db_connection()
    conn.execute('DELETE FROM chat_states WHERE telefone = ?', (telefone_norm,))
    conn.commit()
    conn.close()


def criar_desafio_via_whatsapp(challenger_id, challenged_id, scheduled_date):
    """
    Cria um desafio via WhatsApp.
    Retorna: (sucesso: bool, mensagem: str, challenge_id: int ou None)
    """
    conn = get_db_connection()
    
    try:
        # Buscar dados dos jogadores
        challenger = conn.execute(
            'SELECT * FROM players WHERE id = ? AND active = 1', 
            (challenger_id,)
        ).fetchone()
        
        challenged = conn.execute(
            'SELECT * FROM players WHERE id = ? AND active = 1', 
            (challenged_id,)
        ).fetchone()
        
        if not challenger or not challenged:
            conn.close()
            return False, "Jogador não encontrado ou inativo.", None
        
        # Verificar se desafiado está bloqueado
        if challenged['bloqueado'] == 1:
            motivo = challenged['bloqueio_motivo'] or 'indisponível'
            conn.close()
            return False, f"{challenged['name']} está bloqueado ({motivo}).", None
        
        # Verificar se já existe desafio pendente entre eles
        existing = conn.execute('''
            SELECT id FROM challenges 
            WHERE ((challenger_id = ? AND challenged_id = ?) 
                   OR (challenger_id = ? AND challenged_id = ?))
            AND status IN ('pending', 'accepted')
        ''', (challenger_id, challenged_id, challenged_id, challenger_id)).fetchone()
        
        if existing:
            conn.close()
            return False, "Já existe um desafio pendente entre vocês.", None
        
        # Verificar se algum dos dois já tem desafio ativo
        challenger_busy = conn.execute('''
            SELECT id FROM challenges 
            WHERE (challenger_id = ? OR challenged_id = ?)
            AND status IN ('pending', 'accepted')
        ''', (challenger_id, challenger_id)).fetchone()
        
        if challenger_busy:
            conn.close()
            return False, "Você já tem um desafio ativo. Conclua-o primeiro.", None
        
        challenged_busy = conn.execute('''
            SELECT id FROM challenges 
            WHERE (challenger_id = ? OR challenged_id = ?)
            AND status IN ('pending', 'accepted')
        ''', (challenged_id, challenged_id)).fetchone()
        
        if challenged_busy:
            conn.close()
            return False, f"{challenged['name']} já está em um desafio ativo.", None
        
        # Validar posições (desafiante deve estar abaixo do desafiado)
        if challenger['position'] <= challenged['position']:
            conn.close()
            return False, "Você só pode desafiar jogadores acima de você no ranking.", None
        
        # Validar distância máxima de 8 posições
        if challenger['position'] - challenged['position'] > 8:
            conn.close()
            return False, "Você só pode desafiar jogadores até 8 posições acima.", None
        
        # Criar o desafio
        current_datetime = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        response_deadline = (datetime.now() + timedelta(days=2)).strftime('%Y-%m-%d %H:%M:%S')
        
        conn.execute('''
            INSERT INTO challenges (
                challenger_id, 
                challenged_id, 
                status, 
                scheduled_date, 
                created_at, 
                response_deadline,
                challenger_position_at_creation,
                challenged_position_at_creation
            )
            VALUES (?, ?, 'pending', ?, ?, ?, ?, ?)
        ''', (
            challenger_id, 
            challenged_id, 
            scheduled_date, 
            current_datetime, 
            response_deadline,
            challenger['position'],
            challenged['position']
        ))
        
        challenge_id = conn.execute('SELECT last_insert_rowid()').fetchone()[0]
        
        # Registrar no log
        try:
            conn.execute('''
                INSERT INTO challenge_logs 
                (challenge_id, user_id, modified_by, old_status, new_status, notes, created_at)
                VALUES (?, ?, ?, ?, ?, ?, ?)
            ''', (
                challenge_id, 
                str(challenger_id),
                "WhatsApp Bot",
                None, 
                'pending', 
                f"Desafio criado via WhatsApp. Jogo: {scheduled_date}. Prazo resposta: 2 dias.",
                current_datetime
            ))
        except Exception as e:
            print(f"Erro ao registrar log: {e}")
        
        conn.commit()
        conn.close()
        
        return True, "Desafio criado com sucesso!", challenge_id
        
    except Exception as e:
        conn.rollback()
        conn.close()
        return False, f"Erro ao criar desafio: {str(e)}", None


def get_player_phone(player_id):
    """Busca o telefone de um jogador"""
    conn = get_db_connection()
    player = conn.execute('SELECT telefone FROM players WHERE id = ?', (player_id,)).fetchone()
    conn.close()
    
    if player and player['telefone']:
        return normalizar_telefone(player['telefone'])
    return None


def notificar_desafio_criado_whatsapp(challenge_id):
    """Notifica o desafiado sobre o novo desafio"""
    conn = get_db_connection()
    
    challenge = conn.execute('''
        SELECT c.*, 
               challenger.name as challenger_name,
               challenger.position as challenger_position,
               challenged.name as challenged_name,
               challenged.telefone as challenged_telefone
        FROM challenges c
        JOIN players challenger ON c.challenger_id = challenger.id
        JOIN players challenged ON c.challenged_id = challenged.id
        WHERE c.id = ?
    ''', (challenge_id,)).fetchone()
    
    conn.close()
    
    if not challenge:
        return
    
    # Formatar data
    try:
        data_obj = datetime.strptime(challenge['scheduled_date'], '%Y-%m-%d')
        data_fmt = data_obj.strftime('%d/%m/%Y')
    except:
        data_fmt = challenge['scheduled_date']
    
    # Notificar desafiado
    telefone_desafiado = challenge['challenged_telefone']
    if telefone_desafiado:
        msg = f"""🏌️ *NOVO DESAFIO!*

Você foi desafiado por *{challenge['challenger_name']}* (#{challenge['challenger_position']})

📅 Data proposta: *{data_fmt}*

━━━━━━━━━━━━━━━━━━━━━
*ESCOLHA UMA OPÇÃO:*
━━━━━━━━━━━━━━━━━━━━━

*[4]* ✅ Aceitar a data
*[5]* ❌ Rejeitar (WO - você perde)
*[7]* 📅 Propor 2 novas datas

━━━━━━━━━━━━━━━━━━━━━

⏰ Prazo para responder: *2 dias*"""
        
        telefone_norm = normalizar_telefone(telefone_desafiado)
        enviar_mensagem_whatsapp(f"55{telefone_norm}@s.whatsapp.net", msg)
    
    # Notificar no grupo
    msg_grupo = f"""🏆 *NOVO DESAFIO CRIADO*

⚔️ *{challenge['challenger_name']}* (#{challenge['challenger_position']}) 
    desafiou 
    *{challenge['challenged_name']}*

📅 Data proposta: {data_fmt}

Boa sorte! 🍀"""
    
    enviar_mensagem_whatsapp(WHATSAPP_GRUPO_LIGA, msg_grupo)


def notificar_desafio_aceito_bot(challenge_id):
    """Notifica no grupo que um desafio foi aceito"""
    try:
        conn = get_db_connection()
        
        challenge = conn.execute('''
            SELECT c.*, 
                   challenger.name as challenger_name,
                   challenger.position as challenger_pos,
                   challenged.name as challenged_name,
                   challenged.position as challenged_pos
            FROM challenges c
            JOIN players challenger ON c.challenger_id = challenger.id
            JOIN players challenged ON c.challenged_id = challenged.id
            WHERE c.id = ?
        ''', (challenge_id,)).fetchone()
        
        conn.close()
        
        if not challenge:
            return
        
        # Formatar data
        data_jogo = challenge['scheduled_date']
        try:
            data_obj = datetime.strptime(data_jogo, '%Y-%m-%d')
            data_formatada = data_obj.strftime('%d/%m/%Y')
        except:
            data_formatada = data_jogo
        
        mensagem = f"""✅ *DESAFIO ACEITO!*

🏌️ *{challenge['challenger_name']}* ({challenge['challenger_pos']}º)
      ⚔️ vs ⚔️
🏌️ *{challenge['challenged_name']}* ({challenge['challenged_pos']}º)

📅 Data do jogo: *{data_formatada}*

Boa sorte aos dois! 🏆"""

        enviar_mensagem_whatsapp(WHATSAPP_GRUPO_LIGA, mensagem)
        
    except Exception as e:
        print(f"Erro ao notificar desafio aceito: {e}")





def normalizar_telefone(telefone):
    """Remove caracteres não numéricos e padroniza o telefone"""
    if not telefone:
        return None
    # Remove tudo que não é número
    apenas_numeros = re.sub(r'\D', '', telefone)
    # Remove 55 do início se tiver (código do Brasil)
    if apenas_numeros.startswith('55') and len(apenas_numeros) > 11:
        apenas_numeros = apenas_numeros[2:]
    return apenas_numeros


def extrair_telefone_do_jid(jid):
    """Extrai número de telefone do JID do WhatsApp (ex: 5521999998888@s.whatsapp.net)"""
    if not jid:
        return None
    # Remove sufixo do WhatsApp
    numero = jid.split('@')[0]
    return normalizar_telefone(numero)


def get_player_by_phone(telefone):
    """Busca jogador pelo número de telefone"""
    telefone_normalizado = normalizar_telefone(telefone)
    if not telefone_normalizado:
        return None
    
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Busca considerando variações do número (com/sem 55, com/sem 9)
    cursor.execute("""
        SELECT id, name, position, sexo, telefone
        FROM players 
        WHERE REPLACE(REPLACE(REPLACE(telefone, '-', ''), ' ', ''), '+', '') LIKE ?
           OR REPLACE(REPLACE(REPLACE(telefone, '-', ''), ' ', ''), '+', '') LIKE ?
           OR REPLACE(REPLACE(REPLACE(telefone, '-', ''), ' ', ''), '+', '') LIKE ?
    """, (
        f'%{telefone_normalizado}',
        f'%{telefone_normalizado[-9:]}',  # Últimos 9 dígitos
        f'%{telefone_normalizado[-8:]}'   # Últimos 8 dígitos
    ))
    
    player = cursor.fetchone()
    conn.close()
    
    return dict(player) if player else None


def get_possiveis_desafiados(player_id):
    """Retorna lista de jogadores que podem ser desafiados"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Buscar dados do jogador
    cursor.execute("SELECT position, sexo FROM players WHERE id = ?", (player_id,))
    player = cursor.fetchone()
    
    if not player:
        conn.close()
        return []
    
    posicao_atual = player['position']
    sexo = player['sexo'] or 'masculino'
    
    # Calcular posição mínima (até 8 posições acima)
    posicao_minima = max(1, posicao_atual - 8)
    
    # Buscar possíveis desafiados (mesma categoria, posição superior, não bloqueados)
    cursor.execute("""
        SELECT id, name, position
        FROM players
        WHERE position >= ? 
          AND position < ?
          AND (sexo = ? OR sexo IS NULL OR sexo = '')
          AND (bloqueado = 0 OR bloqueado IS NULL)
          AND id != ?
        ORDER BY position ASC
    """, (posicao_minima, posicao_atual, sexo, player_id))
    
    desafiados = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    # Verificar se jogadores já têm desafios pendentes
    desafiados_disponiveis = []
    for d in desafiados:
        if not tem_desafio_ativo(d['id']) and not tem_desafio_ativo(player_id):
            desafiados_disponiveis.append(d)
    
    return desafiados_disponiveis


def tem_desafio_ativo(player_id):
    """Verifica se jogador tem desafio pendente ou aceito"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT COUNT(*) as count
        FROM challenges
        WHERE (challenger_id = ? OR challenged_id = ?)
          AND status IN ('pending', 'accepted')
    """, (player_id, player_id))
    
    result = cursor.fetchone()
    conn.close()
    
    return result['count'] > 0


def get_desafios_pendentes(player_id):
    """Retorna desafios pendentes do jogador (onde ele é o desafiado)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            c.id,
            c.status,
            c.scheduled_date,
            c.created_at,
            challenger.name as challenger_name,
            challenger.position as challenger_position
        FROM challenges c
        JOIN players challenger ON c.challenger_id = challenger.id
        WHERE c.challenged_id = ?
          AND c.status = 'pending'
        ORDER BY c.created_at DESC
    """, (player_id,))
    
    desafios = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return desafios


def get_meus_desafios(player_id):
    """Retorna todos os desafios ativos do jogador"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    cursor.execute("""
        SELECT 
            c.id,
            c.status,
            c.scheduled_date,
            c.challenger_id,
            c.challenged_id,
            challenger.name as challenger_name,
            challenger.position as challenger_position,
            challenged.name as challenged_name,
            challenged.position as challenged_position
        FROM challenges c
        JOIN players challenger ON c.challenger_id = challenger.id
        JOIN players challenged ON c.challenged_id = challenged.id
        WHERE (c.challenger_id = ? OR c.challenged_id = ?)
          AND c.status IN ('pending', 'accepted')
        ORDER BY c.created_at DESC
    """, (player_id, player_id))
    
    desafios = [dict(row) for row in cursor.fetchall()]
    conn.close()
    
    return desafios


def aceitar_desafio(challenge_id, player_id):
    """Aceita um desafio pendente"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar se o desafio existe e o jogador é o desafiado
    cursor.execute("""
        SELECT id, challenged_id, status
        FROM challenges
        WHERE id = ? AND challenged_id = ? AND status = 'pending'
    """, (challenge_id, player_id))
    
    challenge = cursor.fetchone()
    
    if not challenge:
        conn.close()
        return False, "Desafio não encontrado ou você não é o desafiado."
    
    # Atualizar status
    cursor.execute("""
        UPDATE challenges
        SET status = 'accepted', updated_at = ?
        WHERE id = ?
    """, (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), challenge_id))
    
    conn.commit()
    conn.close()
    
    return True, "Desafio aceito com sucesso!"


def rejeitar_desafio(challenge_id, player_id):
    """Rejeita um desafio (aplica WO - desafiado perde)"""
    conn = get_db_connection()
    cursor = conn.cursor()
    
    # Verificar se o desafio existe e o jogador é o desafiado
    cursor.execute("""
        SELECT c.id, c.challenger_id, c.challenged_id, c.status,
               challenger.position as challenger_pos,
               challenged.position as challenged_pos
        FROM challenges c
        JOIN players challenger ON c.challenger_id = challenger.id
        JOIN players challenged ON c.challenged_id = challenged.id
        WHERE c.id = ? AND c.challenged_id = ? AND c.status = 'pending'
    """, (challenge_id, player_id))
    
    challenge = cursor.fetchone()
    
    if not challenge:
        conn.close()
        return False, "Desafio não encontrado ou você não é o desafiado."
    
    # Aplicar WO - desafiante vence (desafiado rejeitou)
    # O desafiante sobe 1 posição (troca com o desafiado)
    challenger_id = challenge['challenger_id']
    challenged_id = challenge['challenged_id']
    challenger_pos = challenge['challenger_pos']
    challenged_pos = challenge['challenged_pos']
    
    # Trocar posições (desafiante assume posição do desafiado)
    cursor.execute("UPDATE players SET position = ? WHERE id = ?", (challenged_pos, challenger_id))
    cursor.execute("UPDATE players SET position = ? WHERE id = ?", (challenger_pos, challenged_id))
    
    # Atualizar desafio
    cursor.execute("""
        UPDATE challenges
        SET status = 'completed',
            result = 'challenger_win',
            result_type = 'wo_challenged',
            updated_at = ?
        WHERE id = ?
    """, (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), challenge_id))
    
    conn.commit()
    conn.close()
    
    return True, "Desafio rejeitado. WO aplicado - você perdeu a posição."


# ============================================================
# PROCESSADOR DE COMANDOS
# ============================================================

def processar_comando_whatsapp(mensagem, telefone):
    """Processa mensagem recebida e retorna resposta"""
    
    # Normalizar mensagem
    msg = mensagem.lower().strip()
    telefone_normalizado = normalizar_telefone(telefone)
    
    # Buscar jogador pelo telefone
    jogador = get_player_by_phone(telefone)
    
    if not jogador:
        return """❌ *Número não cadastrado*

Seu número de WhatsApp não está vinculado a nenhum jogador da Liga.

Para cadastrar, acesse seu perfil no site e adicione seu número no campo "WhatsApp para Notificações"."""
    
    # ---------------------------------------------------------
    # VERIFICAR SE HÁ ESTADO PENDENTE (conversa em andamento)
    # ---------------------------------------------------------
    estado_atual = get_chat_state(telefone_normalizado)
    
    if estado_atual:
        estado = estado_atual['estado']
        dados = estado_atual['dados']
        
        # ---------------------------------------------------------
        # ESTADO: Selecionando oponente para desafio
        # ---------------------------------------------------------
        if estado == 'selecionando_oponente':
            # Verificar se digitou "0" para cancelar
            if msg == '0' or 'cancelar' in msg:
                clear_chat_state(telefone_normalizado)
                return "❌ Criação de desafio cancelada.\n\n_Digite *0* para ver o menu._"
            
            # Tentar interpretar como número da lista
            try:
                opcao = int(msg)
                possiveis = dados.get('possiveis', [])
                
                if opcao < 1 or opcao > len(possiveis):
                    return f"""⚠️ Opção inválida!

Digite um número de *1* a *{len(possiveis)}* para selecionar o oponente.

Ou digite *0* para cancelar."""
                
                # Oponente selecionado
                oponente = possiveis[opcao - 1]
                
                # Atualizar estado para pedir a data
                set_chat_state(telefone_normalizado, 'informando_data', {
                    'oponente_id': oponente['id'],
                    'oponente_nome': oponente['name'],
                    'oponente_posicao': oponente['position']
                })
                
                # Calcular data máxima (7 dias)
                hoje = datetime.now()
                data_max = hoje + timedelta(days=7)
                
                return f"""✅ Oponente selecionado: *{oponente['name']}* ({oponente['position']}º)

📅 *Qual a data do jogo?*

Digite no formato *DD/MM* (ex: {data_max.strftime('%d/%m')})

A data deve ser nos próximos *7 dias*.
(até {data_max.strftime('%d/%m/%Y')})

_Digite *0* para cancelar._"""
                
            except ValueError:
                return """⚠️ Digite apenas o *número* do oponente.

Exemplo: *1* ou *2* ou *3*

_Digite *0* para cancelar._"""
        
        # ---------------------------------------------------------
        # ESTADO: Informando data do jogo
        # ---------------------------------------------------------
        elif estado == 'informando_data':
            # Verificar se digitou "0" para cancelar
            if msg == '0' or 'cancelar' in msg:
                clear_chat_state(telefone_normalizado)
                return "❌ Criação de desafio cancelada.\n\n_Digite *0* para ver o menu._"
            
            # Tentar interpretar a data
            data_jogo = None
            hoje = datetime.now().date()
            ano_atual = hoje.year
            
            # Formatos aceitos: DD/MM, DD-MM, DD.MM, DD/MM/YYYY
            formatos = [
                (r'^(\d{1,2})[/\-.](\d{1,2})$', '%d/%m'),
                (r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$', '%d/%m/%Y'),
                (r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2})$', '%d/%m/%y'),
            ]
            
            for pattern, fmt in formatos:
                match = re.match(pattern, msg.strip())
                if match:
                    try:
                        if len(match.groups()) == 2:
                            dia, mes = match.groups()
                            data_str = f"{dia}/{mes}/{ano_atual}"
                            data_jogo = datetime.strptime(data_str, '%d/%m/%Y').date()
                            
                            if data_jogo < hoje:
                                data_jogo = datetime.strptime(f"{dia}/{mes}/{ano_atual + 1}", '%d/%m/%Y').date()
                        else:
                            data_str = msg.strip().replace('-', '/').replace('.', '/')
                            data_jogo = datetime.strptime(data_str, fmt).date()
                        break
                    except ValueError:
                        continue
            
            if not data_jogo:
                return """⚠️ Formato de data inválido!

Digite no formato *DD/MM* (ex: 25/02)

_Digite *0* para cancelar._"""
            
            # Validar data
            if data_jogo < hoje:
                return """⚠️ A data não pode ser no passado!

Digite uma data a partir de hoje.

_Digite *0* para cancelar._"""
            
            data_max = hoje + timedelta(days=7)
            if data_jogo > data_max:
                return f"""⚠️ A data não pode ser superior a 7 dias!

Data máxima permitida: *{data_max.strftime('%d/%m/%Y')}*

_Digite *0* para cancelar._"""
            
            # CRIAR O DESAFIO!
            oponente_id = dados['oponente_id']
            oponente_nome = dados['oponente_nome']
            oponente_posicao = dados['oponente_posicao']
            data_formatada = data_jogo.strftime('%Y-%m-%d')
            
            sucesso, mensagem_retorno, challenge_id = criar_desafio_via_whatsapp(
                jogador['id'], 
                oponente_id, 
                data_formatada
            )
            
            # Limpar estado da conversa
            clear_chat_state(telefone_normalizado)
            
            if sucesso:
                # Notificar no grupo e para o desafiado
                try:
                    notificar_desafio_criado_whatsapp(challenge_id)
                except Exception as e:
                    print(f"Erro ao notificar: {e}")
                
                return f"""🎉 *DESAFIO CRIADO COM SUCESSO!*

Você desafiou *{oponente_nome}* ({oponente_posicao}º)

📅 Data do jogo: *{data_jogo.strftime('%d/%m/%Y')}*
⏳ Prazo para resposta: *2 dias*

O desafiado será notificado e deve aceitar ou rejeitar o desafio.

Boa sorte! 🏌️

_Digite *0* para voltar ao menu._"""
            else:
                return f"""❌ *Erro ao criar desafio*

{mensagem_retorno}

_Digite *0* para voltar ao menu._"""
    
    # ---------------------------------------------------------
    # COMANDOS NORMAIS (sem estado pendente)
    # ---------------------------------------------------------
    
    # COMANDO [1]: Minha posição
    if msg == '1' or any(palavra in msg for palavra in ['posição', 'posicao', 'ranking', 'colocação', 'colocacao']):
        return f"""📊 *Sua Posição no Ranking*

Olá, {jogador['name']}!

Você está atualmente na posição *{jogador['position']}º* no ranking da Liga Olímpica de Golfe.

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [2]: Quem posso desafiar
    if msg == '2' or any(palavra in msg for palavra in ['desafiado', 'quem posso', 'possiveis', 'possíveis']):
        possiveis = get_possiveis_desafiados(jogador['id'])
        
        if not possiveis:
            return f"""🎯 *Possíveis Desafiados*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

No momento não há jogadores disponíveis para desafio.

_Digite *0* para voltar ao menu._"""
        
        lista = "\n".join([f"   {p['position']}º - {p['name']}" for p in possiveis])
        
        return f"""🎯 *Possíveis Desafiados*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

Você pode desafiar:
{lista}

📱 Para criar um desafio, digite *6*

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [3]: Meus desafios
    if msg == '3' or (any(palavra in msg for palavra in ['meus desafio', 'meu desafio']) and 'criar' not in msg):
        desafios = get_meus_desafios(jogador['id'])
        
        if not desafios:
            return f"""📋 *Meus Desafios*

Olá, {jogador['name']}!

Você não tem desafios ativos no momento.

_Digite *0* para voltar ao menu._"""
        
        linhas = []
        for d in desafios:
            status_emoji = "⏳" if d['status'] == 'pending' else "✅"
            status_texto = "Pendente" if d['status'] == 'pending' else "Aceito"
            
            if d['challenger_id'] == jogador['id']:
                linhas.append(f"   {status_emoji} #{d['id']} - Você → {d['challenged_name']} ({d['challenged_position']}º) [{status_texto}]")
            else:
                linhas.append(f"   {status_emoji} #{d['id']} - {d['challenger_name']} ({d['challenger_position']}º) → Você [{status_texto}]")
        
        lista = "\n".join(linhas)
        
        pendentes_para_responder = [d for d in desafios if d['status'] == 'pending' and d['challenged_id'] == jogador['id']]
        
        dica = ""
        if pendentes_para_responder:
            dica = "\n\n💡 Para responder, digite *4* (aceitar) ou *5* (rejeitar)"
        
        return f"""📋 *Meus Desafios*

Olá, {jogador['name']}!

Seus desafios ativos:
{lista}{dica}

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [4]: Aceitar desafio
    if msg == '4' or 'aceitar' in msg or 'aceito' in msg:
        numeros = re.findall(r'\d+', msg)
        if numeros and numeros[0] == '4' and len(msg) <= 2:
            numeros = []
        
        desafios_pendentes = get_desafios_pendentes(jogador['id'])
        
        if not desafios_pendentes:
            return """✅ *Aceitar Desafio*

Você não tem nenhum desafio pendente para aceitar.

_Digite *0* para voltar ao menu._"""
        
        if len(desafios_pendentes) == 1 and not numeros:
            desafio = desafios_pendentes[0]
            sucesso, mensagem_retorno = aceitar_desafio(desafio['id'], jogador['id'])
            
            if sucesso:
                try:
                    notificar_desafio_aceito_bot(desafio['id'])
                except:
                    pass
                
                return f"""✅ *Desafio Aceito!*

Você aceitou o desafio de *{desafio['challenger_name']}* (posição {desafio['challenger_position']}º).

📅 Data agendada: {desafio['scheduled_date']}

Boa sorte! 🏌️

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        if numeros:
            challenge_id = int(numeros[0])
            sucesso, mensagem_retorno = aceitar_desafio(challenge_id, jogador['id'])
            
            if sucesso:
                try:
                    notificar_desafio_aceito_bot(challenge_id)
                except:
                    pass
                return f"""✅ *Desafio #{challenge_id} aceito com sucesso!*

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        lista = "\n".join([f"   #{d['id']} - {d['challenger_name']} (posição {d['challenger_position']}º)" for d in desafios_pendentes])
        return f"""✅ *Aceitar Desafio*

Você tem {len(desafios_pendentes)} desafios pendentes:
{lista}

Para aceitar, digite: *4 [número]*
Exemplo: *4 123*

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [5]: Rejeitar desafio
    if msg == '5' or any(palavra in msg for palavra in ['rejeitar', 'rejeito', 'recusar', 'recuso']):
        numeros = re.findall(r'\d+', msg)
        if numeros and numeros[0] == '5' and len(msg) <= 2:
            numeros = []
        
        desafios_pendentes = get_desafios_pendentes(jogador['id'])
        
        if not desafios_pendentes:
            return """❌ *Rejeitar Desafio*

Você não tem nenhum desafio pendente para rejeitar.

_Digite *0* para voltar ao menu._"""
        
        if len(desafios_pendentes) == 1 and not numeros:
            desafio = desafios_pendentes[0]
            sucesso, mensagem_retorno = rejeitar_desafio(desafio['id'], jogador['id'])
            
            if sucesso:
                return f"""⚠️ *Desafio Rejeitado*

Você rejeitou o desafio de *{desafio['challenger_name']}*.

WO aplicado - você perdeu a posição.

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        if numeros:
            challenge_id = int(numeros[0])
            sucesso, mensagem_retorno = rejeitar_desafio(challenge_id, jogador['id'])
            
            if sucesso:
                return f"""⚠️ *Desafio #{challenge_id} rejeitado.* WO aplicado.

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        lista = "\n".join([f"   #{d['id']} - {d['challenger_name']} (posição {d['challenger_position']}º)" for d in desafios_pendentes])
        return f"""❌ *Rejeitar Desafio*

⚠️ *ATENÇÃO*: Rejeitar resulta em WO!

Seus desafios pendentes:
{lista}

Para rejeitar, digite: *5 [número]*
Exemplo: *5 123*

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [6]: Criar desafio - NOVO!
    if msg == '6' or any(palavra in msg for palavra in ['criar desafio', 'desafiar', 'novo desafio', 'quero desafiar']):
        if tem_desafio_ativo(jogador['id']):
            return """⚠️ *Você já tem um desafio ativo!*

Conclua seu desafio atual antes de criar um novo.

Digite *3* para ver seus desafios.

_Digite *0* para voltar ao menu._"""
        
        possiveis = get_possiveis_desafiados(jogador['id'])
        
        if not possiveis:
            return f"""🎯 *Criar Desafio*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

❌ No momento não há jogadores disponíveis para desafio.

_Digite *0* para voltar ao menu._"""
        
        linhas = []
        for i, p in enumerate(possiveis, 1):
            linhas.append(f"   *{i}* - {p['name']} ({p['position']}º)")
        
        lista = "\n".join(linhas)
        
        set_chat_state(telefone_normalizado, 'selecionando_oponente', {
            'possiveis': possiveis
        })
        
        return f"""🎯 *Criar Desafio*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

Selecione quem você quer desafiar:
{lista}

Digite o *número* do oponente (ex: *1*)

_Digite *0* para cancelar._"""
    
    # MENU PRINCIPAL [0]
    return f"""🏌️ *Liga Olímpica de Golfe*

Olá, *{jogador['name']}*!
📊 Posição atual: *{jogador['position']}º*

━━━━━━━━━━━━━━━━━━━━━
*MENU DE OPÇÕES*
━━━━━━━━━━━━━━━━━━━━━

*[1]* 📊 Minha posição
*[2]* 🎯 Quem posso desafiar
*[3]* 📋 Meus desafios
*[4]* ✅ Aceitar desafio
*[5]* ❌ Rejeitar desafio
*[6]* ⚔️ *Criar desafio*

━━━━━━━━━━━━━━━━━━━━━

_Digite o número da opção desejada._"""


# ============================================================
# FUNÇÃO PARA ENVIAR MENSAGEM
# ============================================================

def enviar_mensagem_whatsapp(destinatario, mensagem):
    """Envia mensagem para um número ou grupo"""
    import requests
    
    url = f"{EVOLUTION_API_URL}/message/sendText/{EVOLUTION_INSTANCE}"
    
    headers = {
        "apikey": EVOLUTION_API_KEY,
        "Content-Type": "application/json"
    }
    
    payload = {
        "number": destinatario,
        "text": mensagem
    }
    
    try:
        response = requests.post(url, json=payload, headers=headers, timeout=30)
        return response.status_code == 200 or response.status_code == 201
    except Exception as e:
        print(f"Erro ao enviar WhatsApp: {e}")
        return False


# ============================================================
# CONFIGURAÇÕES (copiar do whatsapp_integration.py)
# ============================================================

EVOLUTION_API_URL = "http://159.89.35.66:8080"
EVOLUTION_API_KEY = "liga-golf-api-key-2024"
EVOLUTION_INSTANCE = "liga-golf"
WHATSAPP_GRUPO_LIGA = "120363403838797386@g.us"


@app.route('/criar-coluna-tipo-membro')
@login_required
def criar_coluna_tipo_membro():
    """Cria a coluna para diferenciar jogadores de membros VIP"""
    if not session.get('is_admin', False):
        flash('Acesso restrito.', 'error')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    try:
        columns = conn.execute("PRAGMA table_info(players)").fetchall()
        column_names = [col[1] for col in columns]
        
        if 'tipo_membro' not in column_names:
            conn.execute("ALTER TABLE players ADD COLUMN tipo_membro TEXT DEFAULT 'jogador'")
            conn.commit()
            flash('✅ Coluna tipo_membro criada com sucesso!', 'success')
        else:
            flash('ℹ️ Coluna já existe.', 'info')
    except Exception as e:
        flash(f'❌ Erro: {str(e)}', 'error')
    finally:
        conn.close()
    
    return redirect(url_for('admin_dashboard'))


# ============================================
# ROTA: Toggle Tipo de Membro (Jogador <-> VIP)
# ============================================
# Adicionar no app.py
#
# O problema: coluna 'position' tem NOT NULL constraint
# Solução: usar position = 0 para VIPs (ao invés de NULL)
# ============================================

# ============================================
# ROTA: Toggle Tipo de Membro (Jogador <-> VIP)
# ============================================
# CORRIGIDO: usa position = 0 e tier = '' (string vazia)
# ao invés de NULL para evitar constraint errors
# ============================================

@app.route('/admin/toggle-tipo-membro/<int:player_id>', methods=['POST'])
@login_required
def toggle_tipo_membro(player_id):
    """Alterna entre Jogador e Membro VIP"""
    if not session.get('is_admin'):
        flash('Acesso negado. Apenas administradores.', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    player = conn.execute('SELECT * FROM players WHERE id = ?', (player_id,)).fetchone()
    
    if not player:
        conn.close()
        flash('Jogador não encontrado.', 'danger')
        return redirect(url_for('index'))
    
    tipo_atual = player['tipo_membro'] or 'jogador'
    sexo = player['sexo'] or 'masculino'
    
    try:
        if tipo_atual == 'jogador':
            # ===== CONVERTER PARA VIP =====
            # 1. Guardar posição atual para reordenar
            posicao_atual = player['position']
            
            # 2. Converter para VIP (position = 0, tier = '' string vazia)
            conn.execute('''
                UPDATE players 
                SET tipo_membro = 'vip', 
                    position = 0,
                    tier = ''
                WHERE id = ?
            ''', (player_id,))
            
            # 3. Reordenar jogadores do mesmo sexo que estavam abaixo
            if posicao_atual and posicao_atual > 0:
                conn.execute('''
                    UPDATE players 
                    SET position = position - 1 
                    WHERE position > ? 
                    AND (sexo = ? OR (sexo IS NULL AND ? = 'masculino'))
                    AND active = 1
                    AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
                ''', (posicao_atual, sexo, sexo))
                
                # 4. Recalcular tiers para o sexo afetado
                recalcular_tiers_por_sexo(conn, sexo)
            
            conn.commit()
            flash(f'{player["name"]} convertido para Membro VIP com sucesso!', 'success')
        
        else:
            # ===== CONVERTER PARA JOGADOR =====
            # 1. Encontrar última posição do ranking do mesmo sexo
            ultima_pos = conn.execute('''
                SELECT COALESCE(MAX(position), 0) as max_pos 
                FROM players 
                WHERE active = 1 
                AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
                AND (sexo = ? OR (sexo IS NULL AND ? = 'masculino'))
                AND position > 0
            ''', (sexo, sexo)).fetchone()['max_pos']
            
            nova_posicao = ultima_pos + 1
            
            # 2. Calcular tier baseado na nova posição
            novo_tier = calcular_tier(nova_posicao)
            
            # 3. Converter para jogador
            conn.execute('''
                UPDATE players 
                SET tipo_membro = 'jogador', 
                    position = ?,
                    tier = ?
                WHERE id = ?
            ''', (nova_posicao, novo_tier, player_id))
            
            conn.commit()
            flash(f'{player["name"]} convertido para Jogador na posição {nova_posicao}!', 'success')
    
    except Exception as e:
        conn.rollback()
        flash(f'Erro ao converter tipo de membro: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('player_detail', player_id=player_id))


def recalcular_tiers_por_sexo(conn, sexo):
    """Recalcula os tiers de todos os jogadores de um sexo"""
    jogadores = conn.execute('''
        SELECT id, position FROM players 
        WHERE active = 1 
        AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
        AND (sexo = ? OR (sexo IS NULL AND ? = 'masculino'))
        AND position > 0
        ORDER BY position
    ''', (sexo, sexo)).fetchall()
    
    for jogador in jogadores:
        novo_tier = calcular_tier(jogador['position'])
        conn.execute('UPDATE players SET tier = ? WHERE id = ?', 
                    (novo_tier, jogador['id']))


def calcular_tier(position):
    """Calcula o tier baseado na posição (pirâmide 5-7-9-11...)"""
    if position is None or position <= 0:
        return ''  # String vazia para VIPs/inativos
    
    # Tier A: posições 1-5
    if position <= 5:
        return 'A'
    # Tier B: posições 6-12 (7 jogadores)
    elif position <= 12:
        return 'B'
    # Tier C: posições 13-21 (9 jogadores)
    elif position <= 21:
        return 'C'
    # Tier D: posições 22-32 (11 jogadores)
    elif position <= 32:
        return 'D'
    # Tier E: posições 33-46 (13 jogadores)
    elif position <= 46:
        return 'E'
    # Tier F: posições 47-62 (15 jogadores)
    elif position <= 62:
        return 'F'
    # Tier G: posições 63+ (17 jogadores)
    else:
        return 'G'



# ============================================
# ROTA ADMIN: Corrigir Banco de Dados VIP
# ============================================
# Adicione esta rota no seu app.py
# Acesse: /admin/corrigir-vip
# ============================================

@app.route('/admin/corrigir-vip')
@login_required
def corrigir_vip():
    """Corrige problemas no banco relacionados ao sistema VIP"""
    if not session.get('is_admin'):
        flash('Acesso negado. Apenas administradores.', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    relatorio = []
    
    try:
        # 1. Corrigir VIPs com position inválido
        cursor = conn.execute('''
            UPDATE players 
            SET position = 0 
            WHERE tipo_membro = 'vip' 
            AND (position IS NULL OR position != 0)
        ''')
        relatorio.append(f"VIPs corrigidos (position → 0): {cursor.rowcount}")
        
        # 2. Reordenar ranking masculino
        jogadores_masc = conn.execute('''
            SELECT id FROM players
            WHERE active = 1
            AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
            AND (sexo = 'masculino' OR sexo IS NULL OR sexo = '')
            ORDER BY 
                CASE WHEN position > 0 THEN position ELSE 9999 END,
                name
        ''').fetchall()
        
        for i, jogador in enumerate(jogadores_masc, 1):
            conn.execute('UPDATE players SET position = ? WHERE id = ?', (i, jogador['id']))
        relatorio.append(f"Ranking masculino reordenado: {len(jogadores_masc)} jogadores")
        
        # 3. Reordenar ranking feminino
        jogadoras_fem = conn.execute('''
            SELECT id FROM players
            WHERE active = 1
            AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
            AND sexo = 'feminino'
            ORDER BY 
                CASE WHEN position > 0 THEN position ELSE 9999 END,
                name
        ''').fetchall()
        
        for i, jogadora in enumerate(jogadoras_fem, 1):
            conn.execute('UPDATE players SET position = ? WHERE id = ?', (i, jogadora['id']))
        relatorio.append(f"Ranking feminino reordenado: {len(jogadoras_fem)} jogadoras")
        
        # 4. Recalcular tiers
        def calcular_tier(pos):
            if pos <= 5: return 'A'
            elif pos <= 12: return 'B'
            elif pos <= 21: return 'C'
            elif pos <= 32: return 'D'
            elif pos <= 46: return 'E'
            elif pos <= 62: return 'F'
            else: return 'G'
        
        jogadores_ativos = conn.execute('''
            SELECT id, position FROM players
            WHERE active = 1
            AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
            AND position > 0
        ''').fetchall()
        
        for jogador in jogadores_ativos:
            novo_tier = calcular_tier(jogador['position'])
            conn.execute('UPDATE players SET tier = ? WHERE id = ?', (novo_tier, jogador['id']))
        relatorio.append(f"Tiers recalculados: {len(jogadores_ativos)} jogadores")
        
        # 5. Limpar tier de VIPs e inativos
        conn.execute("UPDATE players SET tier = NULL WHERE tipo_membro = 'vip'")
        conn.execute("UPDATE players SET tier = NULL WHERE active = 0")
        relatorio.append("Tiers de VIPs e inativos limpos")
        
        # 6. Verificar problemas restantes
        problemas = conn.execute('''
            SELECT id, name, position, tipo_membro, active
            FROM players
            WHERE (
                (active = 1 AND (tipo_membro = 'jogador' OR tipo_membro IS NULL) AND (position IS NULL OR position <= 0))
                OR
                (tipo_membro = 'vip' AND position != 0)
            )
        ''').fetchall()
        
        if problemas:
            relatorio.append(f"⚠️ PROBLEMAS RESTANTES: {len(problemas)}")
            for p in problemas:
                relatorio.append(f"  - {p['name']} (id={p['id']}, pos={p['position']}, tipo={p['tipo_membro']})")
        else:
            relatorio.append("✅ Nenhum problema encontrado!")
        
        conn.commit()
        
        # Exibir relatório
        flash('Correção executada com sucesso!', 'success')
        for linha in relatorio:
            flash(linha, 'info')
            
    except Exception as e:
        conn.rollback()
        flash(f'Erro durante correção: {str(e)}', 'danger')
    finally:
        conn.close()
    
    return redirect(url_for('index'))


# ============================================
# ROTA ALTERNATIVA: Ver diagnóstico sem corrigir
# ============================================

@app.route('/admin/diagnostico-vip')
@login_required  
def diagnostico_vip():
    """Mostra diagnóstico do banco sem fazer alterações"""
    if not session.get('is_admin'):
        flash('Acesso negado.', 'danger')
        return redirect(url_for('index'))
    
    conn = get_db_connection()
    
    # Estatísticas
    stats = {
        'total_jogadores': conn.execute('SELECT COUNT(*) FROM players').fetchone()[0],
        'ativos_masc': conn.execute('''
            SELECT COUNT(*) FROM players 
            WHERE active = 1 AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
            AND (sexo = 'masculino' OR sexo IS NULL OR sexo = '') AND position > 0
        ''').fetchone()[0],
        'ativos_fem': conn.execute('''
            SELECT COUNT(*) FROM players 
            WHERE active = 1 AND (tipo_membro = 'jogador' OR tipo_membro IS NULL)
            AND sexo = 'feminino' AND position > 0
        ''').fetchone()[0],
        'vips': conn.execute("SELECT COUNT(*) FROM players WHERE tipo_membro = 'vip'").fetchone()[0],
        'inativos': conn.execute('SELECT COUNT(*) FROM players WHERE active = 0').fetchone()[0],
    }
    
    # Problemas
    problemas = conn.execute('''
        SELECT id, name, position, tipo_membro, active, sexo
        FROM players
        WHERE (
            (active = 1 AND (tipo_membro = 'jogador' OR tipo_membro IS NULL) AND (position IS NULL OR position <= 0))
            OR
            (tipo_membro = 'vip' AND (position IS NULL OR position != 0))
        )
    ''').fetchall()
    
    # VIPs atuais
    vips = conn.execute('''
        SELECT id, name, position, tipo_membro
        FROM players WHERE tipo_membro = 'vip'
        ORDER BY name
    ''').fetchall()
    
    conn.close()
    
    # Montar HTML simples
    html = f'''
    <!DOCTYPE html>
    <html>
    <head>
        <title>Diagnóstico VIP</title>
        <link href="https://cdn.jsdelivr.net/npm/bootstrap@5.3.0/dist/css/bootstrap.min.css" rel="stylesheet">
    </head>
    <body class="container py-4">
        <h1>🔍 Diagnóstico do Banco de Dados</h1>
        
        <div class="card mb-3">
            <div class="card-header bg-primary text-white">📊 Estatísticas</div>
            <div class="card-body">
                <ul>
                    <li>Total de registros: <strong>{stats['total_jogadores']}</strong></li>
                    <li>Jogadores masculinos ativos: <strong>{stats['ativos_masc']}</strong></li>
                    <li>Jogadoras femininas ativas: <strong>{stats['ativos_fem']}</strong></li>
                    <li>Membros VIP: <strong>{stats['vips']}</strong></li>
                    <li>Inativos: <strong>{stats['inativos']}</strong></li>
                </ul>
            </div>
        </div>
        
        <div class="card mb-3">
            <div class="card-header bg-warning">⚠️ Problemas Encontrados ({len(problemas)})</div>
            <div class="card-body">
                {'<p class="text-success">Nenhum problema encontrado! ✅</p>' if not problemas else ''}
                {'<table class="table table-sm"><thead><tr><th>ID</th><th>Nome</th><th>Position</th><th>Tipo</th><th>Ativo</th></tr></thead><tbody>' + ''.join([f"<tr><td>{p['id']}</td><td>{p['name']}</td><td>{p['position']}</td><td>{p['tipo_membro']}</td><td>{p['active']}</td></tr>" for p in problemas]) + '</tbody></table>' if problemas else ''}
            </div>
        </div>
        
        <div class="card mb-3">
            <div class="card-header bg-warning text-dark">⭐ Membros VIP ({len(vips)})</div>
            <div class="card-body">
                {'<p class="text-muted">Nenhum VIP cadastrado.</p>' if not vips else ''}
                {'<table class="table table-sm"><thead><tr><th>ID</th><th>Nome</th><th>Position</th></tr></thead><tbody>' + ''.join([f"<tr><td>{v['id']}</td><td>{v['name']}</td><td>{v['position']}</td></tr>" for v in vips]) + '</tbody></table>' if vips else ''}
            </div>
        </div>
        
        <div class="d-flex gap-2">
            <a href="{url_for('corrigir_vip')}" class="btn btn-danger" onclick="return confirm('Executar correção automática?')">
                🔧 Executar Correção
            </a>
            <a href="{url_for('index')}" class="btn btn-secondary">Voltar</a>
        </div>
    </body>
    </html>
    '''
    
    return html


@app.route('/admin/corrigir_prazos')
@login_required
def corrigir_prazos():
    if not session.get('is_admin'):
        return "Acesso negado", 403
    
    conn = get_db_connection()
    conn.execute('''
        UPDATE challenges 
        SET response_deadline = datetime(created_at, '+2 days')
        WHERE status = 'pending'
    ''')
    conn.commit()
    conn.close()
    
    flash('Prazos corrigidos para 2 dias!', 'success')
    return redirect(url_for('admin_dashboard'))


# ============================================================
# WEBHOOK DO WHATSAPP - RECEBE MENSAGENS
# ============================================================

@app.route('/webhook/whatsapp', methods=['POST'])
def webhook_whatsapp():
    """Recebe mensagens do WhatsApp via Evolution API"""
    try:
        data = request.json
        
        # Log para debug
        print(f"[Webhook] Dados recebidos: {data}")
        
        # Verificar se é uma mensagem recebida
        if data.get('event') == 'messages.upsert':
            message_data = data.get('data', {})
            
            # Ignorar mensagens enviadas por nós mesmos
            if message_data.get('key', {}).get('fromMe'):
                return jsonify({'status': 'ignored', 'reason': 'own_message'})
            
            # Ignorar mensagens de grupo
            remote_jid = message_data.get('key', {}).get('remoteJid', '')
            if '@g.us' in remote_jid:
                return jsonify({'status': 'ignored', 'reason': 'group_message'})
            
            # Extrair telefone e mensagem
            telefone = extrair_telefone_do_jid(remote_jid)
            
            # Extrair texto da mensagem
            message_content = message_data.get('message', {})
            mensagem = (
                message_content.get('conversation') or 
                message_content.get('extendedTextMessage', {}).get('text') or
                ''
            )
            
            if not mensagem or not telefone:
                return jsonify({'status': 'ignored', 'reason': 'empty_message'})
            
            print(f"[Webhook] Mensagem de {telefone}: {mensagem}")
            
            # Processar comando e obter resposta
            resposta = processar_comando_whatsapp_v2(mensagem, telefone)
            
            # Enviar resposta
            if resposta:
                rodape = "\n\n━━━━━━━━━━━━━━━━━━━━━\n🌐 Visite: www.ligaolimpicadegolfe.com.br"
                enviar_mensagem_whatsapp(f"55{telefone}@s.whatsapp.net", resposta + rodape)
            
            return jsonify({'status': 'processed'})
        
        return jsonify({'status': 'ok'})
        
    except Exception as e:
        print(f"[Webhook] Erro: {e}")
        import traceback
        traceback.print_exc()
        return jsonify({'status': 'error', 'message': str(e)}), 500


# ============================================================
# SISTEMA DE PROPOSTA DE DATAS ALTERNATIVAS
# ============================================================

def salvar_proposta_datas(challenge_id, data1, data2, propositor_id):
    """Salva as datas propostas pelo desafiado"""
    conn = get_db_connection()
    
    conn.execute('''
        UPDATE challenges 
        SET status = 'awaiting_date_confirmation',
            proposed_date_1 = ?,
            proposed_date_2 = ?,
            date_proposed_by = ?,
            date_proposed_at = ?
        WHERE id = ?
    ''', (data1, data2, propositor_id, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), challenge_id))
    
    conn.commit()
    conn.close()


def get_proposta_pendente_para_desafiante(player_id):
    """Busca propostas de data pendentes onde o jogador é o desafiante"""
    conn = get_db_connection()
    
    proposta = conn.execute('''
        SELECT c.*, 
               challenged.name as challenged_name,
               challenged.position as challenged_position
        FROM challenges c
        JOIN players challenged ON c.challenged_id = challenged.id
        WHERE c.challenger_id = ?
        AND c.status = 'awaiting_date_confirmation'
    ''', (player_id,)).fetchone()
    
    conn.close()
    return dict(proposta) if proposta else None


def aceitar_data_proposta(challenge_id, data_escolhida):
    """Aceita uma das datas propostas e confirma o desafio"""
    conn = get_db_connection()
    
    conn.execute('''
        UPDATE challenges 
        SET status = 'accepted',
            scheduled_date = ?,
            updated_at = ?
        WHERE id = ?
    ''', (data_escolhida, datetime.now().strftime('%Y-%m-%d %H:%M:%S'), challenge_id))
    
    conn.commit()
    conn.close()


def cancelar_desafio_sem_penalidade(challenge_id):
    """Cancela o desafio sem prejuízo para nenhuma das partes"""
    conn = get_db_connection()
    
    conn.execute('''
        UPDATE challenges 
        SET status = 'cancelled',
            updated_at = ?
        WHERE id = ?
    ''', (datetime.now().strftime('%Y-%m-%d %H:%M:%S'), challenge_id))
    
    conn.commit()
    conn.close()


def notificar_cancelamento_sem_penalidade(challenge_id):
    """Notifica ambas as partes sobre o cancelamento sem prejuízo"""
    conn = get_db_connection()
    
    challenge = conn.execute('''
        SELECT c.*, 
               challenger.name as challenger_name,
               challenger.telefone as challenger_telefone,
               challenged.name as challenged_name,
               challenged.telefone as challenged_telefone,
               challenged.position as challenged_position
        FROM challenges c
        JOIN players challenger ON c.challenger_id = challenger.id
        JOIN players challenged ON c.challenged_id = challenged.id
        WHERE c.id = ?
    ''', (challenge_id,)).fetchone()
    
    conn.close()
    
    if not challenge:
        return False
    
    # Notificar desafiado
    telefone_desafiado = challenge['challenged_telefone']
    if telefone_desafiado:
        msg_desafiado = f"""❌ *DESAFIO CANCELADO*

*{challenge['challenger_name']}* não aceitou as datas propostas.

O desafio foi cancelado *sem prejuízo* para nenhuma das partes.

Vocês podem criar um novo desafio quando quiserem! 🏌️

_Digite *0* para voltar ao menu._"""
        
        telefone_norm = normalizar_telefone(telefone_desafiado)
        enviar_mensagem_whatsapp(f"55{telefone_norm}@s.whatsapp.net", msg_desafiado)
    
    # Notificar no grupo da liga
    msg_grupo = f"""❌ *DESAFIO CANCELADO*

⚔️ *{challenge['challenger_name']}* vs *{challenge['challenged_name']}*

O desafiante não aceitou as datas propostas.
Desafio cancelado *sem prejuízo* para ambos."""
    
    enviar_mensagem_whatsapp(WHATSAPP_GRUPO_LIGA, msg_grupo)
    
    return True




def notificar_proposta_datas(challenge_id):
    """Notifica o desafiante sobre as datas propostas"""
    conn = get_db_connection()
    
    challenge = conn.execute('''
        SELECT c.*, 
               challenger.name as challenger_name,
               challenger.telefone as challenger_telefone,
               challenged.name as challenged_name,
               challenged.position as challenged_position
        FROM challenges c
        JOIN players challenger ON c.challenger_id = challenger.id
        JOIN players challenged ON c.challenged_id = challenged.id
        WHERE c.id = ?
    ''', (challenge_id,)).fetchone()
    
    conn.close()
    
    if not challenge:
        return False
    
    # Formatar datas
    try:
        data1_obj = datetime.strptime(challenge['proposed_date_1'], '%Y-%m-%d')
        data1_fmt = data1_obj.strftime('%d/%m/%Y')
    except:
        data1_fmt = challenge['proposed_date_1']
    
    try:
        data2_obj = datetime.strptime(challenge['proposed_date_2'], '%Y-%m-%d')
        data2_fmt = data2_obj.strftime('%d/%m/%Y')
    except:
        data2_fmt = challenge['proposed_date_2']
    
    # Notificar desafiante via WhatsApp
    telefone_desafiante = challenge['challenger_telefone']
    if telefone_desafiante:
        msg = f"""📅 *PROPOSTA DE NOVAS DATAS*

*{challenge['challenged_name']}* ({challenge['challenged_position']}º) propôs novas datas para o desafio:

*[A]* {data1_fmt}
*[B]* {data2_fmt}
*[C]* ❌ Cancelar desafio (sem prejuízo)

━━━━━━━━━━━━━━━━━━━━━
Digite *A*, *B* ou *C* para responder.

⏰ _Você tem 2 dias para responder._"""
        
        telefone_normalizado = normalizar_telefone(telefone_desafiante)
        enviar_mensagem_whatsapp(f"55{telefone_normalizado}@s.whatsapp.net", msg)
    
    # Notificar no grupo da liga
    msg_grupo = f"""📅 *PROPOSTA DE NOVAS DATAS*

*{challenge['challenged_name']}* propôs novas datas:

⚔️ *{challenge['challenger_name']}* vs *{challenge['challenged_name']}*

📆 Opção A: {data1_fmt}
📆 Opção B: {data2_fmt}

Aguardando escolha do desafiante... ⏳"""
    
    enviar_mensagem_whatsapp(WHATSAPP_GRUPO_LIGA, msg_grupo)
    
    return True
```

---

## RESUMO DAS ALTERAÇÕES:

1. ✅ **Nova função** `cancelar_desafio_sem_penalidade()` - muda status para `cancelled`
2. ✅ **Nova função** `notificar_cancelamento_sem_penalidade()` - avisa desafiado e grupo
3. ✅ **Modificado** bloco `if msg in ['a', 'b']:` → `if msg in ['a', 'b', 'c']:`
4. ✅ **Modificado** `notificar_proposta_datas()` - inclui opção C na mensagem

---

## FLUXO FINAL:

**Desafiante recebe:**
```
📅 PROPOSTA DE NOVAS DATAS

João (5º) propôs novas datas para o desafio:

[A] 25/02/2026
[B] 27/02/2026
[C] ❌ Cancelar desafio (sem prejuízo)

━━━━━━━━━━━━━━━━━━━━━
Digite A, B ou C para responder.


def notificar_data_confirmada(challenge_id, data_escolhida):
    """Notifica ambos os jogadores que a data foi confirmada"""
    conn = get_db_connection()
    
    challenge = conn.execute('''
        SELECT c.*, 
               challenger.name as challenger_name,
               challenger.telefone as challenger_telefone,
               challenged.name as challenged_name,
               challenged.telefone as challenged_telefone,
               challenged.position as challenged_position
        FROM challenges c
        JOIN players challenger ON c.challenger_id = challenger.id
        JOIN players challenged ON c.challenged_id = challenged.id
        WHERE c.id = ?
    ''', (challenge_id,)).fetchone()
    
    conn.close()
    
    if not challenge:
        return
    
    # Formatar data
    try:
        data_obj = datetime.strptime(data_escolhida, '%Y-%m-%d')
        data_fmt = data_obj.strftime('%d/%m/%Y')
    except:
        data_fmt = data_escolhida
    
    msg = f"""✅ *DATA CONFIRMADA!*

O desafio entre *{challenge['challenger_name']}* e *{challenge['challenged_name']}* está confirmado!

📅 Data: *{data_fmt}*

Boa sorte a ambos! 🏌️"""
    
    # Notificar desafiado
    telefone_desafiado = challenge['challenged_telefone']
    if telefone_desafiado:
        telefone_norm = normalizar_telefone(telefone_desafiado)
        enviar_mensagem_whatsapp(f"55{telefone_norm}@s.whatsapp.net", msg)
    
    # Notificar no grupo
    enviar_mensagem_whatsapp(WHATSAPP_GRUPO_LIGA, msg)


# ============================================================
# FUNÇÃO PROCESSAR_COMANDO_WHATSAPP - VERSÃO ATUALIZADA
# ============================================================
# SUBSTITUI a função processar_comando_whatsapp existente
# ============================================================

def processar_comando_whatsapp_v2(mensagem, telefone):
    """Processa mensagem recebida e retorna resposta - VERSÃO COM PROPOSTA DE DATAS"""
    
    # Normalizar mensagem
    msg = mensagem.lower().strip()
    telefone_normalizado = normalizar_telefone(telefone)
    
    # Buscar jogador pelo telefone
    jogador = get_player_by_phone(telefone)
    
    if not jogador:
        return """❌ *Número não cadastrado*

Seu número de WhatsApp não está vinculado a nenhum jogador da Liga.

Para cadastrar, acesse seu perfil no site e adicione seu número no campo "WhatsApp para Notificações"."""

# ---------------------------------------------------------
    # VERIFICAR SE HÁ PROPOSTA DE DATA PENDENTE (desafiante)
    # ---------------------------------------------------------
    if msg in ['a', 'b', 'c']:
        proposta = get_proposta_pendente_para_desafiante(jogador['id'])
        
        if proposta:
            # Opção C - Cancelar sem prejuízo
            if msg == 'c':
                cancelar_desafio_sem_penalidade(proposta['id'])
                
                # Notificar
                try:
                    notificar_cancelamento_sem_penalidade(proposta['id'])
                except Exception as e:
                    print(f"Erro ao notificar cancelamento: {e}")
                
                return f"""❌ *DESAFIO CANCELADO*

Você optou por não aceitar as datas propostas por *{proposta['challenged_name']}*.

O desafio foi cancelado *sem prejuízo* para nenhuma das partes.

Vocês podem criar um novo desafio quando quiserem! 🏌️

_Digite *0* para voltar ao menu._"""
            
            # Opção A ou B - Aceitar uma das datas
            if msg == 'a':
                data_escolhida = proposta['proposed_date_1']
            else:
                data_escolhida = proposta['proposed_date_2']
            
            # Aceitar a data
            aceitar_data_proposta(proposta['id'], data_escolhida)
            
            # Notificar
            try:
                notificar_data_confirmada(proposta['id'], data_escolhida)
            except Exception as e:
                print(f"Erro ao notificar: {e}")
            
            # Formatar data
            try:
                data_obj = datetime.strptime(data_escolhida, '%Y-%m-%d')
                data_fmt = data_obj.strftime('%d/%m/%Y')
            except:
                data_fmt = data_escolhida
            
            return f"""✅ *DATA CONFIRMADA!*

Você aceitou a data *{data_fmt}* para o desafio contra *{proposta['challenged_name']}*.

O desafio está confirmado! Boa sorte! 🏌️

_Digite *0* para voltar ao menu._"""


    
    # ---------------------------------------------------------
    # VERIFICAR SE HÁ ESTADO PENDENTE (conversa em andamento)
    # ---------------------------------------------------------
    estado_atual = get_chat_state(telefone_normalizado)
    
    if estado_atual:
        estado = estado_atual['estado']
        dados = estado_atual['dados']
        
        # ---------------------------------------------------------
        # ESTADO: Selecionando oponente para desafio
        # ---------------------------------------------------------
        if estado == 'selecionando_oponente':
            if msg == '0' or 'cancelar' in msg:
                clear_chat_state(telefone_normalizado)
                return "❌ Criação de desafio cancelada.\n\n_Digite *0* para ver o menu._"
            
            try:
                opcao = int(msg)
                possiveis = dados.get('possiveis', [])
                
                if opcao < 1 or opcao > len(possiveis):
                    return f"""⚠️ Opção inválida!

Digite um número de *1* a *{len(possiveis)}* para selecionar o oponente.

Ou digite *0* para cancelar."""
                
                oponente = possiveis[opcao - 1]
                
                set_chat_state(telefone_normalizado, 'informando_data', {
                    'oponente_id': oponente['id'],
                    'oponente_nome': oponente['name'],
                    'oponente_posicao': oponente['position']
                })
                
                hoje = datetime.now()
                data_max = hoje + timedelta(days=7)
                
                return f"""✅ Oponente selecionado: *{oponente['name']}* ({oponente['position']}º)

📅 *Qual a data do jogo?*

Digite no formato *DD/MM* (ex: {data_max.strftime('%d/%m')})

A data deve ser nos próximos *7 dias*.
(até {data_max.strftime('%d/%m/%Y')})

_Digite *0* para cancelar._"""
                
            except ValueError:
                return """⚠️ Digite apenas o *número* do oponente.

Exemplo: *1* ou *2* ou *3*

_Digite *0* para cancelar._"""
        
        # ---------------------------------------------------------
        # ESTADO: Informando data do jogo (criar desafio)
        # ---------------------------------------------------------
        elif estado == 'informando_data':
            if msg == '0' or 'cancelar' in msg:
                clear_chat_state(telefone_normalizado)
                return "❌ Criação de desafio cancelada.\n\n_Digite *0* para ver o menu._"
            
            data_jogo = parse_data_input(msg)
            
            if not data_jogo:
                return """⚠️ Formato de data inválido!

Digite no formato *DD/MM* (ex: 25/02)

_Digite *0* para cancelar._"""
            
            hoje = datetime.now().date()
            
            if data_jogo < hoje:
                return """⚠️ A data não pode ser no passado!

Digite uma data a partir de hoje.

_Digite *0* para cancelar._"""
            
            data_max = hoje + timedelta(days=7)
            if data_jogo > data_max:
                return f"""⚠️ A data não pode ser superior a 7 dias!

Data máxima permitida: *{data_max.strftime('%d/%m/%Y')}*

_Digite *0* para cancelar._"""
            
            oponente_id = dados['oponente_id']
            oponente_nome = dados['oponente_nome']
            oponente_posicao = dados['oponente_posicao']
            data_formatada = data_jogo.strftime('%Y-%m-%d')
            
            sucesso, mensagem_retorno, challenge_id = criar_desafio_via_whatsapp(
                jogador['id'], 
                oponente_id, 
                data_formatada
            )
            
            clear_chat_state(telefone_normalizado)
            
            if sucesso:
                try:
                    notificar_desafio_criado_whatsapp(challenge_id)
                except Exception as e:
                    print(f"Erro ao notificar: {e}")
                
                return f"""🎉 *DESAFIO CRIADO COM SUCESSO!*

Você desafiou *{oponente_nome}* ({oponente_posicao}º)

📅 Data do jogo: *{data_jogo.strftime('%d/%m/%Y')}*
⏳ Prazo para resposta: *2 dias*

O desafiado será notificado e pode:
- Aceitar a data
- Rejeitar (perde por WO)
- Propor 2 novas datas

Boa sorte! 🏌️

_Digite *0* para voltar ao menu._"""
            else:
                return f"""❌ *Erro ao criar desafio*

{mensagem_retorno}

_Digite *0* para voltar ao menu._"""
        
        # ---------------------------------------------------------
        # ESTADO: Propondo primeira data alternativa
        # ---------------------------------------------------------
        elif estado == 'propondo_data_1':
            if msg == '0' or 'cancelar' in msg:
                clear_chat_state(telefone_normalizado)
                return "❌ Proposta cancelada.\n\n_Digite *0* para ver o menu._"
            
            data1 = parse_data_input(msg)
            
            if not data1:
                return """⚠️ Formato de data inválido!

Digite no formato *DD/MM* (ex: 25/02)

_Digite *0* para cancelar._"""
            
            hoje = datetime.now().date()
            
            if data1 < hoje:
                return """⚠️ A data não pode ser no passado!

_Digite *0* para cancelar._"""
            
            data_max = hoje + timedelta(days=7)
            if data1 > data_max:
                return f"""⚠️ A data deve ser em até 7 dias!

Data máxima: *{data_max.strftime('%d/%m/%Y')}*

_Digite *0* para cancelar._"""
            
            # Salvar primeira data e pedir segunda
            set_chat_state(telefone_normalizado, 'propondo_data_2', {
                'challenge_id': dados['challenge_id'],
                'challenger_name': dados['challenger_name'],
                'data_1': data1.strftime('%Y-%m-%d')
            })
            
            return f"""✅ Primeira data: *{data1.strftime('%d/%m/%Y')}*

Agora informe a *segunda opção* de data:

(formato DD/MM, máximo 7 dias)

_Digite *0* para cancelar._"""
        
        # ---------------------------------------------------------
        # ESTADO: Propondo segunda data alternativa
        # ---------------------------------------------------------
        elif estado == 'propondo_data_2':
            if msg == '0' or 'cancelar' in msg:
                clear_chat_state(telefone_normalizado)
                return "❌ Proposta cancelada.\n\n_Digite *0* para ver o menu._"
            
            data2 = parse_data_input(msg)
            
            if not data2:
                return """⚠️ Formato de data inválido!

Digite no formato *DD/MM* (ex: 27/02)

_Digite *0* para cancelar._"""
            
            hoje = datetime.now().date()
            
            if data2 < hoje:
                return """⚠️ A data não pode ser no passado!

_Digite *0* para cancelar._"""
            
            data_max = hoje + timedelta(days=7)
            if data2 > data_max:
                return f"""⚠️ A data deve ser em até 7 dias!

Data máxima: *{data_max.strftime('%d/%m/%Y')}*

_Digite *0* para cancelar._"""
            
            # Salvar ambas as datas
            challenge_id = dados['challenge_id']
            data1 = dados['data_1']
            
            salvar_proposta_datas(challenge_id, data1, data2.strftime('%Y-%m-%d'), jogador['id'])
            
            # Formatar data1 para exibição
            try:
                data1_obj = datetime.strptime(data1, '%Y-%m-%d')
                data1_fmt = data1_obj.strftime('%d/%m/%Y')
            except:
                data1_fmt = data1
            
            # Notificar desafiante e grupo
            notificacao_enviada = False
            try:
                notificacao_enviada = notificar_proposta_datas(challenge_id)
            except Exception as e:
                print(f"Erro ao notificar proposta: {e}")
            
            clear_chat_state(telefone_normalizado)
            
            # Mensagem de confirmação para o desafiado
            if notificacao_enviada:
                status_msg = "✅ *" + dados['challenger_name'] + " foi notificado!*"
            else:
                status_msg = "⚠️ Não foi possível notificar automaticamente."
            
            return f"""✅ *PROPOSTA DE DATAS ENVIADA!*

Você propôs as seguintes datas para *{dados['challenger_name']}*:

📆 *Opção A:* {data1_fmt}
📆 *Opção B:* {data2.strftime('%d/%m/%Y')}

━━━━━━━━━━━━━━━━━━━━━
{status_msg}

Ele deve escolher uma das datas em até *2 dias*.

Você será notificado quando a data for confirmada! 🏌️

_Digite *0* para voltar ao menu._"""
    
    # ---------------------------------------------------------
    # COMANDOS NORMAIS (sem estado pendente)
    # ---------------------------------------------------------
    
    # COMANDO [1]: Minha posição
    if msg == '1' or any(palavra in msg for palavra in ['posição', 'posicao', 'ranking', 'colocação', 'colocacao']):
        return f"""📊 *Sua Posição no Ranking*

Olá, {jogador['name']}!

Você está atualmente na posição *{jogador['position']}º* no ranking da Liga Olímpica de Golfe.

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [2]: Quem posso desafiar
    if msg == '2' or any(palavra in msg for palavra in ['quem posso', 'possiveis', 'possíveis']):
        possiveis = get_possiveis_desafiados(jogador['id'])
        
        if not possiveis:
            return f"""🎯 *Possíveis Desafiados*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

No momento não há jogadores disponíveis para desafio.

_Digite *0* para voltar ao menu._"""
        
        lista = "\n".join([f"   {p['position']}º - {p['name']}" for p in possiveis])
        
        return f"""🎯 *Possíveis Desafiados*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

Você pode desafiar:
{lista}

📱 Para criar um desafio, digite *6*

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [3]: Meus desafios
    if msg == '3' or (any(palavra in msg for palavra in ['meus desafio', 'meu desafio']) and 'criar' not in msg):
        desafios = get_meus_desafios(jogador['id'])
        
        # Verificar proposta de datas pendente
        proposta = get_proposta_pendente_para_desafiante(jogador['id'])
        
        if not desafios and not proposta:
            return f"""📋 *Meus Desafios*

Olá, {jogador['name']}!

Você não tem desafios ativos no momento.

_Digite *0* para voltar ao menu._"""
        
        linhas = []
        for d in desafios:
            if d['status'] == 'pending':
                status_emoji = "⏳"
                status_texto = "Aguardando resposta"
            elif d['status'] == 'accepted':
                status_emoji = "✅"
                status_texto = "Aceito"
            elif d['status'] == 'awaiting_date_confirmation':
                status_emoji = "📅"
                status_texto = "Aguardando escolha de data"
            else:
                status_emoji = "❓"
                status_texto = d['status']
            
            if d['challenger_id'] == jogador['id']:
                linhas.append(f"   {status_emoji} #{d['id']} - Você → {d['challenged_name']} ({d['challenged_position']}º) [{status_texto}]")
            else:
                linhas.append(f"   {status_emoji} #{d['id']} - {d['challenger_name']} ({d['challenger_position']}º) → Você [{status_texto}]")
        
        lista = "\n".join(linhas) if linhas else "   Nenhum desafio ativo."
        
        # Adicionar aviso de proposta pendente
        aviso_proposta = ""
        if proposta:
            try:
                data1_obj = datetime.strptime(proposta['proposed_date_1'], '%Y-%m-%d')
                data1_fmt = data1_obj.strftime('%d/%m/%Y')
                data2_obj = datetime.strptime(proposta['proposed_date_2'], '%Y-%m-%d')
                data2_fmt = data2_obj.strftime('%d/%m/%Y')
            except:
                data1_fmt = proposta['proposed_date_1']
                data2_fmt = proposta['proposed_date_2']
            
            aviso_proposta = f"""

📅 *PROPOSTA DE DATAS PENDENTE!*
{proposta['challenged_name']} propôs:
*[A]* {data1_fmt}
*[B]* {data2_fmt}

Digite *A* ou *B* para escolher."""
        
        pendentes_para_responder = [d for d in desafios if d['status'] == 'pending' and d['challenged_id'] == jogador['id']]
        
        dica = ""
        if pendentes_para_responder:
            dica = "\n\n💡 Digite *4* (aceitar), *5* (rejeitar) ou *7* (propor datas)"
        
        return f"""📋 *Meus Desafios*

Olá, {jogador['name']}!

Seus desafios ativos:
{lista}{aviso_proposta}{dica}

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [4]: Aceitar desafio
    if msg == '4' or 'aceitar' in msg or 'aceito' in msg:
        numeros = re.findall(r'\d+', msg)
        if numeros and numeros[0] == '4' and len(msg) <= 2:
            numeros = []
        
        desafios_pendentes = get_desafios_pendentes(jogador['id'])
        
        if not desafios_pendentes:
            return """✅ *Aceitar Desafio*

Você não tem nenhum desafio pendente para aceitar.

_Digite *0* para voltar ao menu._"""
        
        if len(desafios_pendentes) == 1 and not numeros:
            desafio = desafios_pendentes[0]
            sucesso, mensagem_retorno = aceitar_desafio(desafio['id'], jogador['id'])
            
            if sucesso:
                try:
                    notificar_desafio_aceito_bot(desafio['id'])
                except:
                    pass
                
                return f"""✅ *Desafio Aceito!*

Você aceitou o desafio de *{desafio['challenger_name']}* (posição {desafio['challenger_position']}º).

📅 Data agendada: {desafio['scheduled_date']}

Boa sorte! 🏌️

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        if numeros:
            challenge_id = int(numeros[0])
            sucesso, mensagem_retorno = aceitar_desafio(challenge_id, jogador['id'])
            
            if sucesso:
                try:
                    notificar_desafio_aceito_bot(challenge_id)
                except:
                    pass
                return f"""✅ *Desafio #{challenge_id} aceito com sucesso!*

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        lista = "\n".join([f"   #{d['id']} - {d['challenger_name']} (posição {d['challenger_position']}º)" for d in desafios_pendentes])
        return f"""✅ *Aceitar Desafio*

Você tem {len(desafios_pendentes)} desafios pendentes:
{lista}

Para aceitar, digite: *4 [número]*
Exemplo: *4 123*

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [5]: Rejeitar desafio
    if msg == '5' or any(palavra in msg for palavra in ['rejeitar', 'rejeito', 'recusar', 'recuso']):
        numeros = re.findall(r'\d+', msg)
        if numeros and numeros[0] == '5' and len(msg) <= 2:
            numeros = []
        
        desafios_pendentes = get_desafios_pendentes(jogador['id'])
        
        if not desafios_pendentes:
            return """❌ *Rejeitar Desafio*

Você não tem nenhum desafio pendente para rejeitar.

_Digite *0* para voltar ao menu._"""
        
        if len(desafios_pendentes) == 1 and not numeros:
            desafio = desafios_pendentes[0]
            sucesso, mensagem_retorno = rejeitar_desafio(desafio['id'], jogador['id'])
            
            if sucesso:
                return f"""⚠️ *Desafio Rejeitado*

Você rejeitou o desafio de *{desafio['challenger_name']}*.

WO aplicado - você perdeu a posição.

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        if numeros:
            challenge_id = int(numeros[0])
            sucesso, mensagem_retorno = rejeitar_desafio(challenge_id, jogador['id'])
            
            if sucesso:
                return f"""⚠️ *Desafio #{challenge_id} rejeitado.* WO aplicado.

_Digite *0* para voltar ao menu._"""
            else:
                return f"❌ {mensagem_retorno}"
        
        lista = "\n".join([f"   #{d['id']} - {d['challenger_name']} (posição {d['challenger_position']}º)" for d in desafios_pendentes])
        return f"""❌ *Rejeitar Desafio*

⚠️ *ATENÇÃO*: Rejeitar resulta em WO!

Seus desafios pendentes:
{lista}

Para rejeitar, digite: *5 [número]*
Exemplo: *5 123*

💡 Prefere propor novas datas? Digite *7*

_Digite *0* para voltar ao menu._"""
    
    # COMANDO [6]: Criar desafio
    if msg == '6' or any(palavra in msg for palavra in ['criar desafio', 'desafiar', 'novo desafio', 'quero desafiar']):
        if tem_desafio_ativo(jogador['id']):
            return """⚠️ *Você já tem um desafio ativo!*

Conclua seu desafio atual antes de criar um novo.

Digite *3* para ver seus desafios.

_Digite *0* para voltar ao menu._"""
        
        possiveis = get_possiveis_desafiados(jogador['id'])
        
        if not possiveis:
            return f"""🎯 *Criar Desafio*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

❌ No momento não há jogadores disponíveis para desafio.

_Digite *0* para voltar ao menu._"""
        
        linhas = []
        for i, p in enumerate(possiveis, 1):
            linhas.append(f"   *{i}* - {p['name']} ({p['position']}º)")
        
        lista = "\n".join(linhas)
        
        set_chat_state(telefone_normalizado, 'selecionando_oponente', {
            'possiveis': possiveis
        })
        
        return f"""🎯 *Criar Desafio*

Olá, {jogador['name']}!
Você está na posição *{jogador['position']}º*.

Selecione quem você quer desafiar:
{lista}

Digite o *número* do oponente (ex: *1*)

_Digite *0* para cancelar._"""
    
    # COMANDO [7]: Propor novas datas
    if msg == '7' or any(palavra in msg for palavra in ['propor data', 'nova data', 'outras datas', 'propor datas']):
        desafios_pendentes = get_desafios_pendentes(jogador['id'])
        
        if not desafios_pendentes:
            return """📅 *Propor Novas Datas*

Você não tem nenhum desafio pendente para propor novas datas.

_Digite *0* para voltar ao menu._"""
        
        # Se tem apenas um desafio pendente, já inicia o fluxo
        if len(desafios_pendentes) == 1:
            desafio = desafios_pendentes[0]
            
            set_chat_state(telefone_normalizado, 'propondo_data_1', {
                'challenge_id': desafio['id'],
                'challenger_name': desafio['challenger_name']
            })
            
            hoje = datetime.now()
            data_max = hoje + timedelta(days=7)
            
            return f"""📅 *Propor Novas Datas*

Você vai propor novas datas para o desafio de *{desafio['challenger_name']}*.

Informe a *primeira opção* de data:
(formato DD/MM, máximo 7 dias - até {data_max.strftime('%d/%m/%Y')})

_Digite *0* para cancelar._"""
        
        # Se tem múltiplos desafios
        lista = "\n".join([f"   #{d['id']} - {d['challenger_name']} ({d['challenger_position']}º)" for d in desafios_pendentes])
        return f"""📅 *Propor Novas Datas*

Você tem {len(desafios_pendentes)} desafios pendentes:
{lista}

Para propor datas, digite: *7 [número]*
Exemplo: *7 123*

_Digite *0* para voltar ao menu._"""
    
    # MENU PRINCIPAL [0]
    # Verificar se tem proposta de data pendente
    proposta = get_proposta_pendente_para_desafiante(jogador['id'])
    aviso_proposta = ""
    if proposta:
        aviso_proposta = """
━━━━━━━━━━━━━━━━━━━━━
📅 *VOCÊ TEM PROPOSTA DE DATAS!*
Digite *3* para ver e responder
━━━━━━━━━━━━━━━━━━━━━
"""
    
    return f"""🏌️ *Liga Olímpica de Golfe*

Olá, *{jogador['name']}*!
📊 Posição atual: *{jogador['position']}º*
{aviso_proposta}
━━━━━━━━━━━━━━━━━━━━━
*MENU DE OPÇÕES*
━━━━━━━━━━━━━━━━━━━━━

*[1]* 📊 Minha posição
*[2]* 🎯 Quem posso desafiar
*[3]* 📋 Meus desafios
*[4]* ✅ Aceitar desafio
*[5]* ❌ Rejeitar desafio
*[6]* ⚔️ Criar desafio
*[7]* 📅 Propor novas datas

━━━━━━━━━━━━━━━━━━━━━

_Digite o número da opção desejada._"""


def parse_data_input(texto):
    """Converte texto de data para objeto date"""
    texto = texto.strip()
    hoje = datetime.now().date()
    ano_atual = hoje.year
    
    # Formatos aceitos: DD/MM, DD-MM, DD.MM, DD/MM/YYYY
    formatos = [
        (r'^(\d{1,2})[/\-.](\d{1,2})$', 'short'),
        (r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{4})$', 'full'),
        (r'^(\d{1,2})[/\-.](\d{1,2})[/\-.](\d{2})$', 'short_year'),
    ]
    
    for pattern, fmt_type in formatos:
        match = re.match(pattern, texto)
        if match:
            try:
                if fmt_type == 'short':
                    dia, mes = match.groups()
                    data = datetime(ano_atual, int(mes), int(dia)).date()
                    # Se a data já passou, assume ano que vem
                    if data < hoje:
                        data = datetime(ano_atual + 1, int(mes), int(dia)).date()
                    return data
                elif fmt_type == 'full':
                    dia, mes, ano = match.groups()
                    return datetime(int(ano), int(mes), int(dia)).date()
                elif fmt_type == 'short_year':
                    dia, mes, ano = match.groups()
                    ano_completo = 2000 + int(ano)
                    return datetime(ano_completo, int(mes), int(dia)).date()
            except ValueError:
                continue
    
    return None


if __name__ == '__main__':
    # Verificar se o banco de dados existe, caso contrário, importar dados
    if not os.path.exists(DATABASE):
        print("Banco de dados não encontrado. Executando script de importação...")
        import import_data
        import_data.create_database()
        import_data.import_players_data(import_data.cursor)
    
    # Verificar se as colunas active e notes existem na tabela players
    conn = get_db_connection()
    cursor = conn.cursor()
    
    columns_info = cursor.execute('PRAGMA table_info(players)').fetchall()
    column_names = [col[1] for col in columns_info]
    
    if 'active' not in column_names:
        print("Adicionando coluna 'active' à tabela players...")
        cursor.execute('ALTER TABLE players ADD COLUMN active INTEGER DEFAULT 1')
    
    if 'notes' not in column_names:
        print("Adicionando coluna 'notes' à tabela players...")
        cursor.execute('ALTER TABLE players ADD COLUMN notes TEXT')
    
    if 'hcp_last_update' not in column_names:
        print("Adicionando coluna 'hcp_last_update' à tabela players...")
        cursor.execute('ALTER TABLE players ADD COLUMN hcp_last_update DATETIME')
    
    if 'profile_photo' not in column_names:
        print("Adicionando coluna 'profile_photo' à tabela players...")
        cursor.execute('ALTER TABLE players ADD COLUMN profile_photo TEXT DEFAULT NULL')
    
    conn.commit()
    conn.close()

    add_result_type_column()

    create_verification_tokens_table()

    # Criar a tabela de histórico de HCP
    create_hcp_history_table()
    
    # Criar a tabela de histórico diário se não existir
    create_daily_history_table()
    
    # Adicionar coluna de prazo de resposta à tabela de desafios
    add_response_deadline_column()

    # Adicionar coluna de país se não existir
    add_country_column()

    # Criar tabela de configurações do sistema
    create_system_settings_table()

    # Criar configuração de submissão de resultados por jogadores
    create_player_result_setting()
    
    # Garantir que a pasta para fotos existe
    os.makedirs('static/profile_photos', exist_ok=True)
    
    # Verificar e corrigir a estrutura da pirâmide
    print("Realizando verificação inicial da pirâmide...")
    conn = get_db_connection()
    fix_position_gaps(conn)
    update_all_tiers(conn)
    conn.commit()
    conn.close()
    
    # Sincronizar o histórico diário com o histórico de ranking
    print("Sincronizando histórico para o dia atual...")
    sync_ranking_history_tables()
    
    # Criar pasta de templates se não existir
    if not os.path.exists('templates'):
        os.makedirs('templates')
    
    # Criar pasta static se não existir
    if not os.path.exists('static'):
        os.makedirs('static')
    
    create_business_table()

    # ✨ NOVA ADIÇÃO: Auto-correção inicial do ranking feminino
    print("🔧 Executando auto-correção inicial do ranking feminino...")
    auto_fix_female_ranking()
    
    # Modificação: adicionado argumento host='0.0.0.0' para permitir acesso externo
    app.run(debug=True, host='0.0.0.0')

